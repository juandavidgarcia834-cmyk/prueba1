"""Exportación contextual a Excel del módulo Historial.

Diseño:
  • Estructura estrictamente tabular (una columna por dato, sin celdas combinadas).
  • Parámetros numéricos exportados como float real (no texto).
  • Crioscopía con formato 0.000 manteniendo signo negativo, reconocible numéricamente.
  • Freeze panes en fila 1 de todas las hojas.
  • Solo se exportan los datos filtrados que llegan en df_filtrado.

Hojas generadas (contextuales):
  filtro_tipo == "RUTAS"        → General_Rutas + Detalle_Estaciones
  filtro_tipo == "TRANSUIZA"    → Transuiza
  filtro_tipo == "SEGUIMIENTOS" según filtro_subtipo:
        ESTACIONES              → Seg_Estaciones
        ACOMPAÑAMIENTOS         → Acomp_General + Acomp_Detalles
        CONTRAMUESTRAS          → Contramuestras
        TODOS                   → todas las hojas anteriores que apliquen
  filtro_tipo == "TODOS"        → todas las hojas que correspondan a la data
"""
from __future__ import annotations

import io
import json
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.data_utils import load_catalogo, load_seguimientos


# ─────────────────────────────────────────────────────────────────────
# Helpers numéricos (devuelven float/int reales, no texto)
# ─────────────────────────────────────────────────────────────────────
def _f(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        if isinstance(v, float) and pd.isna(v):
            return None
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _i(v: Any) -> int | None:
    f = _f(v)
    return int(f) if f is not None else None


def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def _parse_json_list(raw: Any) -> list[dict]:
    if not raw:
        return []
    if isinstance(raw, list):
        return raw
    try:
        data = json.loads(str(raw))
        return data if isinstance(data, list) else []
    except (ValueError, TypeError):
        return []


# ─────────────────────────────────────────────────────────────────────
# Estilos
# ─────────────────────────────────────────────────────────────────────
_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)

# Formatos numéricos por nombre de columna (sufijo permite distinguir crioscopía)
_FMT_INT = "#,##0"
_FMT_2 = "0.00"
_FMT_3 = "0.000"

_NUM_FORMATS: dict[str, str] = {
    # ── Volumen e intero ────────────────────────
    "Volumen": _FMT_INT,
    "Volumen_Declarado": _FMT_INT,
    "Volumen_Estacion": _FMT_INT,
    "Volumen_Suma_Muestras": _FMT_INT,
    "Diferencia_Volumen": _FMT_INT,
    # ── Porcentajes 2 decimales ─────────────────
    "ST": _FMT_2,
    "ST_Pond": _FMT_2,
    "Diferencia_ST": _FMT_2,
    "Solidos_Totales": _FMT_2,
    "Grasa": _FMT_2,
    "Proteina": _FMT_2,
    "%Agua": _FMT_2,
    "ST_Carrotanque": _FMT_2,
    "ST_Contramuestra": _FMT_2,
    "Diferencia_ST_Carr_Contra": _FMT_2,
    # ── Crioscopía 3 decimales (negativos válidos) ──
    "IC": _FMT_3,
    "IC_Pond": _FMT_3,
    "Diferencia_IC": _FMT_3,
    "Crioscopia": _FMT_3,
}


def _apply_sheet_styling(ws, df: pd.DataFrame) -> None:
    """Estiliza encabezado, freeze panes, formatos numéricos y anchos."""
    if df.empty:
        # Mantener encabezados aún sin datos
        for col_idx, col_name in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=col_idx, value=str(col_name))
            c.fill = _HDR_FILL
            c.font = _HDR_FONT
            c.alignment = _HDR_ALIGN
            c.border = _BORDER
        ws.freeze_panes = "A2"
        return

    n_rows = len(df) + 1  # +1 por encabezado
    n_cols = len(df.columns)

    # Encabezado
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _HDR_ALIGN
        c.border = _BORDER

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # Formato numérico por columna + bordes ligeros
    for col_idx, col_name in enumerate(df.columns, start=1):
        fmt = _NUM_FORMATS.get(str(col_name))
        col_letter = get_column_letter(col_idx)

        # Auto ancho en función del header y de muestras de datos
        try:
            sample_max = df[col_name].astype(str).head(50).map(len).max()
        except Exception:
            sample_max = 0
        width = max(12, min(38, max(len(str(col_name)) + 2, int(sample_max) + 2)))
        ws.column_dimensions[col_letter].width = width

        if fmt:
            for r in range(2, n_rows + 1):
                ws.cell(row=r, column=col_idx).number_format = fmt

    # Bordes y alineación generales (no tocan tipo de dato)
    align_center = Alignment(horizontal="center", vertical="center")
    for r in range(2, n_rows + 1):
        for c_idx in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = _BORDER
            cell.alignment = align_center


# ─────────────────────────────────────────────────────────────────────
# Constructores de DataFrames por tipo
# ─────────────────────────────────────────────────────────────────────
def _df_general_rutas(df_rutas: pd.DataFrame) -> pd.DataFrame:
    """Hoja General_Rutas — una fila por ruta."""
    rows = []
    for _, r in df_rutas.iterrows():
        st_v = _f(r.get("solidos_ruta"))
        st_p = _f(r.get("st_pond"))
        ic_v = _f(r.get("crioscopia_ruta"))
        ic_p = _f(r.get("ic_pond"))
        rows.append(
            {
                "Fecha":         _s(r.get("fecha")),
                "Ruta":          _s(r.get("ruta")),
                "Placa":         _s(r.get("placa")),
                "Conductor":     _s(r.get("conductor")),
                "Volumen":       _i(r.get("volumen_declarado")),
                "ST":            st_v,
                "ST_Pond":       st_p,
                "Diferencia_ST": (round(st_v - st_p, 3) if st_v is not None and st_p is not None else None),
                "IC":            ic_v,
                "IC_Pond":       ic_p,
                "Diferencia_IC": (round(ic_v - ic_p, 4) if ic_v is not None and ic_p is not None else None),
            }
        )
    return pd.DataFrame(rows, columns=[
        "Fecha", "Ruta", "Placa", "Conductor", "Volumen",
        "ST", "ST_Pond", "Diferencia_ST",
        "IC", "IC_Pond", "Diferencia_IC",
    ])


def _df_detalle_estaciones(df_rutas: pd.DataFrame, cat_map: dict[str, str]) -> pd.DataFrame:
    """Hoja Detalle_Estaciones — aplana cada estación repitiendo cabecera."""
    rows = []
    for _, r in df_rutas.iterrows():
        ests = _parse_json_list(r.get("estaciones_json"))
        if not ests:
            continue
        fecha   = _s(r.get("fecha"))
        ruta    = _s(r.get("ruta"))
        placa   = _s(r.get("placa"))
        conduct = _s(r.get("conductor"))
        vol_dec = _i(r.get("volumen_declarado"))
        for e in ests:
            cod = _s(e.get("codigo"))
            rows.append({
                "Fecha":             fecha,
                "Ruta":              ruta,
                "Placa":             placa,
                "Conductor":         conduct,
                "Volumen_Declarado": vol_dec,
                "Codigo_Estacion":   cod,
                "Nombre_Estacion":   cat_map.get(cod, ""),
                "Grasa":             _f(e.get("grasa")),
                "Solidos_Totales":   _f(e.get("solidos")),
                "Proteina":          _f(e.get("proteina")),
                "Crioscopia":        _f(e.get("crioscopia")),
                "%Agua":             _f(e.get("agua_pct")),
                "Volumen_Estacion":  _i(e.get("volumen")),
                "Alcohol":           _s(e.get("alcohol")),
                "Cloruros":          _s(e.get("cloruros")),
                "Neutralizantes":    _s(e.get("neutralizantes")),
                "Observaciones":     _s(e.get("obs")),
            })
    return pd.DataFrame(rows, columns=[
        "Fecha", "Ruta", "Placa", "Conductor", "Volumen_Declarado",
        "Codigo_Estacion", "Nombre_Estacion",
        "Grasa", "Solidos_Totales", "Proteina", "Crioscopia", "%Agua",
        "Volumen_Estacion", "Alcohol", "Cloruros", "Neutralizantes", "Observaciones",
    ])


def _df_transuiza(df_t: pd.DataFrame) -> pd.DataFrame:
    """Hoja Transuiza."""
    rows = []
    for _, r in df_t.iterrows():
        st_carr = _f(r.get("st_carrotanque"))
        st_cm   = _f(r.get("solidos_ruta"))  # ST de la contramuestra/muestra
        diff    = _f(r.get("diferencia_solidos"))
        if diff is None and st_carr is not None and st_cm is not None:
            diff = round(st_carr - st_cm, 3)
        rows.append({
            "Fecha":                     _s(r.get("fecha")),
            "Placa":                     _s(r.get("placa")),
            "ST_Carrotanque":            st_carr,
            "Grasa":                     _f(r.get("grasa_muestra")),
            "ST_Contramuestra":          st_cm,
            "Proteina":                  _f(r.get("proteina_muestra")),
            "Diferencia_ST_Carr_Contra": diff,
        })
    return pd.DataFrame(rows, columns=[
        "Fecha", "Placa", "ST_Carrotanque", "Grasa",
        "ST_Contramuestra", "Proteina", "Diferencia_ST_Carr_Contra",
    ])


def _df_seg_estaciones(df_seg: pd.DataFrame) -> pd.DataFrame:
    """Hoja Seg_Estaciones — datos planos del subtipo ESTACIONES."""
    rows = []
    for _, r in df_seg.iterrows():
        rows.append({
            "Codigo_Estacion": _s(r.get("seg_codigo")),
            "Fecha":           _s(r.get("fecha")),
            "Ruta":            _s(r.get("ruta")),
            "Entregado_Por":   _s(r.get("seg_quien_trajo")),
            "Responsable":     _s(r.get("seg_responsable")),
            "ID_Muestra":      _s(r.get("seg_id_muestra")),
            "Grasa":           _f(r.get("seg_grasa")),
            "ST":              _f(r.get("seg_st")),
            "Proteina":        _f(r.get("seg_proteina")),
            "IC":              _f(r.get("seg_ic")),
            "%Agua":           _f(r.get("seg_agua")),
            "Alcohol":         _s(r.get("seg_alcohol")),
            "Cloruros":        _s(r.get("seg_cloruros")),
            "Neutralizantes":  _s(r.get("seg_neutralizantes")),
            "Observaciones":   _s(r.get("seg_observaciones")),
            "Guardado_En":     _s(r.get("guardado_en")),
        })
    return pd.DataFrame(rows, columns=[
        "Codigo_Estacion", "Fecha", "Ruta", "Entregado_Por", "Responsable",
        "ID_Muestra", "Grasa", "ST", "Proteina", "IC", "%Agua",
        "Alcohol", "Cloruros", "Neutralizantes", "Observaciones", "Guardado_En",
    ])


def _df_acomp_general(df_acomp: pd.DataFrame) -> pd.DataFrame:
    """Hoja Acomp_General — datos cabecera del acompañamiento.

    Acompañamientos NO tienen Conductor (no hay vehículo asociado), pero sí
    registran 'Entregado por' (quién trajo las muestras al laboratorio) y el
    'Responsable' del análisis. Se incluyen ambos.
    """
    rows = []
    for _, r in df_acomp.iterrows():
        st_v = _f(r.get("seg_solidos_ruta"))
        st_p = _f(r.get("seg_st_pond"))
        ic_v = _f(r.get("seg_crioscopia_ruta"))
        ic_p = _f(r.get("seg_ic_pond"))
        rows.append({
            "Fecha":         _s(r.get("fecha")),
            "Ruta":          _s(r.get("ruta")),
            "Entregado_Por": _s(r.get("seg_quien_trajo")),
            "Responsable":   _s(r.get("seg_responsable")),
            "Volumen":       _i(r.get("seg_vol_declarado")),
            "ST":            st_v,
            "ST_Pond":       st_p,
            "Diferencia_ST": (round(st_v - st_p, 3) if st_v is not None and st_p is not None else None),
            "IC":            ic_v,
            "IC_Pond":       ic_p,
            "Diferencia_IC": (round(ic_v - ic_p, 4) if ic_v is not None and ic_p is not None else None),
            "Guardado_En":   _s(r.get("guardado_en")),
        })
    return pd.DataFrame(rows, columns=[
        "Fecha", "Ruta", "Entregado_Por", "Responsable", "Volumen",
        "ST", "ST_Pond", "Diferencia_ST",
        "IC", "IC_Pond", "Diferencia_IC",
        "Guardado_En",
    ])


def _df_acomp_detalles(df_acomp: pd.DataFrame, cat_map: dict[str, str]) -> pd.DataFrame:
    """Hoja Acomp_Detalles — aplana muestras_json.

    Sin columna Conductor: en acompañamientos quien trae las muestras es
    'Entregado_Por' (seg_quien_trajo). Se repiten cabecera por cada muestra.
    """
    rows = []
    for _, r in df_acomp.iterrows():
        muestras = _parse_json_list(r.get("muestras_json"))
        if not muestras:
            continue
        fecha   = _s(r.get("fecha"))
        ruta    = _s(r.get("ruta"))
        quien   = _s(r.get("seg_quien_trajo"))
        respo   = _s(r.get("seg_responsable"))
        vol_dec = _i(r.get("seg_vol_declarado"))
        for m in muestras:
            cod = _s(m.get("ID") or m.get("_id_muestra"))
            rows.append({
                "Fecha":             fecha,
                "Ruta":              ruta,
                "Entregado_Por":     quien,
                "Responsable":       respo,
                "Volumen_Declarado": vol_dec,
                "Codigo_Estacion":   cod,
                "Nombre_Estacion":   cat_map.get(cod, ""),
                "Grasa":             _f(m.get("_grasa")),
                "Solidos_Totales":   _f(m.get("_st")),
                "Proteina":          _f(m.get("_proteina")),
                "Crioscopia":        _f(m.get("_ic")),
                "%Agua":             _f(m.get("_agua")),
                "Volumen_Estacion":  _i(m.get("_volumen")),
                "Alcohol":           _s(m.get("_alcohol")),
                "Cloruros":          _s(m.get("_cloruros")),
                "Neutralizantes":    _s(m.get("_neutralizantes")),
                "Observaciones":     _s(m.get("_obs")),
            })
    return pd.DataFrame(rows, columns=[
        "Fecha", "Ruta", "Entregado_Por", "Responsable", "Volumen_Declarado",
        "Codigo_Estacion", "Nombre_Estacion",
        "Grasa", "Solidos_Totales", "Proteina", "Crioscopia", "%Agua",
        "Volumen_Estacion", "Alcohol", "Cloruros", "Neutralizantes", "Observaciones",
    ])


def _df_contramuestras(df_cm: pd.DataFrame, cat_map: dict[str, str]) -> pd.DataFrame:
    """Hoja Contramuestras — una fila por muestra con datos guardados.

    Estructura solicitada: Código de muestra · Nombre estación · Proveedor
    (quien entregó) · parámetros de calidad (Grasa, ST, Proteína, IC, %Agua,
    Alcohol, Cloruros, Neutralizantes) · Observaciones.
    Se conserva además Fecha, Ruta y Guardado_En para trazabilidad.
    """
    rows = []
    for _, r in df_cm.iterrows():
        muestras = _parse_json_list(r.get("muestras_json"))
        fecha     = _s(r.get("fecha"))
        ruta      = _s(r.get("ruta"))
        proveedor = _s(r.get("seg_quien_trajo"))   # quién entregó las muestras
        respo     = _s(r.get("seg_responsable"))
        guardado  = _s(r.get("guardado_en"))
        if not muestras:
            rows.append({
                "Fecha":           fecha,
                "Codigo":          "",
                "Nombre_Estacion": "",
                "Proveedor":       proveedor,
                "Ruta":            ruta,
                "Responsable":     respo,
                "Grasa":           None,
                "ST":              None,
                "Proteina":        None,
                "IC":              None,
                "%Agua":           None,
                "Alcohol":         "",
                "Cloruros":        "",
                "Neutralizantes":  "",
                "Observaciones":   "",
                "Guardado_En":     guardado,
            })
            continue
        for m in muestras:
            cod = _s(m.get("ID") or m.get("_id_muestra"))
            rows.append({
                "Fecha":           fecha,
                "Codigo":          cod,
                "Nombre_Estacion": cat_map.get(cod, ""),
                "Proveedor":       proveedor,
                "Ruta":            ruta,
                "Responsable":     respo,
                "Grasa":           _f(m.get("_grasa")),
                "ST":              _f(m.get("_st")),
                "Proteina":        _f(m.get("_proteina")),
                "IC":              _f(m.get("_ic")),
                "%Agua":           _f(m.get("_agua")),
                "Alcohol":         _s(m.get("_alcohol")),
                "Cloruros":        _s(m.get("_cloruros")),
                "Neutralizantes":  _s(m.get("_neutralizantes")),
                "Observaciones":   _s(m.get("_obs")),
                "Guardado_En":     guardado,
            })
    return pd.DataFrame(rows, columns=[
        "Fecha", "Codigo", "Nombre_Estacion", "Proveedor", "Ruta", "Responsable",
        "Grasa", "ST", "Proteina", "IC", "%Agua",
        "Alcohol", "Cloruros", "Neutralizantes", "Observaciones", "Guardado_En",
    ])


# ─────────────────────────────────────────────────────────────────────
# Función pública
# ─────────────────────────────────────────────────────────────────────
def historial_to_excel_filtrado(
    df_filtrado: pd.DataFrame,
    fecha_desde,
    fecha_hasta,
    filtro_tipo: str,
    filtro_subtipo: str = "TODOS",
) -> bytes:
    """Genera un Excel contextual según el tipo de registro consultado.

    Args:
        df_filtrado: DataFrame ya filtrado por la UI (lo que el usuario ve).
        fecha_desde/fecha_hasta: rango de fechas activo (solo se usa si
            filtro_tipo='SEGUIMIENTOS' y df_filtrado no contiene seguimientos).
        filtro_tipo: 'RUTAS' | 'TRANSUIZA' | 'SEGUIMIENTOS' | 'TODOS'.
        filtro_subtipo: dentro de SEGUIMIENTOS — 'ESTACIONES' |
            'ACOMPAÑAMIENTOS' | 'CONTRAMUESTRAS' | 'TODOS'.

    Returns:
        Bytes del .xlsx listo para st.download_button.
    """
    if df_filtrado is None:
        df_filtrado = pd.DataFrame()

    filtro_tipo = (filtro_tipo or "TODOS").upper()
    filtro_subtipo = (filtro_subtipo or "TODOS").upper()

    # Catálogo (código → nombre estación) — falla silenciosa si no carga
    try:
        _cat = load_catalogo()
        cat_map = dict(zip(_cat["codigo"], _cat["nombre"]))
    except Exception:
        cat_map = {}

    # Subconjuntos por tipo
    if "tipo_seguimiento" in df_filtrado.columns:
        df_rutas  = df_filtrado[df_filtrado["tipo_seguimiento"] == "RUTAS"].copy()
        df_trans  = df_filtrado[df_filtrado["tipo_seguimiento"] == "TRANSUIZA"].copy()
        df_seg_in = df_filtrado[df_filtrado["tipo_seguimiento"] == "SEGUIMIENTOS"].copy()
    else:
        df_rutas = df_filtrado.copy()
        df_trans = df_filtrado.copy()
        df_seg_in = df_filtrado.copy()

    # Para SEGUIMIENTOS, df_filtrado puede venir de la vista de seguimientos
    # (no incluye tipo_seguimiento). En ese caso usamos el df tal cual.
    if filtro_tipo == "SEGUIMIENTOS" and df_seg_in.empty and not df_filtrado.empty:
        df_seg_in = df_filtrado.copy()

    # Cargar seguimientos completos solo si hace falta
    def _load_seg_filtrado() -> pd.DataFrame:
        try:
            df = load_seguimientos()
        except Exception:
            return pd.DataFrame()
        if "_fecha_dt" in df.columns and fecha_desde and fecha_hasta:
            df = df[
                (df["_fecha_dt"].dt.date >= fecha_desde)
                & (df["_fecha_dt"].dt.date <= fecha_hasta)
            ]
        return df.drop(columns=["_fecha_dt", "_estado"], errors="ignore")

    if filtro_tipo == "TODOS" and df_seg_in.empty:
        df_seg_in = _load_seg_filtrado()

    # ── Construcción de hojas ─────────────────────────────────────────
    sheets: dict[str, pd.DataFrame] = {}

    if filtro_tipo in ("RUTAS", "TODOS") and not df_rutas.empty:
        sheets["General_Rutas"]      = _df_general_rutas(df_rutas)
        det_est = _df_detalle_estaciones(df_rutas, cat_map)
        if not det_est.empty:
            sheets["Detalle_Estaciones"] = det_est

    if filtro_tipo in ("TRANSUIZA", "TODOS") and not df_trans.empty:
        sheets["Transuiza"] = _df_transuiza(df_trans)

    if filtro_tipo in ("SEGUIMIENTOS", "TODOS") and not df_seg_in.empty:
        sub = "sub_tipo_seguimiento"
        if sub in df_seg_in.columns:
            df_est   = df_seg_in[df_seg_in[sub] == "ESTACIONES"].copy()
            df_acomp = df_seg_in[df_seg_in[sub] == "ACOMPAÑAMIENTOS"].copy()
            df_cm    = df_seg_in[df_seg_in[sub] == "CONTRAMUESTRAS"].copy()
        else:
            df_est = df_acomp = df_cm = df_seg_in.copy()

        # Filtro_subtipo restringe SOLO si filtro_tipo=SEGUIMIENTOS
        if filtro_tipo == "SEGUIMIENTOS" and filtro_subtipo != "TODOS":
            if filtro_subtipo == "ESTACIONES":
                df_acomp = df_acomp.iloc[0:0]; df_cm = df_cm.iloc[0:0]
            elif filtro_subtipo == "ACOMPAÑAMIENTOS":
                df_est = df_est.iloc[0:0]; df_cm = df_cm.iloc[0:0]
            elif filtro_subtipo == "CONTRAMUESTRAS":
                df_est = df_est.iloc[0:0]; df_acomp = df_acomp.iloc[0:0]

        if not df_est.empty:
            sheets["Seg_Estaciones"] = _df_seg_estaciones(df_est)
        if not df_acomp.empty:
            sheets["Acomp_General"]  = _df_acomp_general(df_acomp)
            ac_det = _df_acomp_detalles(df_acomp, cat_map)
            if not ac_det.empty:
                sheets["Acomp_Detalles"] = ac_det
        if not df_cm.empty:
            sheets["Contramuestras"] = _df_contramuestras(df_cm, cat_map)

    if not sheets:
        sheets["Sin_datos"] = pd.DataFrame({"Mensaje": ["No hay registros para exportar con los filtros actuales."]})

    # ── Escritura del archivo ─────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]
            _apply_sheet_styling(ws, df)

    buf.seek(0)
    return buf.read()
