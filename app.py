import streamlit as st
from supabase import create_client, Client
from datetime import datetime, date, timezone, timedelta
import base64
import csv
import io
import os
import re
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import altair as alt
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
import streamlit.components.v1 as components

# ── Zona horaria Colombia (UTC-5, sin horario de verano) ─────────────────────
COL_TZ = timezone(timedelta(hours=-5))

def now_col() -> datetime:
    """Hora actual en zona horaria Colombia."""
    return datetime.now(tz=COL_TZ)

# --- CONFIGURACIÓN DE CONEXIÓN A SUPABASE ---
@st.cache_resource
def init_connection():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = init_connection()
# ── Persistencia CSV ─────────────────────────────────────────────────────────
CSV_PATH  = "rutas_historial.csv"
FOTOS_DIR = "fotos"
os.makedirs(FOTOS_DIR, exist_ok=True)
CSV_COLS = [
    "tipo_seguimiento",
    "fecha", "ruta", "placa", "conductor",
    "volumen_declarado", "vol_estaciones", "diferencia",
    "solidos_ruta", "crioscopia_ruta", "st_pond", "ic_pond",
    "num_estaciones", "guardado_en",
    "st_carrotanque", "grasa_muestra", "proteina_muestra", "diferencia_solidos",
    "estaciones_json", "fotos_json",
]

# ── CSV separado para SEGUIMIENTOS ───────────────────────────────────────────
SEG_CSV_PATH = "seguimientos_historial.csv"
SEG_COLS = [
    "sub_tipo_seguimiento", "fecha",
    "seg_codigo", "seg_quien_trajo", "ruta", "seg_responsable",
    "seg_id_muestra", "seg_volumen", "seg_grasa", "seg_st", "seg_ic", "seg_agua",
    "seg_alcohol", "seg_cloruros", "seg_neutralizantes", "seg_observaciones",
    "seg_vol_declarado", "seg_vol_muestras", "seg_diferencia_vol",
    "seg_solidos_ruta", "seg_crioscopia_ruta", "seg_st_pond", "seg_ic_pond",
    "muestras_json", "guardado_en", "fotos_json",
]

DRAFT_PATH      = "borrador_autoguardado.json"
CATALOGO_PATH   = "estaciones_catalogo.csv"
DRAFT_EXACT_KEYS = [
    "continuar", "_tipo_servicio_guardado", "_sub_tipo_seg_guardado",
    "tipo_servicio_select", "sub_tipo_seg_select",
    "_ruta_fg",
    "imagenes_confirmadas", "imagenes_nombres_guardados",
    "trans_imagenes_confirmadas", "trans_imagenes_nombres_guardados",
    "estaciones_guardadas", "form_ver",
    "trans_fecha", "trans_placa", "trans_st_carrotanque",
    "trans_grasa", "trans_st_muestra", "trans_proteina",
    "seg_fecha", "seg_codigo", "seg_quien_trajo", "seg_ruta_acomp",
    "seg_responsable", "seg_quality_key_counter",
    "acomp_muestras", "contra_muestras",
]
DRAFT_PREFIXES = (
    "nue_",
    "fecha_ruta_", "nombre_ruta_", "placa_vehiculo_", "conductor_",
    "volumen_ruta_", "solidos_totales_", "crioscopia_",
    "seg_id_muestra_", "seg_grasa_", "seg_st_", "seg_ic_raw_", "seg_agua_",
    "seg_alcohol_", "seg_cloruros_", "seg_neutralizantes_", "seg_observaciones_",
)


def _draft_encode(value):
    if isinstance(value, datetime):
        return {"__draft_type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"__draft_type": "date", "value": value.isoformat()}
    try:
        json.dumps(value)
        return value
    except TypeError:
        return str(value)


def _draft_decode(value):
    if isinstance(value, dict) and value.get("__draft_type") in ("date", "datetime"):
        raw = value.get("value", "")
        try:
            return datetime.fromisoformat(raw).date()
        except Exception:
            return date.today()
    return value


def restore_draft_state():
    if st.session_state.get("_draft_restored"):
        return
    st.session_state["_draft_restored"] = True
    if not os.path.exists(DRAFT_PATH):
        return
    try:
        with open(DRAFT_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return
    if "tipo_servicio_select" not in data and "_tipo_servicio_guardado" in data:
        data["tipo_servicio_select"] = data.get("_tipo_servicio_guardado")
    if "sub_tipo_seg_select" not in data and "_sub_tipo_seg_guardado" in data:
        data["sub_tipo_seg_select"] = data.get("_sub_tipo_seg_guardado")
    for key, value in data.items():
        if key not in st.session_state:
            st.session_state[key] = _draft_decode(value)


def save_draft_state():
    if st.session_state.pop("_skip_draft_save_once", False):
        return
    data = {}
    for key in DRAFT_EXACT_KEYS:
        if key in st.session_state:
            data[key] = _draft_encode(st.session_state[key])
    for key, value in st.session_state.items():
        if key.startswith(DRAFT_PREFIXES):
            data[key] = _draft_encode(value)
    try:
        with open(DRAFT_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def clear_draft_state():
    st.session_state["_skip_draft_save_once"] = True
    try:
        if os.path.exists(DRAFT_PATH):
            os.remove(DRAFT_PATH)
    except Exception:
        pass


def save_fotos_to_disk(uploaded_files: list, prefix: str) -> list[str]:
    """Guarda imágenes en FOTOS_DIR y retorna la lista de rutas relativas."""
    saved = []
    if not uploaded_files:
        return saved
    ts = now_col().strftime("%Y%m%d_%H%M%S")
    safe_prefix = re.sub(r"[^A-Z0-9_\-]", "_", prefix.upper())
    for i, uf in enumerate(uploaded_files, start=1):
        ext = uf.name.rsplit(".", 1)[-1].lower()
        ext = ext if ext in ("jpg", "jpeg", "png") else "jpg"
        fname = f"{safe_prefix}_{ts}_{i}.{ext}"
        fpath = os.path.join(FOTOS_DIR, fname)
        uf.seek(0)
        with open(fpath, "wb") as fh:
            fh.write(uf.read())
        uf.seek(0)
        saved.append(fpath)
    return saved


def load_historial() -> pd.DataFrame:
    if not os.path.exists(CSV_PATH):
        return pd.DataFrame(columns=CSV_COLS)
    try:
        df = pd.read_csv(CSV_PATH, dtype=str)
        # Columnas nuevas que pueden no existir en CSVs anteriores
        for col in CSV_COLS:
            if col not in df.columns:
                df[col] = ""
        # Tipo de seguimiento: filas vacías → RUTAS (compatibilidad)
        df["tipo_seguimiento"] = df["tipo_seguimiento"].fillna("RUTAS").replace("", "RUTAS")
        for col in ["volumen_declarado", "vol_estaciones", "diferencia", "num_estaciones"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        for col in ["solidos_ruta", "crioscopia_ruta", "st_pond", "ic_pond",
                    "st_carrotanque", "grasa_muestra", "proteina_muestra", "diferencia_solidos"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        if "fecha" in df.columns:
            df["_fecha_dt"] = pd.to_datetime(df["fecha"], format="%d/%m/%Y", errors="coerce")
        return df
    except Exception:
        return pd.DataFrame(columns=CSV_COLS)


def save_ruta_to_csv(row: dict):
    if os.path.exists(CSV_PATH):
        df = load_historial()
        if "_fecha_dt" in df.columns:
            df = df.drop(columns=["_fecha_dt"])
    else:
        df = pd.DataFrame(columns=CSV_COLS)
    # Asegurar que todas las columnas existen en df
    for col in CSV_COLS:
        if col not in df.columns:
            df[col] = ""
    new_row = pd.DataFrame([{k: row.get(k, "") for k in CSV_COLS}])
    df = pd.concat([df[CSV_COLS], new_row], ignore_index=True)
    df.to_csv(CSV_PATH, index=False, encoding="utf-8")


def load_seguimientos() -> pd.DataFrame:
    if not os.path.exists(SEG_CSV_PATH):
        return pd.DataFrame(columns=SEG_COLS)
    try:
        df = pd.read_csv(SEG_CSV_PATH, dtype=str)
        for col in SEG_COLS:
            if col not in df.columns:
                df[col] = ""
        for col in ["seg_grasa", "seg_st", "seg_ic", "seg_agua"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        # Unificación: registros viejos guardados como TERCEROS → ESTACIONES
        if "sub_tipo_seguimiento" in df.columns:
            df["sub_tipo_seguimiento"] = df["sub_tipo_seguimiento"].replace("TERCEROS", "ESTACIONES")
        if "fecha" in df.columns:
            df["_fecha_dt"] = pd.to_datetime(df["fecha"], format="%d/%m/%Y", errors="coerce")
        return df
    except Exception:
        return pd.DataFrame(columns=SEG_COLS)


@st.cache_data(ttl=300)
def load_catalogo() -> pd.DataFrame:
    """Carga el catálogo de estaciones. Retorna DataFrame vacío si no existe."""
    if not os.path.exists(CATALOGO_PATH):
        return pd.DataFrame(columns=["codigo", "nombre", "asesor"])
    try:
        df = pd.read_csv(CATALOGO_PATH, dtype=str)
        df["codigo"] = df["codigo"].str.strip()
        df["nombre"] = df["nombre"].str.strip().str.upper()
        if "asesor" not in df.columns:
            df["asesor"] = ""
        df["asesor"] = df["asesor"].fillna("").str.strip()
        return df.dropna(subset=["codigo", "nombre"])
    except Exception:
        return pd.DataFrame(columns=["codigo", "nombre", "asesor"])


def save_catalogo(df: pd.DataFrame):
    """Guarda el catálogo de estaciones asegurando el orden de columnas."""
    for col in ["codigo", "nombre", "asesor"]:
        if col not in df.columns:
            df[col] = ""
    df[["codigo", "nombre", "asesor"]].to_csv(CATALOGO_PATH, index=False)


def save_seguimiento_to_csv(row: dict):
    if os.path.exists(SEG_CSV_PATH):
        df = load_seguimientos()
        if "_fecha_dt" in df.columns:
            df = df.drop(columns=["_fecha_dt"])
    else:
        df = pd.DataFrame(columns=SEG_COLS)
    for col in SEG_COLS:
        if col not in df.columns:
            df[col] = ""
    new_row = pd.DataFrame([{k: row.get(k, "") for k in SEG_COLS}])
    df = pd.concat([df[SEG_COLS], new_row], ignore_index=True)
    df.to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")


def delete_seg_row(orig_idx: int):
    df = load_seguimientos()
    df = df.drop(index=orig_idx)
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[SEG_COLS].to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")


def delete_seg_rows(orig_indices: list):
    df = load_seguimientos()
    df = df.drop(index=[i for i in orig_indices if i in df.index])
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[SEG_COLS].to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")


def update_seg_row_in_csv(orig_idx: int, new_vals: dict):
    df = load_seguimientos()
    for k, v in new_vals.items():
        if k in df.columns:
            try:
                if pd.api.types.is_numeric_dtype(df[k].dtype):
                    df.at[orig_idx, k] = v
                else:
                    df.at[orig_idx, k] = str(v) if v is not None and v != "" else ""
            except Exception:
                df.at[orig_idx, k] = str(v) if v is not None else ""
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[SEG_COLS].to_csv(SEG_CSV_PATH, index=False, encoding="utf-8")




def delete_row_from_csv(orig_idx: int):
    df = load_historial()
    df = df.drop(index=orig_idx)
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[CSV_COLS].to_csv(CSV_PATH, index=False, encoding="utf-8")


def delete_rows_from_csv(orig_indices: list):
    df = load_historial()
    df = df.drop(index=[i for i in orig_indices if i in df.index])
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[CSV_COLS].to_csv(CSV_PATH, index=False, encoding="utf-8")


def update_row_in_csv(orig_idx: int, new_vals: dict):
    df = load_historial()
    for k, v in new_vals.items():
        if k in df.columns:
            df.at[orig_idx, k] = v
    if "_fecha_dt" in df.columns:
        df = df.drop(columns=["_fecha_dt"])
    df[CSV_COLS].to_csv(CSV_PATH, index=False, encoding="utf-8")


def calcular_estado_calidad(row: dict) -> str:
    """Retorna 'CONFORME' o 'DESVIACIÓN' según los parámetros de la ruta.
    Solo aplica a registros de tipo RUTAS."""
    if str(row.get("tipo_seguimiento", "RUTAS")).strip() != "RUTAS":
        return "CONFORME"
    try:
        st_val = float(str(row.get("solidos_ruta", "")).replace(",", "."))
        if 0 < st_val < 12.60:
            return "DESVIACIÓN"
    except (ValueError, TypeError):
        pass
    try:
        ic_val = float(str(row.get("crioscopia_ruta", "")).replace(",", "."))
        if ic_val > -0.535 or ic_val < -0.550:
            return "DESVIACIÓN"
    except (ValueError, TypeError):
        pass
    return "CONFORME"


def historial_to_excel_filtrado(
    df_filtrado: pd.DataFrame,
    fecha_desde,
    fecha_hasta,
    filtro_tipo: str,
    filtro_subtipo: str = "TODOS",
) -> bytes:
    """Excel multi-hoja respetando los filtros del Historial."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_hdr = PatternFill("solid", fgColor="1F4E79")
    fill_bad = PatternFill("solid", fgColor="FFC7CE")
    fill_alt = PatternFill("solid", fgColor="EEF4FB")
    font_hdr = Font(bold=True, size=10, color="FFFFFF")
    font_bad = Font(bold=True, size=10, color="9C0006")
    bold     = Font(bold=True, size=10)
    normal   = Font(size=10)
    center   = Alignment(horizontal="center", vertical="center")
    bd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _wh(ws, cols, widths):
        for ci, hdr in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=hdr)
            c.fill = fill_hdr; c.font = font_hdr
            c.alignment = center; c.border = bd
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 20

    def _wc(ws, ri, ci, val, fmt=None, bad=False, alt=False):
        v = val if (val is not None and not (isinstance(val, float) and pd.isna(val))) else ""
        c = ws.cell(row=ri, column=ci, value=v)
        c.alignment = center; c.border = bd
        if bad:
            c.font = font_bad; c.fill = fill_bad
        else:
            c.font = normal
            if alt: c.fill = fill_alt
        if fmt: c.number_format = fmt

    # ── Hoja RUTAS ────────────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "RUTAS"):
        df_r = (df_filtrado[df_filtrado["tipo_seguimiento"] == "RUTAS"].copy()
                if "tipo_seguimiento" in df_filtrado.columns else df_filtrado.copy())
        ws1 = wb.create_sheet("RUTAS")
        cols1 = [
            ("TIPO", "tipo_seguimiento"), ("FECHA", "fecha"), ("RUTA", "ruta"),
            ("PLACA", "placa"), ("CONDUCTOR", "conductor"),
            ("VOLUMEN (L)", "volumen_declarado"),
            ("VOL. ESTACIONES (L)", "vol_estaciones"),
            ("DIFERENCIA (L)", "diferencia"),
            ("SÓLIDOS RUTA (%)", "solidos_ruta"),
            ("ST POND", "st_pond"),
            ("CRIOSCOPIA RUTA (°C)", "crioscopia_ruta"),
            ("IC POND", "ic_pond"),
            ("Nº ESTACIONES", "num_estaciones"), ("GUARDADO EN", "guardado_en"),
        ]
        _wh(ws1, [h for h, _ in cols1],
            [10, 12, 18, 10, 18, 14, 18, 12, 16, 10, 18, 10, 12, 18])
        for ri, row in enumerate(df_r.itertuples(index=False), start=2):
            rd = row._asdict()
            desv_st = desv_ic = False
            try:
                v = float(str(rd.get("solidos_ruta","")).replace(",","."))
                if 0 < v < 12.60: desv_st = True
            except Exception: pass
            try:
                v = float(str(rd.get("crioscopia_ruta","")).replace(",","."))
                if v > -0.535: desv_ic = True          # IC menos negativo que -0.535 → agua
            except Exception: pass
            alt = (ri % 2 == 0)
            for ci, (_, col) in enumerate(cols1, 1):
                fmt = "0.00"  if col in ("solidos_ruta","st_pond") else \
                      "0.000" if col in ("crioscopia_ruta","ic_pond") else None
                # Resaltar rojo solo en parámetros fuera de umbral (+ RUTA si hay desviación)
                bad = ((desv_st or desv_ic) and col == "ruta") or \
                      (desv_st and col == "solidos_ruta")          or \
                      (desv_st and col == "st_pond")               or \
                      (desv_ic and col == "crioscopia_ruta")       or \
                      (desv_ic and col == "ic_pond")
                _wc(ws1, ri, ci, rd.get(col,""), fmt=fmt, bad=bad, alt=alt)

    # ── Hoja TRANSUIZA ────────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "TRANSUIZA"):
        df_t = (df_filtrado[df_filtrado["tipo_seguimiento"] == "TRANSUIZA"].copy()
                if "tipo_seguimiento" in df_filtrado.columns else df_filtrado.copy())
        ws2 = wb.create_sheet("TRANSUIZA")
        # Una sola hoja con: FECHA ANÁLISIS, FECHA MUESTRA, GRASA, ST, PROTEÍNA, ST CARROTANQUE
        cols2 = [
            ("FECHA ANÁLISIS",       "guardado_en"),
            ("FECHA MUESTRA",        "fecha"),
            ("GRASA (%)",            "grasa_muestra"),
            ("ST MUESTRA (%)",       "solidos_ruta"),
            ("PROTEÍNA (%)",         "proteina_muestra"),
            ("ST CARROTANQUE (%)",   "st_carrotanque"),
            ("DIFERENCIA SÓLIDOS",   "diferencia_solidos"),
        ]
        _wh(ws2, [h for h, _ in cols2], [20, 16, 14, 16, 14, 20, 18])
        for ri, row in enumerate(df_t.itertuples(index=False), start=2):
            rd = row._asdict(); alt = (ri % 2 == 0)
            for ci, (_, col) in enumerate(cols2, 1):
                fmt = "0.00" if col in ("grasa_muestra", "solidos_ruta",
                                        "proteina_muestra", "st_carrotanque",
                                        "diferencia_solidos") else None
                _wc(ws2, ri, ci, rd.get(col, ""), fmt=fmt, alt=alt)

    # ── Hoja SEGUIMIENTOS ─────────────────────────────────────────────────
    if filtro_tipo in ("TODOS", "SEGUIMIENTOS"):
        if filtro_tipo == "SEGUIMIENTOS":
            df_seg = df_filtrado.copy()
        else:
            df_seg = load_seguimientos()
            if "_fecha_dt" in df_seg.columns:
                df_seg = df_seg[
                    (df_seg["_fecha_dt"].dt.date >= fecha_desde) &
                    (df_seg["_fecha_dt"].dt.date <= fecha_hasta)
                ]
        if filtro_subtipo != "TODOS" and "sub_tipo_seguimiento" in df_seg.columns:
            df_seg = df_seg[df_seg["sub_tipo_seguimiento"] == filtro_subtipo]
        df_seg = df_seg.drop(columns=["_fecha_dt","_estado"], errors="ignore")
        ws3 = wb.create_sheet("SEGUIMIENTOS")
        cols3 = [
            ("SUB-TIPO","sub_tipo_seguimiento"), ("FECHA","fecha"),
            ("CÓDIGO","seg_codigo"), ("ENTREGADO POR","seg_quien_trajo"),
            ("RUTA","ruta"), ("RESPONSABLE","seg_responsable"),
            ("ID MUESTRA","seg_id_muestra"), ("GRASA (%)","seg_grasa"),
            ("ST (%)","seg_st"), ("IC (°C)","seg_ic"), ("AGUA (%)","seg_agua"),
            ("ALCOHOL","seg_alcohol"), ("CLORUROS","seg_cloruros"),
            ("NEUTRALIZANTES","seg_neutralizantes"),
            ("OBSERVACIONES","seg_observaciones"), ("GUARDADO EN","guardado_en"),
        ]
        _wh(ws3, [h for h, _ in cols3],
            [18, 12, 12, 18, 16, 18, 14, 10, 10, 10, 10, 12, 12, 16, 30, 18])
        for ri, row in enumerate(df_seg.itertuples(index=False), start=2):
            rd = row._asdict(); alt = (ri % 2 == 0)
            for ci, (_, col) in enumerate(cols3, 1):
                fmt = "0.00"  if col in ("seg_grasa","seg_st","seg_agua") else \
                      "0.000" if col == "seg_ic" else None
                _wc(ws3, ri, ci, rd.get(col,""), fmt=fmt, alt=alt)

    # ── Hoja ESTACIONES (de las rutas seleccionadas) ─────────────────────
    if filtro_tipo in ("TODOS", "RUTAS"):
        df_re = (df_filtrado[df_filtrado["tipo_seguimiento"] == "RUTAS"].copy()
                 if "tipo_seguimiento" in df_filtrado.columns else df_filtrado.copy())
        ws4 = wb.create_sheet("ESTACIONES")
        # Columnas Hoja 2
        # ci: 1=FECHA,2=RUTA,3=PLACA,4=CONDUCTOR,5=VOL.DECL,6=#EST,7=CÓDIGO,
        #     8=GRASA,9=SÓL.TOT,10=PROTEÍNA,11=CRIOSCOPIA,12=VOLUMEN,
        #     13=ALCOHOL,14=CLORUROS,15=NEUTRALIZANTES,16=%AGUA,
        #     17=OBSERVACIONES,18=ST RUTA,19=IC RUTA,20=ESTADO
        hdrs4 = [
            "FECHA","RUTA","PLACA","CONDUCTOR","VOL. DECLARADO (L)",
            "# ESTACIÓN","CÓDIGO","GRASA (%)","SÓL.TOT. (%)","PROTEÍNA (%)",
            "CRIOSCOPIA (°C)","VOLUMEN (L)","ALCOHOL","CLORUROS","NEUTRALIZANTES",
            "% AGUA","OBSERVACIONES","ST RUTA (%)","IC RUTA (°C)","ESTADO CALIDAD",
        ]
        _wh(ws4, hdrs4,
            [12,18,10,18,16,10,14,10,10,10,14,12,10,10,14,9,26,12,12,14])
        est_ri = 2
        for _, ruta_row in df_re.iterrows():
            raw_json = str(ruta_row.get("estaciones_json","") or "")
            try: ests = json.loads(raw_json) if raw_json.strip() else []
            except Exception: ests = []
            if not ests: continue
            try: st_rv = float(str(ruta_row.get("solidos_ruta","")).replace(",","."))
            except Exception: st_rv = None
            try: ic_rv = float(str(ruta_row.get("crioscopia_ruta","")).replace(",","."))
            except Exception: ic_rv = None
            # Estado global de la ruta (para columna ESTADO CALIDAD)
            desv_st_r = st_rv is not None and 0 < st_rv < 12.60
            desv_ic_r = ic_rv is not None and ic_rv > -0.535
            estado_r  = "DESVIACIÓN" if (desv_st_r or desv_ic_r) else "CONFORME"

            for idx_e, est in enumerate(ests, 1):
                # ── Desviaciones por estación ─────────────────────────────
                # ST estación < 12.6
                try:
                    st_e = float(str(est.get("solidos","")).replace(",","."))
                    desv_st_e = 0 < st_e < 12.60
                except Exception:
                    st_e = None; desv_st_e = False
                # Crioscopia estación > -0.530 (umbral estaciones es -0.530)
                try:
                    ic_e = float(str(est.get("crioscopia","")).replace(",","."))
                    desv_ic_e = ic_e > -0.530
                except Exception:
                    ic_e = None; desv_ic_e = False
                # Agua: detectar valor positivo o marca "+"
                agua_raw = est.get("agua_pct", "")
                desv_agua_e = False
                try:
                    agua_v = float(str(agua_raw).replace(",",".").replace("+",""))
                    if agua_v > 0: desv_agua_e = True
                except Exception:
                    if str(agua_raw).strip() in ("+", "SI", "si", "sí", "SÍ"): desv_agua_e = True

                row_vals = [
                    ruta_row.get("fecha",""),        # 1
                    ruta_row.get("ruta",""),          # 2
                    ruta_row.get("placa",""),         # 3
                    ruta_row.get("conductor",""),     # 4
                    ruta_row.get("volumen_declarado",""),  # 5
                    idx_e,                            # 6
                    est.get("codigo",""),             # 7
                    est.get("grasa"),                 # 8
                    est.get("solidos"),               # 9
                    est.get("proteina"),              # 10
                    est.get("crioscopia"),            # 11
                    est.get("volumen"),               # 12
                    est.get("alcohol",""),            # 13
                    est.get("cloruros",""),           # 14
                    est.get("neutralizantes",""),     # 15
                    agua_raw,                         # 16
                    est.get("obs",""),                # 17
                    st_rv,                            # 18
                    ic_rv,                            # 19
                    estado_r,                         # 20
                ]
                fmts = [None,None,None,None,"#,##0","0",None,
                        "0.00","0.00","0.00","0.000","#,##0",
                        None,None,None,"0.0",None,"0.00","0.000",None]
                alt = (est_ri % 2 == 0)
                for ci_e, (val_e, fmt_e) in enumerate(zip(row_vals, fmts), 1):
                    # Marcar rojo según tipo de desviación por columna
                    bad_e = (
                        (ci_e == 2  and (desv_st_e or desv_ic_e or desv_agua_e)) or  # RUTA
                        (ci_e == 7  and (desv_st_e or desv_ic_e or desv_agua_e)) or  # CÓDIGO
                        (ci_e == 9  and desv_st_e)    or   # SÓL.TOT.
                        (ci_e == 11 and desv_ic_e)    or   # CRIOSCOPIA
                        (ci_e == 16 and desv_agua_e)        # % AGUA
                    )
                    _wc(ws4, est_ri, ci_e, val_e, fmt=fmt_e, bad=bad_e, alt=alt)
                est_ri += 1

    # ── Hoja ACOMPAÑAMIENTOS (muestras expandidas con ponderados) ─────────
    if filtro_tipo in ("TODOS", "SEGUIMIENTOS"):
        if filtro_tipo == "SEGUIMIENTOS":
            df_acomp_xl = df_filtrado.copy()
        else:
            df_acomp_xl = load_seguimientos()
            if "_fecha_dt" in df_acomp_xl.columns:
                df_acomp_xl = df_acomp_xl[
                    (df_acomp_xl["_fecha_dt"].dt.date >= fecha_desde) &
                    (df_acomp_xl["_fecha_dt"].dt.date <= fecha_hasta)
                ]
        df_acomp_xl = df_acomp_xl[
            df_acomp_xl.get("sub_tipo_seguimiento", pd.Series(dtype=str)) == "ACOMPAÑAMIENTOS"
        ] if "sub_tipo_seguimiento" in df_acomp_xl.columns else pd.DataFrame()
        if not df_acomp_xl.empty:
            try:
                _cat_xl = load_catalogo()
                _cat_xl_map = dict(zip(_cat_xl["codigo"], _cat_xl["nombre"]))
            except Exception:
                _cat_xl_map = {}
            ws5 = wb.create_sheet("ACOMPAÑAMIENTOS")
            hdrs5 = [
                "FECHA","RUTA","ENTREGADO POR","RESPONSABLE",
                "VOL. DECLARADO (L)","VOL. SUMA MUESTRAS (L)","DIFERENCIA (L)",
                "ST RUTA (%)","IC RUTA (°C)","ST PONDERADO (%)","IC PONDERADO (°C)",
                "# MUESTRA","CÓDIGO","NOMBRE ESTACIÓN","VOLUMEN (L)",
                "GRASA (%)","ST (%)","PROTEÍNA (%)","IC (°C)","AGUA (%)","POND ST","IC POND",
                "ALCOHOL","CLORUROS","NEUTRALIZANTES","OBSERVACIONES","GUARDADO EN",
            ]
            widths5 = [12,18,18,18,16,18,14,12,14,14,14,10,14,20,10,10,10,10,10,10,10,10,10,10,14,30,18]
            _wh(ws5, hdrs5, widths5)
            _am_ri = 2
            for _, _arow in df_acomp_xl.iterrows():
                _raw_mj = str(_arow.get("muestras_json","") or "")
                try: _muestras_xl = json.loads(_raw_mj) if _raw_mj.strip() else []
                except Exception: _muestras_xl = []
                def _pnxl(x):
                    try: return float(str(x).replace(",","."))
                    except: return None
                try: _st_rv = float(str(_arow.get("seg_solidos_ruta","")).replace(",","."))
                except: _st_rv = None
                try: _ic_rv = float(str(_arow.get("seg_crioscopia_ruta","")).replace(",","."))
                except: _ic_rv = None
                try: _st_pv = float(str(_arow.get("seg_st_pond","")).replace(",","."))
                except: _st_pv = None
                try: _ic_pv = float(str(_arow.get("seg_ic_pond","")).replace(",","."))
                except: _ic_pv = None
                try: _vol_decl_xl = int(float(str(_arow.get("seg_vol_declarado","")).replace(",",".")))
                except: _vol_decl_xl = None
                try: _vol_sum_xl  = int(float(str(_arow.get("seg_vol_muestras","")).replace(",",".")))
                except: _vol_sum_xl = None
                try: _dif_xl = int(float(str(_arow.get("seg_diferencia_vol","")).replace(",",".")))
                except: _dif_xl = None
                _common5 = [
                    _arow.get("fecha",""), _arow.get("ruta",""),
                    _arow.get("seg_quien_trajo",""), _arow.get("seg_responsable",""),
                    _vol_decl_xl, _vol_sum_xl, _dif_xl,
                    _st_rv, _ic_rv, _st_pv, _ic_pv,
                ]
                _alt5 = (_am_ri % 2 == 0)
                if not _muestras_xl:
                    for _ci5, _v5 in enumerate(_common5 + ["—"]*14 + [_arow.get("guardado_en","")], 1):
                        _wc(ws5, _am_ri, _ci5, _v5, alt=_alt5)
                    _am_ri += 1
                    continue
                for _idx5, _am5 in enumerate(_muestras_xl, 1):
                    _cod5 = str(_am5.get("ID","") or "").strip()
                    _vol5 = _pnxl(_am5.get("_volumen"))
                    _st5  = _pnxl(_am5.get("_st"))
                    _ic5  = _pnxl(_am5.get("_ic"))
                    _pst5 = round(_vol5 * _st5, 2) if _vol5 is not None and _st5 is not None else None
                    _pic5 = round(_vol5 * _ic5, 3) if _vol5 is not None and _ic5 is not None else None
                    _row5 = _common5 + [
                        _idx5, _cod5, _cat_xl_map.get(_cod5,""),
                        int(_vol5) if _vol5 is not None else None,
                        _pnxl(_am5.get("_grasa")), _st5, _pnxl(_am5.get("_proteina")), _ic5,
                        _pnxl(_am5.get("_agua")), _pst5, _pic5,
                        _am5.get("_alcohol",""), _am5.get("_cloruros",""),
                        _am5.get("_neutralizantes",""), _am5.get("_obs",""),
                        _arow.get("guardado_en",""),
                    ]
                    _alt5 = (_am_ri % 2 == 0)
                    for _ci5, _v5 in enumerate(_row5, 1):
                        _hdr5 = hdrs5[_ci5-1]
                        _fmt5 = "0.00"  if _hdr5 in ("ST RUTA (%)","ST PONDERADO (%)","GRASA (%)","ST (%)","PROTEÍNA (%)","AGUA (%)","POND ST") else \
                                "0.000" if _hdr5 in ("IC RUTA (°C)","IC PONDERADO (°C)","IC (°C)","IC POND") else None
                        _wc(ws5, _am_ri, _ci5, _v5, fmt=_fmt5, alt=_alt5)
                    _am_ri += 1

    if not wb.sheetnames:
        wb.create_sheet("Sin datos")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



import random as _rnd



_DATOS_LECHE = [
    "¿Sabías que la leche es uno de los alimentos más completos que existen?",
    "La crioscopía es vital para detectar si se añadió agua a la leche.",
    "Los sólidos totales determinan el rendimiento para la producción de leche en polvo.",
    "Una vaca lechera produce en promedio entre 20 y 35 litros de leche al día.",
    "La leche fresca debe conservarse por debajo de 4 °C para mantener su calidad.",
    "Colombia es uno de los principales productores de leche en América Latina.",
    "La acidez óptima de la leche fresca está entre 0.14 % y 0.18 % de ácido láctico.",
    "El contenido de grasa varía según la raza de la vaca y el tipo de alimentación.",
    "La leche contiene los 9 aminoácidos esenciales que el cuerpo humano no puede producir.",
]

def _is_admin() -> bool:
    return st.session_state.get("_rol_usuario") == "ADMINISTRADOR"

st.set_page_config(page_title="QualiLact", page_icon="🧪", layout="wide")
restore_draft_state()

with open("attached_assets/image_1777229405853.png", "rb") as _logo_file:
    _logo_b64 = base64.b64encode(_logo_file.read()).decode("utf-8")

with open("logo_qualilact_brown.png", "rb") as _ql_file:
    _ql_logo_b64 = base64.b64encode(_ql_file.read()).decode("utf-8")

with open("logo_qualilact_cropped.png", "rb") as _ql_crop_file:
    _ql_logo_crop_b64 = base64.b64encode(_ql_crop_file.read()).decode("utf-8")

# ── GUARD DE AUTENTICACIÓN ───────────────────────────────────────────────────
if not st.session_state.get("_logged_in", False):
    st.markdown(
        """<style>
        #MainMenu{visibility:hidden;}
        [data-testid="stHeader"]{display:none;}
        footer{visibility:hidden;}
        [data-testid="stSidebar"]{display:none;}
        .block-container{padding-top:2.5rem!important;}
        </style>""",
        unsafe_allow_html=True,
    )
    _lc, _cm, _rc = st.columns([1, 2, 1])
    with _cm:
        st.markdown(
            f'<div style="background:linear-gradient(145deg,#F0F5FB 0%,#FFFFFF 60%,#EAF2FB 100%);'
            f'border-radius:20px;border:1px solid #D6E4F5;padding:28px 32px 22px;'
            f'box-shadow:0 4px 18px rgba(0,86,163,0.10);margin-bottom:22px;">'
            f'<div style="display:flex;align-items:center;justify-content:center;gap:40px;">'
            f'<img src="data:image/png;base64,{_ql_logo_crop_b64}" '
            f'style="height:165px;width:165px;object-fit:contain;display:block;'
            f'filter:drop-shadow(0 3px 8px rgba(99,79,58,0.22));" alt="QualiLact">'
            f'<div style="width:1px;height:140px;background:linear-gradient(to bottom,transparent,#B8C9DF,transparent);flex-shrink:0;"></div>'
            f'<img src="data:image/png;base64,{_logo_b64}" '
            f'style="height:165px;width:165px;object-fit:contain;display:block;'
            f'filter:drop-shadow(0 3px 8px rgba(0,0,0,0.12));" alt="Nestlé">'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="text-align:center;margin-bottom:24px;">'
            '<h2 style="color:#0056A3;font-weight:800;margin:0;font-size:1.9rem;">QualiLact</h2>'
            '<p style="color:#6B7280;font-size:0.82rem;margin:4px 0 0;">'
            'Control de Calidad Láctea &nbsp;·&nbsp; TROMS - Milksourcing</p></div>',
            unsafe_allow_html=True,
        )
        with st.form("_login_form"):
            _li_u = st.text_input("👤 Usuario", placeholder="Ingrese su usuario")
            _li_p = st.text_input("🔒 Contraseña", type="password", placeholder="Ingrese su contraseña")
            _li_submitted = st.form_submit_button(
                "Iniciar Sesión", type="primary", use_container_width=True
            )

        if _li_submitted:
            try:
                respuesta = supabase.table("usuarios_app").select("*").eq("nombre_usuario", _li_u).eq("contrasena", _li_p).execute()
                
                if len(respuesta.data) > 0:
                    usuario_db = respuesta.data[0] 
                    st.session_state._logged_in      = True
                    st.session_state._rol_usuario    = usuario_db["rol"]
                    st.session_state._nombre_usuario = usuario_db["nombre_usuario"] 
                    st.session_state._usuario_login  = _li_u
                    st.session_state._dato_leche     = _rnd.choice(_DATOS_LECHE)
                    st.session_state._just_logged_in = True
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos. Verifique sus credenciales.")
            except Exception as e:
                st.error(f"Error de conexión: {e}")

        st.markdown(
            '<div style="text-align:center;margin-top:18px;font-size:0.72rem;color:#9CA3AF;">'
            'Acceso restringido — solo personal autorizado Milksourcing</div>',
            unsafe_allow_html=True,
        )
    st.stop()

# ── BIENVENIDA (solo en el primer rerun tras login) ──────────────────────────
if st.session_state.get("_just_logged_in", False):
    st.session_state._just_logged_in = False
    st.info(f"💡 **Dato curioso:** {st.session_state._dato_leche}")

st.markdown(
    f"""
    <div style="
        display: flex;
        justify-content: space-between;
        align-items: center;
        width: 100%;
        padding: 12px 8px 8px 8px;
        border-bottom: 2px solid #0056A3;
    ">
        <div style="display:flex;flex-direction:column;align-items:flex-start;">
            <div style="
                font-size: 2rem;
                font-weight: 800;
                color: #0056A3;
                letter-spacing: 1px;
                line-height: 1.1;
                font-family: 'Segoe UI', sans-serif;
            ">QualiLact</div>
            <div style="
                font-size: 0.9rem;
                color: #6B7280;
                font-weight: 400;
                letter-spacing: 0.5px;
                margin-top: 2px;
                font-family: 'Segoe UI', sans-serif;
            ">Control de Calidad en Leche Fresca</div>
            <div style="
                font-size: 0.72rem;
                color: #634F3A;
                font-weight: 700;
                letter-spacing: 0.4px;
                margin-top: 5px;
                font-family: 'Segoe UI', sans-serif;
            ">TROMS &nbsp;·&nbsp; <span style="font-weight:400;font-style:italic;color:#8B7355;">Transformación Operativa Milksourcing</span></div>
        </div>
        <img src="data:image/png;base64,{_logo_b64}" alt="Nestlé logo"
             style="
                width: min(10vw, 90px);
                max-width: 90px;
                height: auto;
                display: block;
                object-fit: contain;
                margin: 0;
                padding: 0;
             ">
    </div>
    <style>
    @media (max-width: 640px) {{
        div[style*="justify-content: space-between"] {{
            padding: 10px 8px 6px 8px !important;
        }}
        div[style*="justify-content: space-between"] img {{
            width: min(12vw, 46px) !important;
            max-width: 46px !important;
        }}
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    /* ── Indicador de carga: ciclo de iconos ───────────────────── */
    @keyframes _ql_icono {
        0%,  30%  { content: "🐄  Cargando..."; }
        33%, 63%  { content: "🥛  Cargando..."; }
        66%, 96%  { content: "🧪  Cargando..."; }
        100%      { content: "🐄  Cargando..."; }
    }
    /* Ocultar spinner original */
    [data-testid="stStatusWidget"] > div > div { display: none !important; }
    /* Mostrar icono ciclico fijo en parte superior derecha */
    [data-testid="stStatusWidget"]::after {
        content: "🐄  Cargando...";
        position: fixed;
        top: 10px;
        right: 24px;
        font-size: 1rem;
        font-weight: 600;
        color: #0056A3;
        background: #EAF1FA;
        border: 1.5px solid #BDD7EE;
        border-radius: 20px;
        padding: 4px 14px;
        display: block;
        animation: _ql_icono 2.4s linear infinite;
        line-height: 1.6;
        z-index: 9999;
        pointer-events: none;
        white-space: nowrap;
    }


    /* ── Fondo azul claro QualiLact ────────────────────────────── */
    .stApp { background-color: #F0F5F9 !important; }
    section[data-testid="stSidebar"] { background-color: #E8EFF5 !important; }
    /* ── Tarjetas blancas redondeadas ──────────────────────────── */
    div[data-testid="stVerticalBlock"] > div[data-testid="element-container"],
    div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #FFFFFF;
    }
    section.main > div { background-color: #F0F5F9 !important; }

    /* ── Ocultar spinners de number input ──────────────────────── */
    input[type="number"]::-webkit-outer-spin-button,
    input[type="number"]::-webkit-inner-spin-button {
        -webkit-appearance: none; margin: 0;
    }
    input[type="number"] { -moz-appearance: textfield; }
    button[data-testid="stNumberInputStepUp"],
    button[data-testid="stNumberInputStepDown"] { display: none !important; }

    /* ── Inputs: fondo gris claro + bordes redondeados ─────────── */
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stDateInput"] input {
        background-color: #F4F4F4 !important;
        border: 1.5px solid #D1D5DB !important;
        border-radius: 8px !important;
        color: #1F2937 !important;
        font-size: 14px !important;
        padding: 8px 12px !important;
        transition: border-color 0.18s, box-shadow 0.18s;
    }

    /* ── Focus: resaltado azul Nestlé ──────────────────────────── */
    div[data-testid="stTextInput"] input:focus,
    div[data-testid="stNumberInput"] input:focus,
    div[data-testid="stDateInput"] input:focus {
        border-color: #0056A3 !important;
        box-shadow: 0 0 0 3px rgba(0, 86, 163, 0.12) !important;
        background-color: #FFFFFF !important;
        outline: none !important;
    }

    /* ── Labels: gris oscuro, peso semibold ────────────────────── */
    div[data-testid="stTextInput"] label p,
    div[data-testid="stNumberInput"] label p,
    div[data-testid="stDateInput"] label p,
    div[data-testid="stSelectbox"] label p {
        color: #555555 !important;
        font-weight: 600 !important;
        font-size: 12.5px !important;
        letter-spacing: 0.3px;
    }

    /* ── Selectbox: mismos bordes redondeados ──────────────────── */
    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:first-child {
        background-color: #F4F4F4 !important;
        border: 1.5px solid #D1D5DB !important;
        border-radius: 8px !important;
    }

    /* ── Divisores ─────────────────────────────────────────────── */
    hr { border-color: #E5E7EB !important; }

    /* ── Botones primarios: azul Nestlé ────────────────────────── */
    button[kind="primary"], button[data-testid*="primary"] {
        background-color: #0056A3 !important;
        border-color: #0056A3 !important;
        border-radius: 8px !important;
    }
    button[kind="primary"]:hover {
        background-color: #004285 !important;
    }

    /* ── Contenedores con borde ────────────────────────────────── */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        border-radius: 10px !important;
        border-color: #E5E7EB !important;
        background-color: #FAFAFA !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

if "continuar" not in st.session_state:
    st.session_state.continuar = False

if "pagina_activa" not in st.session_state:
    st.session_state.pagina_activa = "REGISTRAR"

for _sk, _sv in [
    ("admin_accion", None), ("admin_idx", None), ("admin_idxs", []), ("admin_from_seg", False),
    ("hist_buscar_ok", False),
    ("tipo_registrar", "RUTAS"), ("sub_tipo_registrar", "ESTACIONES"),
    ("registrar_submenu_open", False),
    ("_sidebar_close", False),
    ("_logged_in", False), ("_rol_usuario", ""), ("_usuario_login", ""),
    ("_nombre_usuario", ""), ("_dato_leche", ""), ("_just_logged_in", False),
]:
    if _sk not in st.session_state:
        st.session_state[_sk] = _sv


def convertir_a_mayusculas(campo):
    st.session_state[campo] = st.session_state[campo].upper()


def sanitizar_nombre_ruta(campo):
    val = st.session_state.get(campo, "")
    st.session_state[campo] = re.sub(r"[^A-ZÁÉÍÓÚÑÜ0-9]", "", val.upper())


def validar_placa():
    _k = f"placa_vehiculo_{st.session_state.get('_ruta_fg', 0)}"
    if _k in st.session_state:
        st.session_state[_k] = re.sub(
            r"[^A-Z0-9]", "", st.session_state[_k].upper()
        )


def activar_siguiente_con_enter():
    # Inyectar solo una vez por sesión de página.
    # window._navInstalled garantiza que no se dupliquen listeners en reruns.
    if st.session_state.get("_nav_js_ok"):
        return
    st.session_state["_nav_js_ok"] = True
    # st.html(unsafe_allow_javascript=True) → inlined in the main page,
    # no iframe, no window.parent needed, no React synthetic-event conflict.
    st.html(
        """
        <script>
        (function(){
          if (window._navInstalled) return;
          window._navInstalled = true;

          // ── Selectores Streamlit (incluye selectbox) ──────────────────
          var ST_FORM = [
            '[data-testid="stTextInput"]',
            '[data-testid="stNumberInput"]',
            '[data-testid="stDateInput"]',
            '[data-testid="stTimeInput"]',
            '[data-testid="stTextArea"]',
            '[data-testid="stSelectbox"]'
          ].join(',');

          function esFormInput(el) {
            return !!(el.closest && el.closest(ST_FORM));
          }
          function esSelectbox(el) {
            return !!(el.closest && el.closest('[data-testid="stSelectbox"]'));
          }
          function selectboxAbierto(el) {
            var c = el.closest('[data-testid="stSelectbox"]');
            if (!c) return false;
            return !!(c.querySelector('[class*="menu"]'));
          }
          function hayDataEditor() {
            var g = document.querySelector('[data-testid="stDataEditor"]');
            if (!g) return false;
            var r = g.getBoundingClientRect();
            return r.width > 0 && r.height > 0;
          }
          function clickBtnGuardar() {
            var vis = function(b) { return b.offsetParent !== null && b.innerText; };
            var btns = Array.from(document.querySelectorAll('button'));
            var btn = btns.find(function(b) { return vis(b) && b.innerText.includes('AGREGAR MUESTRA'); })
                  || btns.find(function(b) { return vis(b) && b.innerText.includes('GUARDAR'); });
            if (btn) setTimeout(function() { btn.click(); }, 120);
          }
          function obtenerInputsVisibles() {
            return Array.from(document.querySelectorAll('input,textarea')).filter(function(el) {
              if (!esFormInput(el)) return false;
              var t = el.getAttribute('type');
              if (t === 'hidden' || t === 'checkbox' || t === 'radio') return false;
              var r = el.getBoundingClientRect();
              return r.width > 0 && r.height > 0;
            });
          }
          function moverFoco(input, delta) {
            var todos = obtenerInputsVisibles();
            var pos = todos.indexOf(input);
            if (pos === -1) return;
            var dest = todos[pos + delta];
            if (dest) setTimeout(function() { dest.focus(); try { dest.select(); } catch(e) {} }, 60);
          }

          var _inGrid = false;

          // focusin: salida del data-editor → ejecuta GUARDAR
          document.addEventListener('focusin', function(e) {
            if (!_inGrid) return;
            var t = e.target, tag = t.tagName;
            if ((tag === 'INPUT' || tag === 'TEXTAREA') && !esFormInput(t)) return;
            if (tag === 'CANVAS' && t.closest('[data-testid="stDataEditor"]')) return;
            _inGrid = false; clickBtnGuardar();
          });

          // ── keydown capture-phase (antes que React) ───────────────────
          document.addEventListener('keydown', function(e) {
            var a = document.activeElement;
            if (!a) return;
            var tag = a.tagName;
            if (tag !== 'INPUT' && tag !== 'TEXTAREA') return;

            if (esFormInput(a)) {
              var esSel = esSelectbox(a);
              // Si el dropdown está abierto dejar que React lo maneje
              if (esSel && selectboxAbierto(a)) return;

              var esNum = a.getAttribute('type') === 'number';
              var todos = obtenerInputsVisibles();
              var pos   = todos.indexOf(a);
              var esUlt = pos !== -1 && pos === todos.length - 1;
              var ph    = (a.placeholder || '').toLowerCase();

              if (e.key === 'Enter') {
                e.preventDefault(); e.stopPropagation();
                if (ph.includes('observaciones') || esUlt) clickBtnGuardar();
                else moverFoco(a, 1);
              } else if (e.key === 'ArrowRight' && !esSel && (esNum || a.selectionStart >= a.value.length)) {
                e.preventDefault(); e.stopPropagation(); moverFoco(a, 1);
              } else if (e.key === 'ArrowLeft' && !esSel && (esNum || a.selectionStart <= 0)) {
                e.preventDefault(); e.stopPropagation(); moverFoco(a, -1);
              }
              return;
            }

            // ── data-editor grid ──────────────────────────────────────
            if (!hayDataEditor()) return;
            var isNum = a.getAttribute('type') === 'number';
            var tab = function(sh) {
              a.dispatchEvent(new KeyboardEvent('keydown', {
                key: 'Tab', code: 'Tab', keyCode: 9, which: 9,
                shiftKey: !!sh, bubbles: true, cancelable: true, composed: true
              }));
            };
            if (e.key === 'Enter') {
              e.preventDefault(); e.stopPropagation(); _inGrid = true; tab(false);
            } else if (e.key === 'ArrowRight' && (isNum || a.selectionStart >= a.value.length)) {
              e.preventDefault(); e.stopPropagation(); tab(false);
            } else if (e.key === 'ArrowLeft' && (isNum || a.selectionStart <= 0)) {
              e.preventDefault(); e.stopPropagation(); tab(true);
            }
          }, true); // capture phase
        })();
        </script>
        """,
        unsafe_allow_javascript=True,
    )

# ── SIDEBAR DE NAVEGACIÓN ────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        f'<div style="text-align:center;padding:16px 14px 18px;background:linear-gradient(160deg,#EAF2FB 0%,#F6F9FC 100%);border-radius:14px;border-bottom:3px solid #0056A3;margin-bottom:22px;box-shadow:0 3px 12px rgba(0,86,163,0.10);"><div style="margin:4px auto 8px;"><img src="data:image/png;base64,{_ql_logo_b64}" style="width:185px;height:auto;display:block;margin:0 auto;filter:drop-shadow(0 3px 8px rgba(99,79,58,0.25));" alt="QualiLact logo"></div><div style="font-size:0.74rem;color:#6B7280;letter-spacing:0.3px;font-family:Segoe UI,sans-serif;margin-top:0px;margin-bottom:10px;">Control de Calidad L&#225;ctea</div><div style="height:1px;background:linear-gradient(90deg,transparent,#B8966A,transparent);margin:0 18px 12px;">&nbsp;</div><div style="display:inline-block;background:#634F3A;color:#FFF8F0;font-size:0.65rem;font-weight:800;letter-spacing:2px;padding:3px 14px;border-radius:20px;font-family:Segoe UI,sans-serif;margin-bottom:6px;box-shadow:0 2px 6px rgba(99,79,58,0.30);">TROMS</div><div style="font-size:0.63rem;color:#8B7355;font-style:italic;font-family:Segoe UI,sans-serif;line-height:1.5;">Transformaci&#243;n Operativa Milk Sourcing</div></div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <style>
        /* Sidebar nav buttons — full width, styled */
        [data-testid="stSidebar"] .stButton > button {
            width: 100%;
            border-radius: 8px;
            font-weight: 600;
            font-size: 0.95rem;
            padding: 10px 14px;
            margin-bottom: 6px;
            border: 2px solid transparent;
            background: #F0F4FA;
            color: #1F2937;
            transition: all 0.15s;
        }
        [data-testid="stSidebar"] .stButton > button:hover {
            background: #DBEAFE;
            border-color: #0056A3;
            color: #0056A3;
        }
        /* Botón activo (tipo primary = página seleccionada) */
        [data-testid="stSidebar"] .stButton > button[kind="primary"] {
            background: #0056A3 !important;
            color: #FFFFFF !important;
            border-color: #003D7A !important;
        }
        [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
            background: #003D7A !important;
            color: #FFFFFF !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ── Botón REGISTRAR ──────────────────────────────────────────────────────
    _reg_active = st.session_state.pagina_activa == "REGISTRAR"
    _arrow = "▾" if st.session_state.registrar_submenu_open else "▸"
    _reg_lbl = f"📝 **Registrar** {_arrow}" if _reg_active else f"📝 Registrar {_arrow}"
    if st.button(
        _reg_lbl,
        key="_nav_REGISTRAR",
        width='stretch',
        type="primary" if _reg_active else "secondary",
    ):
        if st.session_state.pagina_activa == "REGISTRAR":
            st.session_state.registrar_submenu_open = not st.session_state.registrar_submenu_open
        else:
            st.session_state.pagina_activa = "REGISTRAR"
            st.session_state.registrar_submenu_open = True
        st.rerun()

    # ── Sub-menú inline (justo debajo de Registrar) ──────────────────────────
    if st.session_state.registrar_submenu_open:
        _tipo_opts = ["RUTAS", "TRANSUIZA", "SEGUIMIENTOS"]
        _t_icons = {"RUTAS": "🚛", "TRANSUIZA": "🏭", "SEGUIMIENTOS": "🔬"}
        st.markdown(
            "<div style='margin:4px 0 4px 10px;border-left:3px solid #0056A3;"
            "padding-left:8px;'>"
            "<span style='font-size:0.68rem;font-weight:700;color:#6B7280;"
            "letter-spacing:.06em;'>TIPO DE ANÁLISIS</span></div>",
            unsafe_allow_html=True,
        )
        for _t in _tipo_opts:
            _t_active = (
                st.session_state.pagina_activa == "REGISTRAR"
                and st.session_state.tipo_registrar == _t
            )
            _t_lbl = f"  {_t_icons[_t]} **{_t}**" if _t_active else f"  {_t_icons[_t]} {_t}"
            if st.button(
                _t_lbl,
                key=f"_subnav_{_t}",
                width='stretch',
                type="primary" if _t_active else "secondary",
            ):
                st.session_state.pagina_activa = "REGISTRAR"
                st.session_state.tipo_registrar = _t
                st.session_state.registrar_submenu_open = False
                st.session_state._sidebar_close = True
                st.rerun()

    # ── Botón HISTORIAL ──────────────────────────────────────────────────────
    _hist_active = st.session_state.pagina_activa == "HISTORIAL"
    _hist_lbl = "🗂️ **Historial**" if _hist_active else "🗂️ Historial"
    if st.button(
        _hist_lbl,
        key="_nav_HISTORIAL",
        width='stretch',
        type="primary" if _hist_active else "secondary",
    ):
        st.session_state.pagina_activa = "HISTORIAL"
        st.session_state.registrar_submenu_open = False
        st.session_state._sidebar_close = True
        st.rerun()

    # ── Botón DASHBOARD ──────────────────────────────────────────────────────
    _dash_active = st.session_state.pagina_activa == "DASHBOARD"
    _dash_lbl = "📊 **Dashboard**" if _dash_active else "📊 Dashboard"
    if st.button(
        _dash_lbl,
        key="_nav_DASHBOARD",
        width='stretch',
        type="primary" if _dash_active else "secondary",
    ):
        st.session_state.pagina_activa = "DASHBOARD"
        st.session_state.registrar_submenu_open = False
        st.session_state._sidebar_close = True
        st.rerun()

    _est_active = st.session_state.pagina_activa == "REGISTRAR" and st.session_state.tipo_registrar == "ESTACIONES"
    _est_lbl = "🏷️ **Estaciones**" if _est_active else "🏷️ Estaciones"
    if st.button(
        _est_lbl,
        key="_nav_ESTACIONES",
        width='stretch',
        type="primary" if _est_active else "secondary",
    ):
        st.session_state.pagina_activa = "REGISTRAR"
        st.session_state.tipo_registrar = "ESTACIONES"
        st.session_state.registrar_submenu_open = False
        st.session_state.cat_accion = None
        st.session_state._sidebar_close = True
        st.rerun()

    st.markdown("<hr style='border-color:#E5E7EB;margin:18px 0;'>", unsafe_allow_html=True)
    st.markdown(
        f"<div style='font-size:0.72rem;color:#9CA3AF;text-align:center;'>"
        f"Sección activa:<br><strong style='color:#0056A3;'>"
        f"{st.session_state.pagina_activa}</strong></div>",
        unsafe_allow_html=True,
    )
    _sb_hora = now_col().hour
    _sb_saludo = "Buenos días" if _sb_hora < 12 else ("Buenas tardes" if _sb_hora < 18 else "Buenas noches")
    _sb_rol_icon = "🔑" if st.session_state._rol_usuario == "ADMINISTRADOR" else "👤"
    _sb_rol_color = "#0056A3" if st.session_state._rol_usuario == "ADMINISTRADOR" else "#059669"
    _sb_rol_label = "Administrador" if st.session_state._rol_usuario == "ADMINISTRADOR" else "Operador"
    _es_admin = st.session_state._rol_usuario == "ADMINISTRADOR"
    _sb_nombre_html = (
        f"<div style='font-size:0.95rem;font-weight:800;color:#1E293B;margin-bottom:6px;'>"
        f"Abimelec T.</div>"
        if _es_admin else ""
    )
    st.markdown(
        f"<div style='background:linear-gradient(135deg,#EAF2FB 0%,#F0F7FF 100%);"
        f"border-radius:14px;border:1px solid #C8DEFA;padding:12px 14px 10px;"
        f"margin-bottom:10px;box-shadow:0 2px 8px rgba(0,86,163,0.08);'>"
        f"<div style='font-size:0.72rem;color:#6B7280;margin-bottom:4px;'>{_sb_saludo}</div>"
        f"{_sb_nombre_html}"
        f"<div style='display:inline-flex;align-items:center;gap:5px;"
        f"background:{_sb_rol_color};color:#fff;font-size:0.62rem;font-weight:700;"
        f"letter-spacing:.06em;padding:2px 8px;border-radius:20px;'>"
        f"{_sb_rol_icon}&nbsp;{_sb_rol_label}</div>"
        f"<div style='font-size:0.65rem;color:#9CA3AF;margin-top:5px;'>"
        f"{st.session_state._usuario_login}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )
    if st.button("🚪 Cerrar Sesión", key="_btn_logout", width='stretch', type="secondary"):
        for _k in list(st.session_state.keys()):
            del st.session_state[_k]
        st.rerun()


if st.session_state._sidebar_close:
    st.session_state._sidebar_close = False
    import streamlit.components.v1 as _stc
    _stc.html(
        """<script>
        (function(){
            var attempts = 0;
            function tryClose(){
                var doc = window.parent.document;
                // Strategy 1: known data-testid attributes
                var btn = doc.querySelector('[data-testid="stSidebarCollapseButton"] button') ||
                          doc.querySelector('button[aria-label="Close sidebar"]') ||
                          doc.querySelector('[data-testid="stSidebarCollapseButton"]');
                // Strategy 2: icon-only button (no text) inside the sidebar — that is the collapse arrow
                if(!btn){
                    var allBtns = doc.querySelectorAll('section[data-testid="stSidebar"] button');
                    for(var i=0; i<allBtns.length; i++){
                        if(allBtns[i].textContent.trim()===''){
                            btn = allBtns[i]; break;
                        }
                    }
                }
                if(btn){ btn.click(); }
                else if(attempts < 12){ attempts++; setTimeout(tryClose, 150); }
            }
            setTimeout(tryClose, 80);
        })();
        </script>""",
        height=1,
    )

if st.session_state.pagina_activa == "REGISTRAR":
    # ── Variables de tipo desde el sidebar ───────────────────────────────────
    tipo_servicio = st.session_state.tipo_registrar
    sub_tipo_seg = st.session_state.sub_tipo_registrar
    st.session_state["_tipo_servicio_guardado"] = tipo_servicio
    st.session_state["_sub_tipo_seg_guardado"] = sub_tipo_seg

    if tipo_servicio == "RUTAS":
        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;
                            margin-bottom:6px;">
                  <span style="font-size:1.35rem;">📋</span>
                  <span style="font-size:1.35rem;font-weight:700;
                               color:#0056A3;letter-spacing:.5px;
                               font-family:'Segoe UI',sans-serif;">
                    SEGUIMIENTO DE RUTAS
                  </span>
                </div>""",
            unsafe_allow_html=True,
        )

        # ── Generación de claves para reset limpio al guardar ─────────
        if "_ruta_fg" not in st.session_state:
            st.session_state._ruta_fg = 0
        _fg = st.session_state._ruta_fg

        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📋 Datos de Identificación
               </div>""",
            unsafe_allow_html=True,
        )
        r1c1, r1c2 = st.columns(2)
        fecha_ruta = r1c1.date_input(
            "📅  FECHA DE LA RUTA", now_col(),
            key=f"fecha_ruta_{_fg}", format="DD/MM/YYYY",
        )
        nombre_ruta = r1c2.text_input(
            "📍  NOMBRE DE LA RUTA", placeholder="ESCRIBA AQUÍ...",
            key=f"nombre_ruta_{_fg}", on_change=sanitizar_nombre_ruta,
            args=(f"nombre_ruta_{_fg}",),
        )
        r2c1, r2c2, r2c3 = st.columns(3)
        placa = r2c1.text_input(
            "🚚  PLACA DE VEHÍCULO", placeholder="AAA000",
            key=f"placa_vehiculo_{_fg}", on_change=validar_placa,
        )
        conductor = r2c2.text_input(
            "👤  CONDUCTOR", placeholder="NOMBRE COMPLETO",
            key=f"conductor_{_fg}", on_change=convertir_a_mayusculas,
            args=(f"conductor_{_fg}",),
        )
        volumen = r2c3.number_input(
            "📦  VOLUMEN (L)", min_value=0, value=None, step=1,
            format="%d", placeholder="DIGITE VOLUMEN", key=f"volumen_ruta_{_fg}",
        )
        activar_siguiente_con_enter()

        st.markdown("---")

        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 🧪 Análisis de Calidad de Ruta
               </div>""",
            unsafe_allow_html=True,
        )
        cq1, cq2 = st.columns(2)

        with cq1:
            solidos_raw = st.text_input(
                "SÓLIDOS TOTALES (%)",
                key=f"solidos_totales_{_fg}",
                placeholder="Ej: 12.80",
            )
            try:
                solidos_totales = float(solidos_raw.replace(",", ".")) if solidos_raw else None
            except ValueError:
                solidos_totales = None
                st.warning("⚠️ Ingrese un número válido")

            if solidos_totales is not None and 0 < solidos_totales < 12.60:
                st.error("🚨 ALERTA: SÓLIDOS POR DEBAJO DE 12.60%")
                st.markdown(
                    """
                    <style>
                    div[data-testid="stTextInput"]:has(input[aria-label="SÓLIDOS TOTALES (%)"]) input {
                        border: 2px solid red !important;
                        background-color: #fff0f0 !important;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )
            elif solidos_totales is not None and solidos_totales >= 12.60:
                st.success("✅ Sólidos dentro del parámetro")

        with cq2:
            crioscopia_raw = st.text_input(
                "CRIOSCOPIA (°C)",
                key=f"crioscopia_{_fg}",
                value="-0.",
                placeholder="-0.530",
            )
            try:
                crioscopia = float(crioscopia_raw.replace(",", ".")) if crioscopia_raw not in ("", "-", "-0", "-0.") else None
            except ValueError:
                crioscopia = None
                st.warning("⚠️ Ingrese un número válido")

            if crioscopia is not None and crioscopia > -0.535:
                st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MAYOR A -0.535)")
            elif crioscopia is not None and crioscopia < -0.550:
                st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MENOR A -0.550)")
            elif crioscopia is not None:
                st.success("✅ Crioscopia dentro del parámetro")

        st.markdown("---")

        # ── Imágenes de Muestras ───────────────────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📷 Imágenes de Muestras
               </div>""",
            unsafe_allow_html=True,
        )
        if "imagenes_confirmadas" not in st.session_state:
            st.session_state.imagenes_confirmadas = False
        if "imagenes_nombres_guardados" not in st.session_state:
            st.session_state.imagenes_nombres_guardados = []

        imagenes_subidas = st.file_uploader(
            "ADJUNTAR IMÁGENES DE MUESTRAS DE LA RUTA",
            type=["png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key=f"imagenes_muestras_{_fg}",
            label_visibility="visible",
        )

        if imagenes_subidas:
            nombres_actuales = [f.name for f in imagenes_subidas]
            if nombres_actuales != st.session_state.imagenes_nombres_guardados:
                st.session_state.imagenes_confirmadas = False

            # ── Miniaturas en cuadrícula HTML (bordes redondeados) ────
            confirmed = st.session_state.imagenes_confirmadas
            thumb_html = "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:8px 0;'>"
            for img in imagenes_subidas:
                raw_bytes = img.read()
                b64 = base64.b64encode(raw_bytes).decode()
                ext = img.name.rsplit(".", 1)[-1].lower()
                mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
                nombre_corto = img.name if len(img.name) <= 16 else img.name[:14] + "…"
                check_html = (
                    "<div style='color:#16a34a;font-size:12px;"
                    "text-align:center;font-weight:600;'>✅ Guardada</div>"
                    if confirmed else
                    f"<div style='font-size:10px;color:#888;text-align:center;'>{nombre_corto}</div>"
                )
                border_color = "#16a34a" if confirmed else "#D1D5DB"
                thumb_html += (
                    f"<div style='display:flex;flex-direction:column;"
                    f"align-items:center;gap:4px;'>"
                    f"<img src='data:{mime};base64,{b64}' "
                    f"style='width:150px;height:150px;object-fit:cover;"
                    f"border-radius:10px;border:2px solid {border_color};"
                    f"box-shadow:0 2px 6px rgba(0,0,0,0.08);background:#F4F4F4;'/>"
                    f"{check_html}</div>"
                )
                img.seek(0)  # reset cursor para uso posterior
            thumb_html += "</div>"
            st.markdown(thumb_html, unsafe_allow_html=True)

            # ── Botón guardar / confirmación ───────────────────────
            if not st.session_state.imagenes_confirmadas:
                st.markdown("<div style='margin-top:8px;'></div>",
                            unsafe_allow_html=True)
                if st.button("💾 GUARDAR IMÁGENES",
                             width='content'):
                    st.session_state.imagenes_confirmadas = True
                    st.session_state.imagenes_nombres_guardados = nombres_actuales
                    st.rerun()
            else:
                st.success("✅ Imágenes guardadas correctamente.")
        else:
            st.session_state.imagenes_confirmadas = False
            st.caption("No se han adjuntado imágenes.")

        st.markdown("---")

        if "estaciones_guardadas" not in st.session_state:
            st.session_state.estaciones_guardadas = []
        if "form_ver" not in st.session_state:
            st.session_state.form_ver = 0

        def parse_num(val, default=None):
            if val is None:
                return default
            try:
                return float(str(val).replace(",", "."))
            except ValueError:
                return default

        # ── Formulario nueva estación ──────────────────────────────────
        with st.container(border=True):
            v = st.session_state.form_ver
            num_nueva = len(st.session_state.estaciones_guardadas) + 1

            # Lookup nombre desde catálogo con el código actual en session_state
            _cat_r = load_catalogo()
            _cat_r_cod = dict(zip(_cat_r["codigo"], _cat_r["nombre"]))
            _cod_actual = st.session_state.get(f"nue_codigo_{v}", "").strip()
            _nom_actual = _cat_r_cod.get(_cod_actual) or _cat_r_cod.get(_cod_actual.upper(), "")
            if _nom_actual:
                st.markdown(
                    f"**Agregar Estación — #{num_nueva}"
                    f"&nbsp;&nbsp;<span style='color:#0056A3;font-size:1em;'>"
                    f"· {_nom_actual}</span>**",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(f"**Agregar Estación — #{num_nueva}**")

            if f"nue_crio_{v}" not in st.session_state:
                st.session_state[f"nue_crio_{v}"] = "-0."

            f1, f2, f3, f4, f5, f6 = st.columns([1.5, 1, 1, 1, 1.5, 1])
            form_codigo   = f1.text_input("CÓDIGO", key=f"nue_codigo_{v}",
                                          placeholder="CÓDIGO",
                                          on_change=convertir_a_mayusculas,
                                          args=(f"nue_codigo_{v}",))
            form_grasa    = f2.number_input("GRASA (%)", key=f"nue_grasa_{v}",
                                            min_value=0.0, max_value=100.0,
                                            step=0.01, format="%.2f",
                                            value=None, placeholder="0.00")
            form_solidos  = f3.number_input("SÓL. TOT. (%)", key=f"nue_solidos_{v}",
                                            min_value=0.0, max_value=100.0,
                                            step=0.01, format="%.2f",
                                            value=None, placeholder="0.00")
            form_proteina = f4.number_input("PROTEÍNA (%)", key=f"nue_proteina_{v}",
                                            min_value=0.0, max_value=100.0,
                                            step=0.01, format="%.2f",
                                            value=None, placeholder="0.00")
            form_crio_raw = f5.text_input("CRIOSCOPIA (°C)", key=f"nue_crio_{v}",
                                          placeholder="-0.530")
            form_vol      = f6.number_input("VOLUMEN (L)", key=f"nue_vol_{v}",
                                            min_value=0, step=1,
                                            value=None, placeholder="0")

            form_crio_val = (parse_num(form_crio_raw)
                             if form_crio_raw not in ("", "-", "-0", "-0.")
                             else None)

            if form_solidos is not None and 0 < form_solidos < 12.60:
                st.error("🚨 SÓLIDOS POR DEBAJO DE 12.60%")

            form_agua_pct = None
            if form_crio_val is not None and form_crio_val > -0.530:
                aw1, aw2 = st.columns([2, 1])
                aw1.warning("💧 ALERTA: PRESENCIA DE AGUA — CRIOSCOPIA MAYOR A -0.530")
                form_agua_pct = aw2.number_input(
                    "% AGUA AÑADIDA", key=f"nue_agua_{v}",
                    min_value=0.0, max_value=100.0, step=0.1,
                    format="%.1f", value=None, placeholder="0.0")
            elif form_crio_val is not None and form_crio_val < -0.550:
                st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MENOR A -0.550)")

            q1, q2, q3, q4 = st.columns([0.6, 0.6, 0.8, 2])
            form_alcohol  = q1.selectbox("ALCOHOL", options=["N/A", "+", "-"],
                                         key=f"nue_alcohol_{v}")
            form_cloruros = q2.selectbox("CLORUROS", options=["N/A", "+", "-"],
                                         key=f"nue_cloruros_{v}")
            form_neutral  = q3.selectbox("NEUTRALIZANTES", options=["N/A", "+", "-"],
                                         key=f"nue_neutral_{v}")
            form_obs = q4.text_input("OBSERVACIONES", key=f"nue_obs_{v}",
                                     placeholder="Ingrese observaciones...")

            activar_siguiente_con_enter()

            if st.button("💾 GUARDAR", type="primary",
                         width='stretch'):
                st.session_state.estaciones_guardadas.append({
                    "codigo":         form_codigo,
                    "grasa":          form_grasa,
                    "solidos":        form_solidos,
                    "proteina":       form_proteina,
                    "crioscopia":     form_crio_raw if form_crio_val is not None else None,
                    "volumen":        form_vol,
                    "alcohol":        form_alcohol,
                    "cloruros":       form_cloruros,
                    "neutralizantes": form_neutral,
                    "agua_pct":       form_agua_pct,
                    "obs":            form_obs,
                })
                st.session_state.form_ver += 1
                st.rerun()


        st.markdown("---")

        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📦 Calidad por Estación
               </div>""",
            unsafe_allow_html=True,
        )

        # ── Data editor ────────────────────────────────────────────────
        EDITOR_COLS = ["codigo", "grasa", "solidos", "proteina", "crioscopia",
                       "agua_pct", "volumen", "alcohol", "cloruros",
                       "neutralizantes", "obs"]

        if st.session_state.estaciones_guardadas:
            df_est = pd.DataFrame(st.session_state.estaciones_guardadas,
                                  columns=EDITOR_COLS)
        else:
            df_est = pd.DataFrame(columns=EDITOR_COLS)

        for c in ["grasa", "solidos", "proteina", "agua_pct"]:
            df_est[c] = pd.to_numeric(df_est[c], errors="coerce")
        df_est["volumen"] = pd.to_numeric(df_est["volumen"],
                                          errors="coerce").astype("Int64")

        # Columna NOMBRE ESTACIÓN derivada del catálogo (solo lectura)
        _cat_tab = load_catalogo()
        _cat_tab_map = dict(zip(_cat_tab["codigo"], _cat_tab["nombre"]))
        df_est["nombre_estacion"] = df_est["codigo"].apply(
            lambda c: _cat_tab_map.get(str(c).strip(), "")
                      if pd.notna(c) else ""
        )

        _nv = st.session_state.get("_est_nombre_ver", 0)
        edited = st.data_editor(
            df_est,
            num_rows="dynamic",
            width='stretch',
            key=f"de_est_{st.session_state.form_ver}_{_nv}",
            column_config={
                "codigo":        st.column_config.TextColumn("CÓDIGO"),
                "grasa":         st.column_config.NumberColumn(
                                     "GRASA (%)", format="%.2f",
                                     min_value=0.0, max_value=100.0),
                "solidos":       st.column_config.NumberColumn(
                                     "SÓL.TOT. (%)", format="%.2f",
                                     min_value=0.0, max_value=100.0),
                "proteina":      st.column_config.NumberColumn(
                                     "PROTEÍNA (%)", format="%.2f",
                                     min_value=0.0, max_value=100.0),
                "crioscopia":    st.column_config.TextColumn("CRIOSCOPIA (°C)"),
                "volumen":       st.column_config.NumberColumn(
                                     "VOLUMEN (L)", format="%d",
                                     min_value=0, step=1),
                "alcohol":       st.column_config.SelectboxColumn(
                                     "ALCOHOL", options=["N/A", "+", "-"],
                                     required=True),
                "cloruros":      st.column_config.SelectboxColumn(
                                     "CLORUROS", options=["N/A", "+", "-"],
                                     required=True),
                "neutralizantes":st.column_config.SelectboxColumn(
                                     "NEUTRALIZANTES", options=["N/A", "+", "-"],
                                     required=True),
                "agua_pct":      st.column_config.NumberColumn(
                                     "% AGUA", format="%.1f",
                                     min_value=0.0, max_value=100.0),
                "obs":           st.column_config.TextColumn("OBSERVACIONES"),
                "nombre_estacion": st.column_config.TextColumn(
                                     "NOMBRE ESTACIÓN", disabled=True),
            },
            hide_index=True,
        )

        # Sincronizar ediciones/eliminaciones de vuelta al estado (sin nombre_estacion)
        _prev_codes = list(df_est["codigo"].fillna("").astype(str).str.strip())
        raw = json.loads(edited.to_json(orient="records"))
        st.session_state.estaciones_guardadas = [
            {k: v for k, v in r.items() if k != "nombre_estacion"}
            for r in raw
            if any(v is not None and str(v).strip() != ""
                   for k, v in r.items() if k != "nombre_estacion")
        ]
        _new_codes = [str(r.get("codigo") or "").strip()
                      for r in st.session_state.estaciones_guardadas]
        if _prev_codes != _new_codes:
            st.session_state["_est_nombre_ver"] = st.session_state.get("_est_nombre_ver", 0) + 1
            st.rerun()

        st.markdown("---")
        # ── Reconciliación de volúmenes ────────────────────────────────
        st.markdown("---")
        vol_ruta = volumen if volumen is not None else 0
        vol_est_total = 0
        for e in st.session_state.estaciones_guardadas:
            v_e = e.get("volumen")
            if v_e is not None:
                try:
                    vol_est_total += int(v_e)
                except (ValueError, TypeError):
                    pass

        col_res1, col_res2, col_res3 = st.columns(3)
        col_res1.metric("VOLUMEN DECLARADO DE RUTA (L)",
                        f"{int(vol_ruta):,}" if vol_ruta else "—")
        col_res2.metric("VOLUMEN SUMA ESTACIONES (L)",
                        f"{int(vol_est_total):,}" if vol_est_total else "—")
        diferencia = vol_est_total - vol_ruta if vol_ruta else 0
        col_res3.metric("DIFERENCIA (L)",
                        f"{int(diferencia):+,}" if vol_ruta else "—")

        if vol_ruta and vol_est_total and vol_ruta != vol_est_total:
            st.warning(
                f"⚠️ El volumen de estaciones ({int(vol_est_total):,} L) no coincide "
                f"con el volumen declarado de la ruta ({int(vol_ruta):,} L). "
                f"Diferencia: {int(diferencia):+,} L"
            )
        elif vol_ruta and vol_est_total and vol_ruta == vol_est_total:
            st.success("✅ El volumen de estaciones coincide con el volumen de la ruta.")

        # ── Ponderados a nivel de ruta (usados en exportación y guardado) ────
        _ests = st.session_state.estaciones_guardadas
        _pond_st, _pond_ic = [], []
        for _e in _ests:
            _v  = parse_num(_e.get("volumen"))
            _s  = parse_num(_e.get("solidos"))
            _cr = parse_num(_e.get("crioscopia"))
            _pond_st.append(round(_v * _s,  2) if _v is not None and _s  is not None else None)
            _pond_ic.append(round(_v * _cr, 3) if _v is not None and _cr is not None else None)
        _vol_total = sum(parse_num(_e.get("volumen")) or 0 for _e in _ests)
        _st_pond = round(sum(x for x in _pond_st if x is not None) / _vol_total, 2) if _vol_total else None
        _ic_pond = round(sum(x for x in _pond_ic if x is not None) / _vol_total, 3) if _vol_total else None

        # ── GUARDAR RUTA EN HISTORIAL ──────────────────────────────────
        st.markdown("---")
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 💾 Guardar en Historial
               </div>""",
            unsafe_allow_html=True,
        )

        if "ruta_guardada_ok" not in st.session_state:
            st.session_state.ruta_guardada_ok = False

        if st.button(
            "💾  GUARDAR RUTA",
            type="primary",
            width='stretch',
            key="btn_guardar_ruta",
        ):
            _fotos_prefix = f"RUTAS_{(placa or 'SIN_PLACA').upper()}"
            _fotos_paths  = save_fotos_to_disk(imagenes_subidas or [], _fotos_prefix)
            save_ruta_to_csv({
                "tipo_seguimiento": "RUTAS",
                "fecha":            fecha_ruta.strftime("%d/%m/%Y") if fecha_ruta else "",
                "ruta":             nombre_ruta or "",
                "placa":            placa or "",
                "conductor":        conductor or "",
                "volumen_declarado": int(volumen) if volumen else "",
                "vol_estaciones":   int(vol_est_total) if vol_est_total else "",
                "diferencia":       int(diferencia) if vol_ruta else "",
                "solidos_ruta":     solidos_totales if solidos_totales is not None else "",
                "crioscopia_ruta":  crioscopia if crioscopia is not None else "",
                "st_pond":          _st_pond if _st_pond is not None else "",
                "ic_pond":          _ic_pond if _ic_pond is not None else "",
                "num_estaciones":   len(st.session_state.estaciones_guardadas),
                "guardado_en":      now_col().strftime("%d/%m/%Y %H:%M"),
                "estaciones_json":  json.dumps(
                    st.session_state.estaciones_guardadas, ensure_ascii=False
                ),
                "fotos_json":       json.dumps(_fotos_paths, ensure_ascii=False),
            })
            # ── Limpiar todos los campos para la siguiente ruta ────────
            # Incrementar generación → todos los widgets de identificación
            # y calidad obtienen una clave nueva y se renderizan vacíos
            st.session_state._ruta_fg = (
                st.session_state.get("_ruta_fg", 0) + 1
            )
            # Limpiar campos de estación (nue_*)
            for _k in list(st.session_state.keys()):
                if _k.startswith("nue_"):
                    st.session_state.pop(_k, None)
            st.session_state.pop("imagenes_muestras", None)
            st.session_state.estaciones_guardadas          = []
            st.session_state.form_ver                      = (
                st.session_state.get("form_ver", 0) + 1
            )
            st.session_state.imagenes_confirmadas          = False
            st.session_state.imagenes_nombres_guardados    = []
            st.session_state.ruta_guardada_ok              = True
            clear_draft_state()
            st.rerun()

        if st.session_state.ruta_guardada_ok:
            st.success(
                "✅ Ruta guardada exitosamente en el historial. "
                "Puedes consultarla en **📊 Historial de Rutas** al final de la página."
            )
            st.session_state.ruta_guardada_ok = False

    elif tipo_servicio == "TRANSUIZA":
        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
                 <span style="font-size:1.35rem;">🚛</span>
                 <span style="font-size:1.35rem;font-weight:700;color:#0056A3;
                              letter-spacing:.5px;font-family:'Segoe UI',sans-serif;">
                   SEGUIMIENTO TRANSUIZA
                 </span>
               </div>""",
            unsafe_allow_html=True,
        )

        # ── Datos de identificación ────────────────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📋 Datos de Identificación
               </div>""",
            unsafe_allow_html=True,
        )
        _trans_fg = st.session_state.get("_trans_fg", 0)

        with st.container(border=True):
            tc1, tc2, tc3 = st.columns(3)
            trans_fecha = tc1.date_input(
                "📅 FECHA", now_col(), key=f"trans_fecha_{_trans_fg}", format="DD/MM/YYYY"
            )
            _trans_placa_key = f"trans_placa_{_trans_fg}"
            trans_placa = tc2.text_input(
                "🚚 PLACA DEL VEHÍCULO", placeholder="AAA000", key=_trans_placa_key,
                on_change=lambda k=_trans_placa_key: st.session_state.__setitem__(
                    k, re.sub(r"[^A-Z0-9]", "", st.session_state.get(k, "").upper())
                ),
            )
            trans_st_carrotanque = tc3.number_input(
                "🏷️ ST DEL CARROTANQUE (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key=f"trans_st_carrotanque_{_trans_fg}",
            )
        activar_siguiente_con_enter()

        # ── Calidad de la muestra ──────────────────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 🧪 Calidad de la Muestra
               </div>""",
            unsafe_allow_html=True,
        )
        with st.container(border=True):
            qc1, qc2, qc3 = st.columns(3)
            trans_grasa = qc1.number_input(
                "GRASA (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key=f"trans_grasa_{_trans_fg}",
            )
            trans_st_muestra = qc2.number_input(
                "ST MUESTRA (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key=f"trans_st_muestra_{_trans_fg}",
            )
            trans_proteina = qc3.number_input(
                "PROTEÍNA (%)", min_value=0.0, max_value=100.0,
                step=0.01, format="%.2f", value=None, placeholder="0.00",
                key=f"trans_proteina_{_trans_fg}",
            )

            # ── Diferencia de Sólidos (automática) ────────────────────
            if trans_st_carrotanque is not None and trans_st_muestra is not None:
                dif_solidos = round(trans_st_carrotanque - trans_st_muestra, 2)
                color_dif = "#9C0006" if abs(dif_solidos) > 0.5 else "#006100"
                st.markdown(
                    f"""<div style="margin-top:12px;padding:12px 16px;
                        background:#F8FAFC;border-radius:10px;
                        border:1.5px solid #D1D5DB;text-align:center;">
                        <div style="font-size:11px;font-weight:600;color:#6B7280;
                                    letter-spacing:.4px;margin-bottom:4px;">
                            DIFERENCIA DE SÓLIDOS (ST Carrotanque − ST Muestra)
                        </div>
                        <div style="font-size:2rem;font-weight:800;color:{color_dif};">
                            {dif_solidos:+.2f} %
                        </div>
                    </div>""",
                    unsafe_allow_html=True,
                )
            else:
                dif_solidos = None
                st.info("💡 Ingrese ST del Carrotanque y ST Muestra para calcular la diferencia.")

        activar_siguiente_con_enter()

        # ── Imágenes de Muestras (TRANSUIZA) ──────────────────────────
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 📷 Imágenes de Muestras
               </div>""",
            unsafe_allow_html=True,
        )
        if "trans_imagenes_confirmadas" not in st.session_state:
            st.session_state.trans_imagenes_confirmadas = False
        if "trans_imagenes_nombres_guardados" not in st.session_state:
            st.session_state.trans_imagenes_nombres_guardados = []

        trans_imagenes_subidas = st.file_uploader(
            "ADJUNTAR IMÁGENES DE MUESTRAS TRANSUIZA",
            type=["png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key=f"trans_imagenes_muestras_{_trans_fg}",
            label_visibility="visible",
        )

        if trans_imagenes_subidas:
            _t_nombres = [f.name for f in trans_imagenes_subidas]
            if _t_nombres != st.session_state.trans_imagenes_nombres_guardados:
                st.session_state.trans_imagenes_confirmadas = False

            _t_confirmed = st.session_state.trans_imagenes_confirmadas
            _t_thumb_html = "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:8px 0;'>"
            for _timg in trans_imagenes_subidas:
                _t_raw = _timg.read()
                _t_b64 = base64.b64encode(_t_raw).decode()
                _t_ext = _timg.name.rsplit(".", 1)[-1].lower()
                _t_mime = "image/jpeg" if _t_ext in ("jpg", "jpeg") else "image/png"
                _t_nombre_corto = _timg.name if len(_timg.name) <= 16 else _timg.name[:14] + "…"
                _t_check_html = (
                    "<div style='color:#16a34a;font-size:12px;"
                    "text-align:center;font-weight:600;'>✅ Guardada</div>"
                    if _t_confirmed else
                    f"<div style='font-size:10px;color:#888;text-align:center;'>{_t_nombre_corto}</div>"
                )
                _t_border = "#16a34a" if _t_confirmed else "#D1D5DB"
                _t_thumb_html += (
                    f"<div style='display:flex;flex-direction:column;"
                    f"align-items:center;gap:4px;'>"
                    f"<img src='data:{_t_mime};base64,{_t_b64}' "
                    f"style='width:150px;height:150px;object-fit:cover;"
                    f"border-radius:10px;border:2px solid {_t_border};"
                    f"box-shadow:0 2px 6px rgba(0,0,0,0.08);background:#F4F4F4;'/>"
                    f"{_t_check_html}</div>"
                )
                _timg.seek(0)
            _t_thumb_html += "</div>"
            st.markdown(_t_thumb_html, unsafe_allow_html=True)

            if not st.session_state.trans_imagenes_confirmadas:
                st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
                if st.button("💾 GUARDAR IMÁGENES", key=f"btn_trans_guardar_imgs_{_trans_fg}",
                             width='content'):
                    st.session_state.trans_imagenes_confirmadas = True
                    st.session_state.trans_imagenes_nombres_guardados = _t_nombres
                    st.rerun()
            else:
                st.success("✅ Imágenes guardadas correctamente.")
        else:
            st.session_state.trans_imagenes_confirmadas = False
            st.caption("No se han adjuntado imágenes.")

        # ── Guardar Transuiza ──────────────────────────────────────────
        st.markdown("---")
        st.markdown(
            """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                           margin:14px 0 6px 0;letter-spacing:.4px;
                           border-left:4px solid #0056A3;padding-left:10px;">
                 💾 Guardar en Historial
               </div>""",
            unsafe_allow_html=True,
        )
        if "trans_guardado_ok" not in st.session_state:
            st.session_state.trans_guardado_ok = False

        if st.button("💾  GUARDAR TRANSUIZA", type="primary",
                     width='stretch', key="btn_guardar_trans"):
            if not trans_placa:
                st.warning("⚠️ Ingrese la placa del vehículo.")
            else:
                _t_fotos_prefix = f"TRANS_{(trans_placa or 'SIN_PLACA').upper()}"
                _t_fotos_paths  = save_fotos_to_disk(
                    trans_imagenes_subidas or [], _t_fotos_prefix
                )
                save_ruta_to_csv({
                    "tipo_seguimiento": "TRANSUIZA",
                    "fecha":            trans_fecha.strftime("%d/%m/%Y") if trans_fecha else "",
                    "ruta":             "ENTRERIOS",
                    "placa":            (trans_placa or "").upper(),
                    "st_carrotanque":   trans_st_carrotanque if trans_st_carrotanque is not None else "",
                    "solidos_ruta":     trans_st_muestra if trans_st_muestra is not None else "",
                    "grasa_muestra":    trans_grasa if trans_grasa is not None else "",
                    "proteina_muestra": trans_proteina if trans_proteina is not None else "",
                    "diferencia_solidos": dif_solidos if dif_solidos is not None else "",
                    "guardado_en":      now_col().strftime("%d/%m/%Y %H:%M"),
                    "fotos_json":       json.dumps(_t_fotos_paths, ensure_ascii=False),
                })
                _next_fg = _trans_fg + 1
                for _k in list(st.session_state.keys()):
                    if _k.startswith("trans_") or _k == f"trans_imagenes_muestras_{_trans_fg}":
                        st.session_state.pop(_k, None)
                st.session_state["_trans_fg"]  = _next_fg
                st.session_state.trans_guardado_ok = True
                clear_draft_state()
                st.rerun()

        if st.session_state.trans_guardado_ok:
            st.success("✅ Registro TRANSUIZA guardado en el historial.")
            st.session_state.trans_guardado_ok = False

    elif tipo_servicio == "SEGUIMIENTOS":

        # ── Encabezado ────────────────────────────────────────────────────────
        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                  <span style="font-size:1.35rem;">🔬</span>
                  <span style="font-size:1.35rem;font-weight:700;color:#0056A3;
                               letter-spacing:.5px;font-family:'Segoe UI',sans-serif;">
                    SEGUIMIENTOS
                  </span>
                </div>""",
            unsafe_allow_html=True,
        )

        # ── Pestañas de sub-tipo ─────────────────────────────────────────────
        _seg_sub_vals  = ["ESTACIONES", "ACOMPAÑAMIENTOS", "CONTRAMUESTRAS SOLICITADAS"]
        _seg_tab_icons = {"ESTACIONES": "🏭",
                          "ACOMPAÑAMIENTOS": "👥", "CONTRAMUESTRAS SOLICITADAS": "🧪"}
        _seg_tab_labels = [
            "🏭 ESTACIONES",
            "👥 ACOMPAÑAMIENTOS", "🧪 CONTRAMUESTRAS",
        ]
        _seg_tabs = st.tabs(_seg_tab_labels)

        for _ti, (_tab_ctx, _sub) in enumerate(zip(_seg_tabs, _seg_sub_vals)):
            with _tab_ctx:
                # Contador de generación de formulario — cambia la key de todos los
                # widgets de ACOMPAÑAMIENTOS al guardar, forzando un reset real.
                _fg_seg = st.session_state.get(f"_fg_seg_{_ti}", 0)

                # ── Datos de Identificación ───────────────────────────────
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         📋 Datos de Identificación
                       </div>""",
                    unsafe_allow_html=True,
                )
                with st.container(border=True):
                    if _sub == "ESTACIONES":
                        _cat_df = load_catalogo()
                        _cat_map_cod = dict(zip(_cat_df["codigo"], _cat_df["nombre"]))
                        _cat_map_nom = dict(zip(_cat_df["nombre"], _cat_df["codigo"]))

                        # callbacks de llenado cruzado
                        def _fill_nombre_from_cod(_ti=_ti, _m=_cat_map_cod):
                            cod = st.session_state.get(f"seg_codigo_{_ti}", "").strip()
                            nom = _m.get(cod) or _m.get(cod.upper())
                            if nom:
                                st.session_state[f"seg_nombre_{_ti}"] = nom

                        def _fill_cod_from_nombre(_ti=_ti, _m=_cat_map_nom):
                            raw = st.session_state.get(f"seg_nombre_{_ti}", "")
                            nom = raw.upper()
                            st.session_state[f"seg_nombre_{_ti}"] = nom
                            cod = _m.get(nom.strip())
                            if cod:
                                st.session_state[f"seg_codigo_{_ti}"] = cod

                        sid1, sid2, sid3 = st.columns([1, 1, 2])
                        seg_fecha = sid1.date_input(
                            "📅 FECHA", now_col(),
                            key=f"seg_fecha_{_ti}", format="DD/MM/YYYY",
                        )
                        seg_codigo = sid2.text_input(
                            "🔖 CÓDIGO", placeholder="Ej: 6085",
                            key=f"seg_codigo_{_ti}",
                            on_change=_fill_nombre_from_cod,
                        )
                        st.session_state.setdefault(f"seg_nombre_{_ti}", "")
                        sid3.text_input(
                            "📍 NOMBRE ESTACIÓN", placeholder="Ej: LUCERITO",
                            key=f"seg_nombre_{_ti}",
                            on_change=_fill_cod_from_nombre,
                        )
                        activar_siguiente_con_enter()

                        seg_quien_trajo = ""
                        seg_ruta_acomp  = ""
                        seg_responsable = ""

                    elif _sub == "ACOMPAÑAMIENTOS":
                        sa1, sa2 = st.columns(2)
                        seg_fecha = sa1.date_input(
                            "📅 FECHA", now_col(),
                            key=f"seg_fecha_{_ti}_{_fg_seg}", format="DD/MM/YYYY",
                        )
                        seg_ruta_acomp = sa2.text_input(
                            "📍 NOMBRE DE LA RUTA", placeholder="ESCRIBA AQUÍ...",
                            key=f"seg_ruta_acomp_{_ti}_{_fg_seg}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_ruta_acomp_{_ti}_{_fg_seg}",),
                        )
                        sb1, sb2 = st.columns(2)
                        seg_quien_trajo = sb1.text_input(
                            "👤 ENTREGADO POR", placeholder="NOMBRE COMPLETO",
                            key=f"seg_quien_trajo_{_ti}_{_fg_seg}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_quien_trajo_{_ti}_{_fg_seg}",),
                        )
                        seg_vol_declarado_id = sb2.number_input(
                            "🪣 VOLUMEN (L)", min_value=0, step=1,
                            value=None, placeholder="DIGITE VOLUMEN",
                            key=f"seg_vol_declarado_{_ti}_{_fg_seg}",
                        )
                        seg_codigo      = ""
                        seg_responsable = ""

                    else:  # CONTRAMUESTRAS SOLICITADAS
                        _cat_df_ct   = load_catalogo()
                        _cat_ct_cod  = dict(zip(_cat_df_ct["codigo"], _cat_df_ct["nombre"]))
                        _cat_ct_nom  = dict(zip(_cat_df_ct["nombre"],  _cat_df_ct["codigo"]))

                        def _fill_nom_ct(_ti=_ti, _fg=_fg_seg, _m=_cat_ct_cod):
                            cod = st.session_state.get(f"seg_id_muestra_ct_{_ti}_{_fg}", "").strip()
                            st.session_state[f"seg_id_muestra_ct_{_ti}_{_fg}"] = cod.upper()
                            nom = _m.get(cod) or _m.get(cod.upper())
                            if nom:
                                st.session_state[f"seg_nom_muestra_ct_{_ti}_{_fg}"] = nom

                        def _fill_cod_ct(_ti=_ti, _fg=_fg_seg, _m=_cat_ct_nom):
                            raw = st.session_state.get(f"seg_nom_muestra_ct_{_ti}_{_fg}", "")
                            nom = raw.upper()
                            st.session_state[f"seg_nom_muestra_ct_{_ti}_{_fg}"] = nom
                            cod = _m.get(nom.strip())
                            if cod:
                                st.session_state[f"seg_id_muestra_ct_{_ti}_{_fg}"] = cod

                        sc1, sc2, sc3, sc4 = st.columns([1.5, 1.5, 1, 2])
                        seg_fecha = sc1.date_input(
                            "📅 FECHA DE LAS MUESTRAS", now_col(),
                            key=f"seg_fecha_{_ti}_{_fg_seg}", format="DD/MM/YYYY",
                        )
                        seg_responsable = sc2.text_input(
                            "👤 ENTREGADO POR", placeholder="NOMBRE...",
                            key=f"seg_responsable_{_ti}_{_fg_seg}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_responsable_{_ti}_{_fg_seg}",),
                        )
                        st.session_state.setdefault(f"seg_id_muestra_ct_{_ti}_{_fg_seg}", "")
                        seg_id_muestra_ct = sc3.text_input(
                            "🔖 CÓDIGO", placeholder="Ej: 6085",
                            key=f"seg_id_muestra_ct_{_ti}_{_fg_seg}",
                            on_change=_fill_nom_ct,
                        )
                        st.session_state.setdefault(f"seg_nom_muestra_ct_{_ti}_{_fg_seg}", "")
                        sc4.text_input(
                            "📍 NOMBRE ESTACIÓN", placeholder="Ej: LUCERITO",
                            key=f"seg_nom_muestra_ct_{_ti}_{_fg_seg}",
                            on_change=_fill_cod_ct,
                        )
                        seg_codigo      = ""
                        seg_quien_trajo = ""
                        seg_ruta_acomp  = ""

                activar_siguiente_con_enter()

                # ── ACOMPAÑAMIENTOS: sección análisis justo tras identificación ─
                if _sub == "ACOMPAÑAMIENTOS":
                    st.markdown("---")
                    st.markdown(
                        """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                       margin:10px 0 6px 0;letter-spacing:.4px;
                                       border-left:4px solid #0056A3;padding-left:10px;">
                             🔬 Análisis de Calidad de Ruta
                           </div>""",
                        unsafe_allow_html=True,
                    )
                    _aq1, _aq2 = st.columns(2)
                    seg_solidos_ruta = _aq1.number_input(
                        "SÓLIDOS TOTALES (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None, placeholder="Ej: 12.80",
                        key=f"seg_solidos_ruta_{_ti}_{_fg_seg}",
                    )
                    seg_crios_raw = _aq2.text_input(
                        "CRIOSCOPIA (°C)", value="-0.", placeholder="-0.530",
                        key=f"seg_crios_raw_{_ti}_{_fg_seg}",
                    )
                    try:
                        seg_crioscopia_ruta = float(seg_crios_raw.replace(",", ".")) if seg_crios_raw.strip() else None
                    except Exception:
                        seg_crioscopia_ruta = None
                    st.markdown("---")

                # ── Parámetros de Calidad ─────────────────────────────────
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         🧪 Parámetros de Calidad
                       </div>""",
                    unsafe_allow_html=True,
                )
                _qk = st.session_state.get(f"seg_quality_key_counter_{_ti}", 0)
                with st.container(border=True):
                    seg_id_muestra = ""
                    seg_volumen    = None
                    seg_proveedor  = ""
                    if _sub == "ACOMPAÑAMIENTOS":
                        _cat_df_m  = load_catalogo()
                        _cat_m_cod = dict(zip(_cat_df_m["codigo"], _cat_df_m["nombre"]))
                        _cat_m_nom = dict(zip(_cat_df_m["nombre"], _cat_df_m["codigo"]))

                        def _fill_nom_m(_ti=_ti, _qk=_qk, _m=_cat_m_cod):
                            cod = st.session_state.get(f"seg_id_muestra_{_ti}_{_qk}", "").strip()
                            st.session_state[f"seg_id_muestra_{_ti}_{_qk}"] = cod.upper()
                            nom = _m.get(cod) or _m.get(cod.upper())
                            if nom:
                                st.session_state[f"seg_nom_muestra_{_ti}_{_qk}"] = nom

                        def _fill_cod_m(_ti=_ti, _qk=_qk, _m=_cat_m_nom):
                            raw = st.session_state.get(f"seg_nom_muestra_{_ti}_{_qk}", "")
                            nom = raw.upper()
                            st.session_state[f"seg_nom_muestra_{_ti}_{_qk}"] = nom
                            cod = _m.get(nom.strip())
                            if cod:
                                st.session_state[f"seg_id_muestra_{_ti}_{_qk}"] = cod

                        _idc1, _idc2, _idc3 = st.columns([1, 2, 1])
                        seg_id_muestra = _idc1.text_input(
                            "🔖 CÓDIGO", placeholder="Ej: 6085",
                            key=f"seg_id_muestra_{_ti}_{_qk}",
                            on_change=_fill_nom_m,
                        )
                        st.session_state.setdefault(f"seg_nom_muestra_{_ti}_{_qk}", "")
                        _idc2.text_input(
                            "📍 NOMBRE ESTACIÓN", placeholder="Ej: LUCERITO",
                            key=f"seg_nom_muestra_{_ti}_{_qk}",
                            on_change=_fill_cod_m,
                        )
                        seg_volumen = _idc3.number_input(
                            "VOLUMEN (L)", min_value=0, step=1,
                            value=None, placeholder="0",
                            key=f"seg_volumen_{_ti}_{_qk}",
                        )
                        activar_siguiente_con_enter()
                    elif _sub == "CONTRAMUESTRAS SOLICITADAS":
                        # Código ya capturado en el encabezado — leer de session_state
                        seg_id_muestra = st.session_state.get(f"seg_id_muestra_ct_{_ti}_{_fg_seg}", "").strip()
                        seg_proveedor = st.text_input(
                            "🏭 PROVEEDOR", placeholder="Nombre proveedor...",
                            key=f"seg_proveedor_{_ti}_{_qk}",
                            on_change=convertir_a_mayusculas,
                            args=(f"seg_proveedor_{_ti}_{_qk}",),
                        )
                        activar_siguiente_con_enter()

                    sq1, sq2, sq3, sq4 = st.columns(4)
                    seg_grasa = sq1.number_input(
                        "GRASA (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None,
                        placeholder="0.00", key=f"seg_grasa_{_ti}_{_qk}",
                    )
                    seg_st = sq2.number_input(
                        "ST (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None,
                        placeholder="0.00", key=f"seg_st_{_ti}_{_qk}",
                    )
                    seg_proteina = sq3.number_input(
                        "PROTEÍNA (%)", min_value=0.0, max_value=100.0,
                        step=0.01, format="%.2f", value=None,
                        placeholder="0.00", key=f"seg_proteina_{_ti}_{_qk}",
                    )
                    with sq4:
                        seg_ic_raw = st.text_input(
                            "IC (°C)", key=f"seg_ic_raw_{_ti}_{_qk}",
                            value="-0.", placeholder="-0.530",
                        )
                        try:
                            seg_ic = float(seg_ic_raw.replace(",", ".")) \
                                if seg_ic_raw not in ("", "-", "-0", "-0.") else None
                        except ValueError:
                            seg_ic = None
                            st.warning("⚠️ Ingrese un número válido")
                    _ic_fuera = seg_ic is not None and seg_ic > -0.530
                    _ic_bajo  = seg_ic is not None and seg_ic < -0.550
                    seg_agua = None
                    if _ic_fuera:
                        _aw1, _aw2 = st.columns([1, 2])
                        with _aw1:
                            seg_agua = st.number_input(
                                "💧 AGUA ADICIONADA (%)",
                                min_value=0.0, max_value=100.0,
                                step=0.01, format="%.2f", value=None,
                                placeholder="0.00", key=f"seg_agua_{_ti}_{_qk}",
                            )

                    if seg_st is not None and seg_st > 0 and seg_st < 12.60:
                        st.error("🚨 ALERTA: ST FUERA DE RANGO (MENOR A 12.60%)")
                    elif seg_st is not None and seg_st > 0:
                        st.success("✅ ST dentro del parámetro")

                    if _ic_fuera:
                        st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MAYOR A -0.530)")
                    elif _ic_bajo:
                        st.error("🚨 ALERTA: CRIOSCOPIA FUERA DE RANGO (MENOR A -0.550)")
                    elif seg_ic is not None:
                        st.success("✅ Crioscopia dentro del parámetro")

                    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                    sq5, sq6, sq7 = st.columns(3)
                    opciones_tri = ["N/A", "NEGATIVO (−)", "POSITIVO (+)"]
                    seg_alcohol        = sq5.selectbox("ALCOHOL",        opciones_tri, key=f"seg_alcohol_{_ti}_{_qk}")
                    seg_cloruros       = sq6.selectbox("CLORUROS",       opciones_tri, key=f"seg_cloruros_{_ti}_{_qk}")
                    seg_neutralizantes = sq7.selectbox("NEUTRALIZANTES", opciones_tri, key=f"seg_neutralizantes_{_ti}_{_qk}")

                    _positivos = [p for p, v in [
                        ("ALCOHOL", seg_alcohol), ("CLORUROS", seg_cloruros),
                        ("NEUTRALIZANTES", seg_neutralizantes),
                    ] if v == "POSITIVO (+)"]
                    if _positivos:
                        st.error(f"🚨 ALERTA: {', '.join(_positivos)} POSITIVO(S) — ADULTERACIÓN")

                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                    seg_observaciones = st.text_area(
                        "📝 OBSERVACIONES", placeholder="ESCRIBA AQUÍ...",
                        key=f"seg_observaciones_{_ti}_{_qk}", height=90,
                    )

                activar_siguiente_con_enter()

                # ── ACOMPAÑAMIENTOS: agregar muestra ─────────────────────
                if _sub == "ACOMPAÑAMIENTOS":
                    _acomp_key = "acomp_muestras"
                    if _acomp_key not in st.session_state:
                        st.session_state[_acomp_key] = []
                    if st.button("➕  AGREGAR MUESTRA", width='stretch',
                                 key=f"btn_agregar_muestra_{_ti}"):
                        st.session_state[_acomp_key].append({
                            "ID": seg_id_muestra or "",
                            "VOLUMEN (L)": int(seg_volumen) if seg_volumen is not None else "",
                            "GRASA (%)": f"{seg_grasa:.2f}" if seg_grasa is not None else "",
                            "ST (%)":    f"{seg_st:.2f}"    if seg_st    is not None else "",
                            "IC (°C)":   f"{seg_ic:.3f}"    if seg_ic    is not None else "",
                            "AGUA (%)":  f"{seg_agua:.2f}"  if seg_agua  is not None else "",
                            "ALCOHOL":   seg_alcohol, "CLORUROS": seg_cloruros,
                            "NEUTRALIZANTES": seg_neutralizantes,
                            "OBS":       seg_observaciones or "",
                            "_volumen": seg_volumen,
                            "_grasa": seg_grasa, "_st": seg_st, "_ic": seg_ic,
                            "_proteina": seg_proteina,
                            "_agua": seg_agua, "_alcohol": seg_alcohol,
                            "_cloruros": seg_cloruros, "_neutralizantes": seg_neutralizantes,
                            "_obs": seg_observaciones or "",
                        })
                        st.session_state[f"seg_quality_key_counter_{_ti}"] = _qk + 1
                        st.rerun()
                    if st.session_state[_acomp_key]:
                        st.markdown(
                            f"""<div style="font-size:0.9rem;font-weight:700;color:#0056A3;
                                           margin:10px 0 4px 0;">
                                 📋 {len(st.session_state[_acomp_key])} muestra(s) registrada(s)
                               </div>""",
                            unsafe_allow_html=True,
                        )
                        _cat_pa  = load_catalogo()
                        _cat_pa_map = dict(zip(_cat_pa["codigo"], _cat_pa["nombre"]))
                        _rows_pa = []
                        for m in st.session_state[_acomp_key]:
                            _cod_pa = str(m.get("ID", "") or "").strip()
                            _nom_pa = _cat_pa_map.get(_cod_pa, "")
                            _row_pa = {"ID": _cod_pa, "NOMBRE ESTACIÓN": _nom_pa}
                            _row_pa.update({k: v for k, v in m.items()
                                            if not k.startswith("_") and k != "ID"})
                            _rows_pa.append(_row_pa)
                        df_prev_a = pd.DataFrame(_rows_pa)
                        st.dataframe(df_prev_a, width='stretch', hide_index=True)

                # ── ACOMPAÑAMIENTOS: calcular ponderados para guardado ────
                if _sub == "ACOMPAÑAMIENTOS":
                    seg_vol_declarado = seg_vol_declarado_id
                    _acomp_vols_f = [
                        (lambda v: v if v is not None else None)(
                            (lambda s: float(s.replace(",", "."))
                             if s and str(s).strip() not in ("", "None", "nan") else None
                            )(str(m.get("_volumen", "") or ""))
                        )
                        for m in st.session_state.get(_acomp_key, [])
                    ]
                    _acomp_vol_muestras = int(sum(v for v in _acomp_vols_f if v is not None))
                    _acomp_pond_st, _acomp_pond_ic = [], []
                    for _am in st.session_state.get(_acomp_key, []):
                        def _pn_form(x):
                            try: return float(str(x).replace(",", "."))
                            except: return None
                        _av  = _pn_form(_am.get("_volumen"))
                        _ast = _pn_form(_am.get("_st"))
                        _aic = _pn_form(_am.get("_ic"))
                        _acomp_pond_st.append(_av * _ast if _av is not None and _ast is not None else None)
                        _acomp_pond_ic.append(_av * _aic if _av is not None and _aic is not None else None)
                    _acomp_vol_total_f = sum(v for v in _acomp_vols_f if v is not None)
                    _acomp_st_pond = (round(sum(x for x in _acomp_pond_st if x is not None) / _acomp_vol_total_f, 2)
                                      if _acomp_vol_total_f else None)
                    _acomp_ic_pond = (round(sum(x for x in _acomp_pond_ic if x is not None) / _acomp_vol_total_f, 3)
                                      if _acomp_vol_total_f else None)
                    _acomp_diferencia_vol = (int(seg_vol_declarado) - _acomp_vol_muestras
                                             if seg_vol_declarado is not None else None)

                # ── CONTRAMUESTRAS: agregar muestra ──────────────────────
                if _sub == "CONTRAMUESTRAS SOLICITADAS":
                    _contra_key = "contra_muestras"
                    if _contra_key not in st.session_state:
                        st.session_state[_contra_key] = []
                    if st.button("➕  AGREGAR CONTRAMUESTRA", width='stretch',
                                 key=f"btn_agregar_contra_{_ti}"):
                        st.session_state[_contra_key].append({
                            "ID":        seg_id_muestra or "",
                            "PROVEEDOR": seg_proveedor or "",
                            "GRASA (%)": f"{seg_grasa:.2f}" if seg_grasa is not None else "",
                            "ST (%)":    f"{seg_st:.2f}"    if seg_st    is not None else "",
                            "IC (°C)":   f"{seg_ic:.3f}"    if seg_ic    is not None else "",
                            "AGUA (%)":  f"{seg_agua:.2f}"  if seg_agua  is not None else "",
                            "ALCOHOL":   seg_alcohol, "CLORUROS": seg_cloruros,
                            "NEUTRALIZANTES": seg_neutralizantes,
                            "OBS":       seg_observaciones or "",
                            "_proveedor": seg_proveedor or "",
                            "_grasa": seg_grasa, "_st": seg_st, "_ic": seg_ic,
                            "_proteina": seg_proteina,
                            "_agua": seg_agua, "_alcohol": seg_alcohol,
                            "_cloruros": seg_cloruros, "_neutralizantes": seg_neutralizantes,
                            "_obs": seg_observaciones or "",
                        })
                        st.session_state[f"seg_quality_key_counter_{_ti}"] = _qk + 1
                        st.rerun()
                    if st.session_state[_contra_key]:
                        st.markdown(
                            f"""<div style="font-size:0.9rem;font-weight:700;color:#0056A3;
                                           margin:10px 0 4px 0;">
                                 📋 {len(st.session_state[_contra_key])} contramuestra(s) registrada(s)
                               </div>""",
                            unsafe_allow_html=True,
                        )
                        _cat_cp     = load_catalogo()
                        _cat_cp_map = dict(zip(_cat_cp["codigo"].astype(str), _cat_cp["nombre"]))
                        _rows_cp    = []
                        for _mc in st.session_state[_contra_key]:
                            _cod_cp = str(_mc.get("ID", "") or "").strip()
                            _nom_cp = _cat_cp_map.get(_cod_cp, "")
                            _row_cp = {"ID": _cod_cp, "NOMBRE ESTACIÓN": _nom_cp}
                            _row_cp.update({k: v for k, v in _mc.items()
                                            if not k.startswith("_") and k != "ID"})
                            _rows_cp.append(_row_cp)
                        df_prev_c = pd.DataFrame(_rows_cp)
                        st.dataframe(df_prev_c, width='stretch', hide_index=True)

                # ── Imágenes de Muestras (SEGUIMIENTOS) ──────────────────
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         📷 Imágenes de Muestras
                       </div>""",
                    unsafe_allow_html=True,
                )
                _s_imgs_conf_key   = f"seg_imgs_confirmadas_{_ti}"
                _s_imgs_noms_key   = f"seg_imgs_nombres_{_ti}"
                _s_img_gen_key     = f"_seg_img_gen_{_ti}"
                if _s_imgs_conf_key not in st.session_state:
                    st.session_state[_s_imgs_conf_key] = False
                if _s_imgs_noms_key not in st.session_state:
                    st.session_state[_s_imgs_noms_key] = []
                _s_img_gen = st.session_state.get(_s_img_gen_key, 0)

                _s_imgs_subidas = st.file_uploader(
                    "ADJUNTAR IMÁGENES",
                    type=["png", "jpg", "jpeg"],
                    accept_multiple_files=True,
                    key=f"seg_imgs_uploader_{_ti}_{_s_img_gen}",
                    label_visibility="visible",
                )
                if _s_imgs_subidas:
                    _s_nombres = [f.name for f in _s_imgs_subidas]
                    if _s_nombres != st.session_state[_s_imgs_noms_key]:
                        st.session_state[_s_imgs_conf_key] = False
                    _s_confirmed = st.session_state[_s_imgs_conf_key]
                    _s_thumb = "<div style='display:flex;flex-wrap:wrap;gap:10px;margin:8px 0;'>"
                    for _si in _s_imgs_subidas:
                        _sb64 = base64.b64encode(_si.read()).decode()
                        _sext = _si.name.rsplit(".", 1)[-1].lower()
                        _smime = "image/jpeg" if _sext in ("jpg", "jpeg") else "image/png"
                        _snc   = _si.name if len(_si.name) <= 16 else _si.name[:14] + "…"
                        _schk  = ("<div style='color:#16a34a;font-size:12px;"
                                  "text-align:center;font-weight:600;'>✅ Guardada</div>"
                                  if _s_confirmed else
                                  f"<div style='font-size:10px;color:#888;text-align:center;'>{_snc}</div>")
                        _sbrd  = "#16a34a" if _s_confirmed else "#D1D5DB"
                        _s_thumb += (f"<div style='display:flex;flex-direction:column;"
                                     f"align-items:center;gap:4px;'>"
                                     f"<img src='data:{_smime};base64,{_sb64}' "
                                     f"style='width:150px;height:150px;object-fit:cover;"
                                     f"border-radius:10px;border:2px solid {_sbrd};"
                                     f"box-shadow:0 2px 6px rgba(0,0,0,0.08);'/>"
                                     f"{_schk}</div>")
                        _si.seek(0)
                    _s_thumb += "</div>"
                    st.markdown(_s_thumb, unsafe_allow_html=True)
                    if not st.session_state[_s_imgs_conf_key]:
                        if st.button("💾 GUARDAR IMÁGENES",
                                     key=f"btn_seg_save_imgs_{_ti}",
                                     width='content'):
                            st.session_state[_s_imgs_conf_key] = True
                            st.session_state[_s_imgs_noms_key] = _s_nombres
                            st.rerun()
                    else:
                        st.success("✅ Imágenes guardadas correctamente.")
                else:
                    st.session_state[_s_imgs_conf_key] = False
                    st.caption("No se han adjuntado imágenes.")

                # ── Guardar ───────────────────────────────────────────────
                st.markdown("---")
                st.markdown(
                    """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                                   margin:14px 0 6px 0;letter-spacing:.4px;
                                   border-left:4px solid #0056A3;padding-left:10px;">
                         💾 Guardar en Historial
                       </div>""",
                    unsafe_allow_html=True,
                )
                if st.button(
                    f"💾  GUARDAR {_sub}", type="primary",
                    width='stretch', key=f"btn_guardar_seg_{_ti}",
                ):
                    ts = now_col().strftime("%d/%m/%Y %H:%M")
                    # ── Guardar fotos a disco ──────────────────────────────
                    _s_fotos_prefix = f"SEG_{_sub[:3].upper()}_{(seg_codigo or seg_quien_trajo or seg_responsable or 'X').replace(' ','_')[:12]}"
                    _s_fotos_paths  = save_fotos_to_disk(
                        _s_imgs_subidas or [], _s_fotos_prefix
                    )
                    base = {
                        "tipo_seguimiento":     "SEGUIMIENTOS",
                        "sub_tipo_seguimiento": _sub,
                        "fecha":                seg_fecha.strftime("%d/%m/%Y") if seg_fecha else "",
                        "seg_codigo":           seg_codigo,
                        "seg_quien_trajo":      seg_quien_trajo,
                        "ruta":                 seg_ruta_acomp,
                        "seg_responsable":      seg_responsable,
                        "guardado_en":          ts,
                        "fotos_json":           json.dumps(_s_fotos_paths, ensure_ascii=False),
                    }
                    def _guardar_lista_muestras(lista):
                        for m in lista:
                            save_seguimiento_to_csv({**base,
                                "seg_id_muestra":     m.get("ID", ""),
                                "seg_volumen":        m.get("_volumen", ""),
                                "seg_grasa":          m.get("_grasa", ""),
                                "seg_st":             m.get("_st", ""),
                                "seg_ic":             m.get("_ic", ""),
                                "seg_agua":           m.get("_agua", ""),
                                "seg_alcohol":        m.get("_alcohol", ""),
                                "seg_cloruros":       m.get("_cloruros", ""),
                                "seg_neutralizantes": m.get("_neutralizantes", ""),
                                "seg_observaciones":  m.get("_obs", ""),
                            })
                    if _sub == "ACOMPAÑAMIENTOS" and st.session_state.get("acomp_muestras"):
                        # Recalcular ponderados en el momento de guardar
                        _sv_muestras = st.session_state.acomp_muestras
                        def _pn_sv(x):
                            try: return float(str(x).replace(",", "."))
                            except: return None
                        _sv_vols = [_pn_sv(m.get("_volumen")) for m in _sv_muestras]
                        _sv_vol_tot = sum(v for v in _sv_vols if v is not None)
                        _sv_pond_st = [(_pn_sv(m.get("_volumen")) or 0) * (_pn_sv(m.get("_st")) or 0)
                                       for m in _sv_muestras
                                       if _pn_sv(m.get("_volumen")) is not None and _pn_sv(m.get("_st")) is not None]
                        _sv_pond_ic = [(_pn_sv(m.get("_volumen")) or 0) * (_pn_sv(m.get("_ic")) or 0)
                                       for m in _sv_muestras
                                       if _pn_sv(m.get("_volumen")) is not None and _pn_sv(m.get("_ic")) is not None]
                        _sv_st_pond = round(sum(_sv_pond_st) / _sv_vol_tot, 2) if _sv_vol_tot and _sv_pond_st else ""
                        _sv_ic_pond = round(sum(_sv_pond_ic) / _sv_vol_tot, 3) if _sv_vol_tot and _sv_pond_ic else ""
                        _sv_vol_dec = int(seg_vol_declarado) if seg_vol_declarado is not None else ""
                        _sv_vol_m   = int(_sv_vol_tot) if _sv_vol_tot else ""
                        _sv_dif_vol = (int(seg_vol_declarado) - int(_sv_vol_tot)
                                       if seg_vol_declarado is not None and _sv_vol_tot else "")
                        # Una sola fila con todas las muestras en JSON (igual que RUTAS)
                        save_seguimiento_to_csv({**base,
                            "seg_vol_declarado":   _sv_vol_dec,
                            "seg_vol_muestras":    _sv_vol_m,
                            "seg_diferencia_vol":  _sv_dif_vol,
                            "seg_solidos_ruta":    seg_solidos_ruta if seg_solidos_ruta is not None else "",
                            "seg_crioscopia_ruta": seg_crioscopia_ruta if seg_crioscopia_ruta is not None else "",
                            "seg_st_pond":         _sv_st_pond,
                            "seg_ic_pond":         _sv_ic_pond,
                            "muestras_json":       json.dumps(_sv_muestras, ensure_ascii=False),
                        })
                        st.session_state.acomp_muestras = []
                    elif _sub == "CONTRAMUESTRAS SOLICITADAS" and st.session_state.get("contra_muestras"):
                        _cm_list = st.session_state.contra_muestras
                        save_seguimiento_to_csv({**base,
                            "seg_id_muestra": _cm_list[0].get("ID", "") if _cm_list else "",
                            "muestras_json":  json.dumps(_cm_list, ensure_ascii=False),
                        })
                        st.session_state.contra_muestras = []
                    else:
                        save_seguimiento_to_csv({**base,
                            "seg_id_muestra":    seg_id_muestra or "",
                            "seg_grasa":         seg_grasa if seg_grasa is not None else "",
                            "seg_st":            seg_st    if seg_st    is not None else "",
                            "seg_ic":            seg_ic    if seg_ic    is not None else "",
                            "seg_agua":          seg_agua  if seg_agua  is not None else "",
                            "seg_alcohol":       seg_alcohol,
                            "seg_cloruros":      seg_cloruros,
                            "seg_neutralizantes": seg_neutralizantes,
                            "seg_observaciones": seg_observaciones or "",
                        })
                    # Incrementar contadores de generación → fuerza widgets nuevos (vacios)
                    st.session_state[f"_fg_seg_{_ti}"]                 = _fg_seg + 1
                    st.session_state[f"seg_quality_key_counter_{_ti}"] = _qk + 1
                    st.session_state[_s_img_gen_key]                   = _s_img_gen + 1
                    st.session_state[_s_imgs_conf_key]                 = False
                    st.session_state[_s_imgs_noms_key]                 = []
                    st.session_state[f"seg_guardado_ok_{_ti}"] = True
                    clear_draft_state()
                    st.rerun()

                if st.session_state.get(f"seg_guardado_ok_{_ti}"):
                    st.success(f"✅ Seguimiento {_sub} guardado en el historial.")
                    st.session_state[f"seg_guardado_ok_{_ti}"] = False


    if st.sidebar.button("REINICIAR FORMULARIO"):
        st.session_state.continuar = False
        clear_draft_state()
        st.rerun()

    # ══════════════════════════════════════════════════════════════════
    if tipo_servicio == "ESTACIONES":
        # ── Inicializar estados ────────────────────────────────────────
        if "cat_accion" not in st.session_state:
            st.session_state.cat_accion = None
        if "cat_sel_idx" not in st.session_state:
            st.session_state.cat_sel_idx = None

        st.markdown(
            """<div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
                  <span style="font-size:1.35rem;">🏷️</span>
                  <span style="font-size:1.35rem;font-weight:700;color:#0056A3;
                               letter-spacing:.5px;font-family:'Segoe UI',sans-serif;">
                    CATÁLOGO DE ESTACIONES
                  </span>
                </div>""",
            unsafe_allow_html=True,
        )

        # ── Estados de navegación ──────────────────────────────────────
        if "cat_nav_codigo" not in st.session_state:
            st.session_state.cat_nav_codigo = None

        # ── Filtros ────────────────────────────────────────────────────
        _cat_df = load_catalogo().reset_index(drop=True)
        _asesores_lista = ["Todos"] + sorted(
            _cat_df["asesor"].dropna().unique().tolist()
        )
        _f1, _f2, _cnt_col = st.columns([2.5, 1.8, 0.7])
        with _f1:
            _cat_q = st.text_input(
                "🔍 Código / Nombre",
                key="cat_buscar_input",
                placeholder="Buscar por código o nombre…",
                label_visibility="collapsed",
            ).strip().upper()
        with _f2:
            _cat_asesor = st.selectbox(
                "Asesor",
                options=_asesores_lista,
                index=0,
                key="cat_filtro_asesor",
                label_visibility="collapsed",
            )

        # Aplicar ambos filtros combinados
        _cat_filt = _cat_df.copy()
        if _cat_q:
            _mask_q = (
                _cat_filt["codigo"].str.upper().str.contains(_cat_q, na=False) |
                _cat_filt["nombre"].str.upper().str.contains(_cat_q, na=False)
            )
            _cat_filt = _cat_filt[_mask_q]
        if _cat_asesor != "Todos":
            _cat_filt = _cat_filt[_cat_filt["asesor"] == _cat_asesor]
        _cat_filt = _cat_filt.reset_index(drop=True)
        _n_filt = len(_cat_filt)
        with _cnt_col:
            st.markdown(
                f"<div style='padding:8px 0 0 2px;font-size:13px;color:#6B7280;'>"
                f"{_n_filt} reg.</div>",
                unsafe_allow_html=True,
            )

        # Calcular índice de navegación dentro del filtro
        _nav_cod = st.session_state.cat_nav_codigo
        if _nav_cod and _nav_cod in _cat_filt["codigo"].values:
            _nav_pos = int(_cat_filt[_cat_filt["codigo"] == _nav_cod].index[0])
        else:
            _nav_pos = 0 if _n_filt > 0 else None

        # ── Tabla del catálogo (filtrada, altura dinámica) ─────────────
        _cat_vis = _cat_filt[["codigo", "nombre", "asesor"]].rename(columns={
            "codigo": "CÓDIGO", "nombre": "NOMBRE", "asesor": "ASESOR"
        })
        _tabla_h = max(42 + _n_filt * 35, 70) if _n_filt > 0 else 70
        _tabla_h = min(_tabla_h, 280)
        _cat_sel_widget = st.dataframe(
            _cat_vis,
            width='stretch',
            hide_index=True,
            on_select="rerun",
            selection_mode="single-row",
            height=_tabla_h,
            column_config={
                "CÓDIGO": st.column_config.TextColumn(width="small"),
                "NOMBRE": st.column_config.TextColumn(width="large"),
                "ASESOR": st.column_config.TextColumn(width="medium"),
            },
        )
        # Sincronizar selección por clic en tabla → actualiza navegación
        _cat_filas = (_cat_sel_widget.selection.rows
                      if _cat_sel_widget and hasattr(_cat_sel_widget, "selection")
                      else [])
        if _cat_filas:
            _clicked = _cat_filas[0]
            if _clicked < _n_filt:
                _nav_pos = _clicked
                st.session_state.cat_nav_codigo = _cat_filt.iloc[_nav_pos]["codigo"]
                _nav_cod = st.session_state.cat_nav_codigo

        # ── Barra de navegación + botones ─────────────────────────────
        _nv1, _nv2, _nv3, _nvsep, _nb1, _nb2, _nb3 = st.columns(
            [0.5, 0.5, 2.5, 0.2, 1, 1, 1]
        )
        with _nv1:
            if st.button("◀", key="cat_prev", width='stretch',
                         disabled=(_nav_pos is None or _nav_pos <= 0)):
                _nav_pos = max(0, _nav_pos - 1)
                st.session_state.cat_nav_codigo = _cat_filt.iloc[_nav_pos]["codigo"]
                st.rerun()
        with _nv2:
            if st.button("▶", key="cat_next", width='stretch',
                         disabled=(_nav_pos is None or _nav_pos >= _n_filt - 1)):
                _nav_pos = min(_n_filt - 1, _nav_pos + 1)
                st.session_state.cat_nav_codigo = _cat_filt.iloc[_nav_pos]["codigo"]
                st.rerun()
        with _nv3:
            if _nav_pos is not None and _n_filt > 0:
                _r_info = _cat_filt.iloc[_nav_pos]
                st.markdown(
                    f"<div style='padding:6px 0 0 6px;font-size:13px;color:#0056A3;"
                    f"font-weight:600;white-space:nowrap;overflow:hidden;"
                    f"text-overflow:ellipsis;'>"
                    f"<span style='color:#6B7280;font-weight:400;'>"
                    f"{_nav_pos + 1}/{_n_filt}</span> &nbsp;"
                    f"<strong>{_r_info['codigo']}</strong> — {_r_info['nombre']}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
        with _nb1:
            if st.button("➕ Nueva", key="cat_btn_nueva",
                         width='stretch', type="primary"):
                st.session_state.cat_accion     = "nueva"
                st.session_state.cat_nav_codigo = None
                st.rerun()
        with _nb2:
            if _is_admin():
                if st.button("✏️ Modificar", key="cat_btn_mod",
                             width='stretch',
                             disabled=(_nav_pos is None or _n_filt == 0)):
                    st.session_state.cat_accion     = "modificar"
                    if _nav_pos is not None:
                        st.session_state.cat_nav_codigo = _cat_filt.iloc[_nav_pos]["codigo"]
                    st.rerun()
        with _nb3:
            if _is_admin():
                if st.button("🗑️ Eliminar", key="cat_btn_del",
                             width='stretch',
                             disabled=(_nav_pos is None or _n_filt == 0)):
                    st.session_state.cat_accion     = "eliminar"
                    if _nav_pos is not None:
                        st.session_state.cat_nav_codigo = _cat_filt.iloc[_nav_pos]["codigo"]
                    st.rerun()

        st.markdown("<hr style='border-color:#E5E7EB;margin:8px 0 14px 0;'>",
                    unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════
        # FORMULARIO: NUEVA ESTACIÓN — 3 campos + Guardar
        # ══════════════════════════════════════════════════════════════
        if st.session_state.cat_accion == "nueva":
            st.markdown(
                "<div id='nueva-est-anchor' style='font-size:1rem;font-weight:700;"
                "color:#0056A3;margin:0 0 10px 0;border-left:4px solid #0056A3;"
                "padding-left:10px;'>➕ NUEVA ESTACIÓN</div>",
                unsafe_allow_html=True,
            )
            _fg = st.session_state.get("_cat_nueva_fg", 0)

            # ── Campos (sin st.form para que el selectbox reaccione) ───
            _nc1, _nc2, _nc3 = st.columns(3)
            with _nc1:
                _new_cod = st.text_input(
                    "CÓDIGO *", key=f"nv_cod_{_fg}",
                    placeholder="Ej: 6008"
                ).strip().upper()
            with _nc2:
                _new_nom = st.text_input(
                    "NOMBRE DE LA ESTACIÓN *", key=f"nv_nom_{_fg}",
                    placeholder="Ej: LA JAGUA"
                ).strip().upper()
            with _nc3:
                _ase_opciones = sorted(
                    load_catalogo()["asesor"].str.strip().str.upper()
                    .dropna().unique().tolist()
                ) + ["— OTRO —"]
                _ase_idx_default = next(
                    (i for i, v in enumerate(_ase_opciones)
                     if "JUAN ORTEGA" in v), 0
                )
                _ase_sel = st.selectbox(
                    "ASESOR", options=_ase_opciones,
                    key=f"nv_ase_sel_{_fg}",
                    index=_ase_idx_default,
                )
            # Campo "OTRO" aparece debajo si se elige esa opción
            _ase_otro_visible = (_ase_sel == "— OTRO —")
            if _ase_otro_visible:
                _new_ase = st.text_input(
                    "NOMBRE DEL ASESOR *", key=f"nv_ase_otro_{_fg}",
                    placeholder="Escribe el nombre del asesor…"
                ).strip().upper()
            else:
                _new_ase = _ase_sel

            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            _gb1, _gb2, _ = st.columns([1, 1, 4])
            with _gb1:
                _btn_guardar = st.button(
                    "💾 Guardar", type="primary",
                    width='stretch', key="cat_guardar_nv"
                )
            with _gb2:
                _btn_cancel = st.button(
                    "✖ Cancelar", width='stretch',
                    key="cat_cancelar_nv"
                )

            # JS: Enter/→ entre CÓDIGO y NOMBRE; blur → mayúsculas
            components.html("""<script>
(function tryAttach(n){
  var doc = window.parent.document;
  var lblTexts = ['CÓDIGO *', 'NOMBRE DE LA ESTACIÓN *'];
  var inputs = lblTexts.map(function(txt){
    var cs = Array.from(doc.querySelectorAll('[data-testid="stTextInput"]'));
    var c = cs.find(function(el){
      var lbl = el.querySelector('label, p');
      return lbl && lbl.textContent.trim() === txt;
    });
    return c ? c.querySelector('input') : null;
  });
  // También incluir "NOMBRE DEL ASESOR *" si está visible
  var aOtroC = Array.from(doc.querySelectorAll('[data-testid="stTextInput"]')).find(function(el){
    var lbl=el.querySelector('label, p');
    return lbl && lbl.textContent.trim()==='NOMBRE DEL ASESOR *';
  });
  if(aOtroC) inputs.push(aOtroC.querySelector('input'));
  var missing = inputs.some(function(i){return !i;});
  if(missing && n>0){setTimeout(function(){tryAttach(n-1);},300);return;}
  inputs = inputs.filter(Boolean);
  if(inputs.length===0) return;
  var guardarBtn = Array.from(doc.querySelectorAll('button')).find(
    function(b){return b.textContent.trim().indexOf('Guardar')>=0;}
  );
  function toUpper(inp){
    var nS=Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype,'value');
    nS.set.call(inp,inp.value.toUpperCase());
    inp.dispatchEvent(new Event('input',{bubbles:true}));
  }
  inputs.forEach(function(inp,i){
    if(inp._qlnav) return;
    inp._qlnav = true;
    inp.addEventListener('keydown',function(e){
      if(e.key==='Enter'||(e.key==='ArrowRight'&&inp.selectionStart===inp.value.length)){
        e.preventDefault();
        toUpper(inp);
        if(i<inputs.length-1){
          inputs[i+1].focus();
          inputs[i+1].setSelectionRange(0,inputs[i+1].value.length);
        } else if(guardarBtn){guardarBtn.focus();}
      } else if(e.key==='ArrowLeft'&&inp.selectionStart===0&&i>0){
        e.preventDefault();
        toUpper(inp);
        inputs[i-1].focus();
        var l=inputs[i-1].value.length;
        inputs[i-1].setSelectionRange(l,l);
      }
    });
    inp.addEventListener('blur',function(){toUpper(inp);});
  });
})(25);
</script>""", height=0)

            # ── Lógica ────────────────────────────────────────────────
            if _btn_cancel:
                st.session_state.cat_accion = None
                st.rerun()

            if _btn_guardar:
                if not _new_cod:
                    st.error("El CÓDIGO es obligatorio.")
                elif not _new_nom:
                    st.error("El NOMBRE es obligatorio.")
                elif _ase_otro_visible and not _new_ase:
                    st.error("Escribe el nombre del asesor.")
                elif _new_cod in load_catalogo()["codigo"].values:
                    st.error(f"El código {_new_cod} ya existe en el catálogo.")
                else:
                    _df_save = load_catalogo()
                    _df_save = pd.concat([_df_save, pd.DataFrame([{
                        "codigo": _new_cod, "nombre": _new_nom,
                        "asesor": _new_ase.upper()
                    }])], ignore_index=True)
                    save_catalogo(_df_save)
                    load_catalogo.clear()
                    st.session_state["_cat_nueva_fg"] = _fg + 1
                    st.session_state.cat_accion       = None
                    st.session_state.cat_nav_codigo   = _new_cod
                    st.success(f"✅ Estación {_new_cod} — {_new_nom} agregada.")
                    st.rerun()

        # ══════════════════════════════════════════════════════════════
        # FORMULARIO: MODIFICAR ESTACIÓN
        # ══════════════════════════════════════════════════════════════
        elif st.session_state.cat_accion == "modificar":
            _cod_m = st.session_state.cat_nav_codigo
            _df_m  = load_catalogo().reset_index(drop=True)
            _rows_m = _df_m[_df_m["codigo"] == _cod_m]
            if _cod_m and not _rows_m.empty:
                _idx_m = _rows_m.index[0]
                _row_m = _df_m.iloc[_idx_m]
                st.markdown(
                    "<div style='font-size:1rem;font-weight:700;color:#0056A3;"
                    "margin:0 0 10px 0;border-left:4px solid #F59E0B;"
                    "padding-left:10px;'>✏️ MODIFICAR ESTACIÓN</div>",
                    unsafe_allow_html=True,
                )
                _mc1, _mc2, _mc3 = st.columns(3)
                with _mc1:
                    _mod_cod = st.text_input("CÓDIGO *",
                                             value=_row_m["codigo"],
                                             key="cat_mod_cod").strip().upper()
                with _mc2:
                    _mod_nom = st.text_input("NOMBRE *",
                                             value=_row_m["nombre"],
                                             key="cat_mod_nom").strip().upper()
                with _mc3:
                    _mod_ase = st.text_input("ASESOR",
                                             value=_row_m.get("asesor", ""),
                                             key="cat_mod_ase").strip().upper()

                _ma1, _ma2, _ = st.columns([1, 1, 4])
                with _ma1:
                    if st.button("💾 Guardar cambios", key="cat_guardar_mod",
                                 type="primary", width='stretch'):
                        if not _mod_cod or not _mod_nom:
                            st.error("CÓDIGO y NOMBRE son obligatorios.")
                        else:
                            _df_m.at[_idx_m, "codigo"] = _mod_cod
                            _df_m.at[_idx_m, "nombre"] = _mod_nom
                            _df_m.at[_idx_m, "asesor"] = _mod_ase
                            save_catalogo(_df_m)
                            load_catalogo.clear()
                            st.session_state.cat_accion    = None
                            st.session_state.cat_nav_codigo = _mod_cod
                            st.success(f"✅ Estación actualizada: {_mod_cod} — {_mod_nom}")
                            st.rerun()
                with _ma2:
                    if st.button("✖ Cancelar", key="cat_cancelar_mod",
                                 width='stretch'):
                        st.session_state.cat_accion = None
                        st.rerun()

        # ══════════════════════════════════════════════════════════════
        # CONFIRMAR: ELIMINAR ESTACIÓN
        # ══════════════════════════════════════════════════════════════
        elif st.session_state.cat_accion == "eliminar":
            _cod_e = st.session_state.cat_nav_codigo
            _df_e  = load_catalogo().reset_index(drop=True)
            _rows_e = _df_e[_df_e["codigo"] == _cod_e]
            if _cod_e and not _rows_e.empty:
                _idx_e = _rows_e.index[0]
                _row_e = _df_e.iloc[_idx_e]
                st.warning(
                    f"⚠️ ¿Eliminar la estación **{_row_e['codigo']} — {_row_e['nombre']}**?"
                    f" Esta acción no se puede deshacer."
                )
                _ea1, _ea2, _ = st.columns([1, 1, 4])
                with _ea1:
                    if st.button("🗑️ Confirmar eliminación", key="cat_confirmar_del",
                                 type="primary", width='stretch'):
                        _df_e = _df_e.drop(index=_idx_e).reset_index(drop=True)
                        save_catalogo(_df_e)
                        load_catalogo.clear()
                        st.session_state.cat_accion    = None
                        st.session_state.cat_nav_codigo = None
                        st.success("✅ Estación eliminada del catálogo.")
                        st.rerun()
                with _ea2:
                    if st.button("✖ Cancelar", key="cat_cancelar_del",
                                 width='stretch'):
                        st.session_state.cat_accion = None
                        st.rerun()


elif st.session_state.pagina_activa == "HISTORIAL":
    st.markdown("---")

    # ── HISTORIAL DE RUTAS ──────────────────────────────────────────────────────
    st.markdown(
        """<div style="font-size:1rem;font-weight:700;color:#0056A3;
                       margin:14px 0 6px 0;letter-spacing:.4px;
                       border-left:4px solid #0056A3;padding-left:10px;">
             📊 Historial de Rutas
           </div>""",
        unsafe_allow_html=True,
    )

    df_hist = load_historial()
    if df_hist.empty:
        st.info("No hay rutas guardadas aún. Complete el formulario y presione **GUARDAR RUTA** para registrar datos aquí.")
    else:
        # ── Filtros ───────────────────────────────────────────────────
        st.markdown(
            "<div style='font-weight:600;color:#374151;margin-bottom:8px;'>"
            "🔍 Filtros de búsqueda</div>",
            unsafe_allow_html=True,
        )

        # ── Fila 1: Fechas ────────────────────────────────────────────
        _hoy_col = now_col().date()
        ff1, ff2, _ = st.columns([2, 2, 4])
        with ff1:
            fecha_desde = st.date_input(
                "FECHA DESDE", value=_hoy_col,
                format="DD/MM/YYYY", key="hist_desde",
            )
        with ff2:
            fecha_hasta = st.date_input(
                "FECHA HASTA", value=_hoy_col,
                format="DD/MM/YYYY", key="hist_hasta",
            )

        # ── Fila 2: REGISTRO · SUB-TIPO (solo SEGUIMIENTOS) ──────────
        filtro_subtipo = "TODOS"
        _fr2a, _fr2b, _ = st.columns([2, 2, 2])
        with _fr2a:
            filtro_tipo = st.selectbox(
                "REGISTRO",
                ["TODOS", "RUTAS", "TRANSUIZA", "SEGUIMIENTOS"],
                key="hist_tipo",
            )
        with _fr2b:
            if filtro_tipo == "SEGUIMIENTOS":
                filtro_subtipo = st.selectbox(
                    "📋 SUB-TIPO",
                    ["TODOS", "ESTACIONES", "ACOMPAÑAMIENTOS", "CONTRAMUESTRAS SOLICITADAS"],
                    key="hist_subtipo",
                )

        # ── Fila 3: RUTA · CÓDIGO · PLACA ────────────────────────────
        # RUTA: combina rutas_historial + seguimientos
        _df_seg_src = load_seguimientos()
        _rutas_rutas = (
            df_hist["ruta"].dropna().replace("", pd.NA).dropna().unique().tolist()
            if "ruta" in df_hist.columns else []
        )
        _rutas_segs = (
            _df_seg_src["ruta"].dropna().replace("", pd.NA).dropna().unique().tolist()
            if "ruta" in _df_seg_src.columns else []
        )
        rutas_unicas = ["TODAS"] + sorted(set(_rutas_rutas) | set(_rutas_segs))

        # CÓDIGO: solo seguimientos
        _codigos_todos = (
            ["TODOS"] + sorted(
                _df_seg_src["seg_codigo"].dropna().replace("", pd.NA).dropna().unique().tolist()
            )
            if "seg_codigo" in _df_seg_src.columns else ["TODOS"]
        )

        # PLACA: solo rutas_historial
        placas_unicas = (
            ["TODAS"] + sorted(
                df_hist["placa"].dropna().replace("", pd.NA).dropna().unique().tolist()
            )
            if "placa" in df_hist.columns else ["TODAS"]
        )

        fr1, fr2, fr3 = st.columns([2, 2, 2])
        with fr1:
            filtro_ruta = st.selectbox("📍 RUTA", rutas_unicas, key="hist_ruta")
        with fr2:
            filtro_codigo = st.selectbox("🔖 CÓDIGO", _codigos_todos, key="hist_codigo_seg")
        with fr3:
            filtro_placa = st.selectbox("🚛 PLACA", placas_unicas, key="hist_placa")

        _fbrow, _ = st.columns([1, 5])
        with _fbrow:
            if st.button("🔍 BUSCAR", type="primary",
                         key="btn_buscar_hist", width='stretch'):
                st.session_state.hist_buscar_ok = True
                st.rerun()

        # ── Aplicar filtros ───────────────────────────────────────────
        if not st.session_state.hist_buscar_ok and not st.session_state.get("admin_accion"):
            st.info("Selecciona los filtros y presiona **🔍 BUSCAR** para ver el historial.")
        else:
            # Cuando se elige un CÓDIGO específico → buscar siempre en seguimientos
            # (independiente del REGISTRO seleccionado)
            _buscar_por_codigo = filtro_codigo != "TODOS"

            if _buscar_por_codigo or filtro_tipo in ("TODOS", "SEGUIMIENTOS"):
                # Fuente: seguimientos_historial.csv
                df_filtrado = load_seguimientos()
                if "_fecha_dt" in df_filtrado.columns:
                    df_filtrado = df_filtrado[
                        (df_filtrado["_fecha_dt"].dt.date >= fecha_desde) &
                        (df_filtrado["_fecha_dt"].dt.date <= fecha_hasta)
                    ]
                # Sub-tipo solo si REGISTRO = SEGUIMIENTOS y no hay búsqueda por código
                if not _buscar_por_codigo and filtro_subtipo != "TODOS" \
                        and "sub_tipo_seguimiento" in df_filtrado.columns:
                    df_filtrado = df_filtrado[
                        df_filtrado["sub_tipo_seguimiento"] == filtro_subtipo
                    ]
                if _buscar_por_codigo and "seg_codigo" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["seg_codigo"] == filtro_codigo]
                if filtro_ruta != "TODAS" and "ruta" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["ruta"] == filtro_ruta]
            else:
                df_filtrado = df_hist.copy()
                if "_fecha_dt" in df_filtrado.columns:
                    df_filtrado = df_filtrado[
                        (df_filtrado["_fecha_dt"].dt.date >= fecha_desde) &
                        (df_filtrado["_fecha_dt"].dt.date <= fecha_hasta)
                    ]
                if filtro_tipo != "TODOS" and "tipo_seguimiento" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["tipo_seguimiento"] == filtro_tipo]
                if filtro_placa != "TODAS" and "placa" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["placa"] == filtro_placa]
                if filtro_ruta != "TODAS" and "ruta" in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado["ruta"] == filtro_ruta]

            # ── Estado de calidad ─────────────────────────────────────────
            df_filtrado = df_filtrado.copy()
            df_filtrado["_estado"] = df_filtrado.apply(
                lambda r: calcular_estado_calidad(r.to_dict()), axis=1
            )

            RED = "background-color:#FFC7CE;color:#9C0006;font-weight:700"

            # Columnas reutilizables para secciones de seguimientos
            _SEG_COLS = ["sub_tipo_seguimiento", "fecha", "seg_codigo", "seg_quien_trajo",
                         "ruta", "seg_responsable", "seg_id_muestra",
                         "seg_grasa", "seg_st", "seg_ic", "seg_agua",
                         "seg_alcohol", "seg_cloruros", "seg_neutralizantes",
                         "seg_observaciones", "guardado_en"]
            _SEG_LBLS = {
                "sub_tipo_seguimiento": "SUB-TIPO", "fecha": "FECHA",
                "seg_codigo": "CÓDIGO", "seg_quien_trajo": "ENTREGADO POR",
                "ruta": "RUTA", "seg_responsable": "RESPONSABLE",
                "seg_id_muestra": "ID MUESTRA", "seg_grasa": "GRASA (%)",
                "seg_st": "ST (%)", "seg_ic": "IC (°C)", "seg_agua": "AGUA (%)",
                "seg_alcohol": "ALCOHOL", "seg_cloruros": "CLORUROS",
                "seg_neutralizantes": "NEUTRALIZANTES",
                "seg_observaciones": "OBSERVACIONES", "guardado_en": "GUARDADO EN",
            }
            _SEG_CFG = {
                "GRASA (%)": st.column_config.NumberColumn(format="%.2f"),
                "ST (%)":    st.column_config.NumberColumn(format="%.2f"),
                "IC (°C)":   st.column_config.NumberColumn(format="%.3f"),
                "AGUA (%)":  st.column_config.NumberColumn(format="%.2f"),
            }

            def _render_seccion_cod(titulo, emoji, df_sec):
                """Renderiza una sección con tabla o mensaje vacío."""
                st.markdown(
                    f"<div style='font-size:1rem;font-weight:700;color:#0056A3;"
                    f"margin:14px 0 6px 0;border-left:4px solid #0056A3;"
                    f"padding-left:10px;letter-spacing:.4px;'>"
                    f"{emoji} {titulo} — {len(df_sec)} registro(s)</div>",
                    unsafe_allow_html=True,
                )
                if df_sec.empty:
                    st.info("No hay datos para este código en esta sección.")
                else:
                    _vis = [c for c in _SEG_COLS if c in df_sec.columns]
                    df_d = df_sec[_vis].rename(columns=_SEG_LBLS).reset_index(drop=True)
                    st.dataframe(df_d, width='stretch',
                                 hide_index=True, column_config=_SEG_CFG)

            # ═══════════════════════════════════════════════════════════════
            # MODO CÓDIGO MAESTRO: secciones separadas por tipo
            # ═══════════════════════════════════════════════════════════════
            if _buscar_por_codigo:
                import json as _json_hist

                # Base de seguimientos ya filtrada por fecha + código
                _df_seg_cod = df_filtrado  # cargado arriba con filtro de código

                # Sección ESTACIONES
                _df_est_cod = _df_seg_cod[
                    _df_seg_cod.get("sub_tipo_seguimiento",
                                    pd.Series(dtype=str)) == "ESTACIONES"
                ] if "sub_tipo_seguimiento" in _df_seg_cod.columns else pd.DataFrame()
                # fallback correcto:
                if "sub_tipo_seguimiento" in _df_seg_cod.columns:
                    _df_est_cod   = _df_seg_cod[_df_seg_cod["sub_tipo_seguimiento"] == "ESTACIONES"]
                    _df_acomp_cod = _df_seg_cod[_df_seg_cod["sub_tipo_seguimiento"] == "ACOMPAÑAMIENTOS"]
                    _df_ct_cod    = _df_seg_cod[_df_seg_cod["sub_tipo_seguimiento"] == "CONTRAMUESTRAS SOLICITADAS"]
                else:
                    _df_est_cod = _df_acomp_cod = _df_ct_cod = pd.DataFrame()

                # Sección RUTAS: código aparece en estaciones_json
                _df_rutas_src = df_hist.copy()
                if "_fecha_dt" in _df_rutas_src.columns:
                    _df_rutas_src = _df_rutas_src[
                        (_df_rutas_src["_fecha_dt"].dt.date >= fecha_desde) &
                        (_df_rutas_src["_fecha_dt"].dt.date <= fecha_hasta)
                    ]
                _idxs_ruta_cod = []
                for _ri, _rrow in _df_rutas_src.iterrows():
                    try:
                        _ests = _json_hist.loads(
                            _rrow.get("estaciones_json", "[]") or "[]"
                        )
                        if any(
                            str(_e.get("codigo", "")).strip().upper()
                            == str(filtro_codigo).strip().upper()
                            for _e in _ests
                        ):
                            _idxs_ruta_cod.append(_ri)
                    except Exception:
                        pass
                _df_rutas_cod = _df_rutas_src.loc[_idxs_ruta_cod]

                _total_cod = len(_df_est_cod) + len(_df_acomp_cod) + len(_df_ct_cod) + len(_df_rutas_cod)
                st.markdown(
                    f"<div style='font-weight:700;color:#374151;margin:4px 0 10px 0;'>"
                    f"🔍 Resultados para código <span style='color:#0056A3;'>{filtro_codigo}</span>"
                    f" — {_total_cod} registro(s) en total</div>",
                    unsafe_allow_html=True,
                )

                _render_seccion_cod("ESTACIONES",              "🏗️", _df_est_cod)
                _render_seccion_cod("ACOMPAÑAMIENTOS",         "🚌", _df_acomp_cod)
                _render_seccion_cod("CONTRAMUESTRAS SOLICITADAS", "🧪", _df_ct_cod)

                # RUTAS: columnas propias
                st.markdown(
                    "<div style='font-size:1rem;font-weight:700;color:#0056A3;"
                    "margin:14px 0 6px 0;border-left:4px solid #0056A3;"
                    "padding-left:10px;letter-spacing:.4px;'>"
                    f"🚛 RUTAS — {len(_df_rutas_cod)} registro(s)</div>",
                    unsafe_allow_html=True,
                )
                if _df_rutas_cod.empty:
                    st.info("No hay datos para este código en esta sección.")
                else:
                    _rcols = ["fecha", "ruta", "placa", "conductor",
                              "volumen_declarado", "vol_estaciones",
                              "solidos_ruta", "crioscopia_ruta", "guardado_en"]
                    _rlbls = {
                        "fecha": "FECHA", "ruta": "RUTA", "placa": "PLACA",
                        "conductor": "CONDUCTOR",
                        "volumen_declarado": "VOL. DECL. (L)",
                        "vol_estaciones": "VOL. EST. (L)",
                        "solidos_ruta": "ST RUTA (%)",
                        "crioscopia_ruta": "IC RUTA (°C)",
                        "guardado_en": "GUARDADO EN",
                    }
                    _rcfg = {
                        "ST RUTA (%)":  st.column_config.NumberColumn(format="%.2f"),
                        "IC RUTA (°C)": st.column_config.NumberColumn(format="%.3f"),
                        "VOL. DECL. (L)": st.column_config.NumberColumn(format="%d"),
                        "VOL. EST. (L)":  st.column_config.NumberColumn(format="%d"),
                    }
                    _rvis = [c for c in _rcols if c in _df_rutas_cod.columns]
                    st.dataframe(
                        _df_rutas_cod[_rvis].rename(columns=_rlbls).reset_index(drop=True),
                        width='stretch', hide_index=True, column_config=_rcfg,
                    )

                # Excel con todo lo encontrado para este código
                if _total_cod > 0:
                    _ts = now_col().strftime('%Y%m%d_%H%M')
                    _cx, _ = st.columns([1, 3])
                    with _cx:
                        st.download_button(
                            label="⬇️ DESCARGAR REPORTE EXCEL",
                            data=historial_to_excel_filtrado(
                                df_filtrado, fecha_desde, fecha_hasta,
                                filtro_tipo, filtro_subtipo
                            ),
                            file_name=f"historial_qualilact_{_ts}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary", width='stretch',
                        )

                # Sin panel de acciones en modo multi-sección
                sel = None
                sel_orig_idxs = []
                sel_orig_idx  = None

            # ═══════════════════════════════════════════════════════════════
            # MODO NORMAL: tabla única con acciones
            # ═══════════════════════════════════════════════════════════════
            else:
                n_desv = (df_filtrado["_estado"] == "DESVIACIÓN").sum()
                col_info, col_alerta = st.columns([3, 1])
                with col_info:
                    st.markdown(f"**{len(df_filtrado)} registro(s) encontrado(s)**")
                with col_alerta:
                    if n_desv:
                        st.markdown(
                            f"<div style='text-align:right;font-size:13px;"
                            f"color:#9C0006;font-weight:600;'>"
                            f"⚠️ {n_desv} desviación(es)</div>",
                            unsafe_allow_html=True,
                        )

            # ── Columnas y styler según tipo de seguimiento ───────────────
            if not _buscar_por_codigo and filtro_tipo == "TRANSUIZA" and not _buscar_por_codigo:
                # Columnas exclusivas TRANSUIZA
                cols_sel = ["tipo_seguimiento", "fecha", "placa",
                            "st_carrotanque", "grasa_muestra", "solidos_ruta",
                            "proteina_muestra", "diferencia_solidos", "guardado_en"]
                col_labels = {
                    "tipo_seguimiento": "TIPO", "fecha": "FECHA", "placa": "PLACA",
                    "st_carrotanque": "ST CARROTANQUE (%)",
                    "grasa_muestra": "GRASA (%)", "solidos_ruta": "ST MUESTRA (%)",
                    "proteina_muestra": "PROTEÍNA (%)",
                    "diferencia_solidos": "DIF. SÓLIDOS", "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "ST CARROTANQUE (%)": st.column_config.NumberColumn(format="%.2f"),
                    "GRASA (%)":          st.column_config.NumberColumn(format="%.2f"),
                    "ST MUESTRA (%)":     st.column_config.NumberColumn(format="%.2f"),
                    "PROTEÍNA (%)":       st.column_config.NumberColumn(format="%.2f"),
                    "DIF. SÓLIDOS":       st.column_config.NumberColumn(format="%.2f"),
                }
                def resaltar_celdas(row):
                    return [""] * len(row)

            elif filtro_tipo == "RUTAS" and not _buscar_por_codigo:
                # Columnas completas RUTAS
                cols_sel = ["tipo_seguimiento", "fecha", "ruta", "placa", "conductor",
                            "volumen_declarado", "vol_estaciones", "diferencia",
                            "solidos_ruta", "crioscopia_ruta", "st_pond", "ic_pond",
                            "num_estaciones", "guardado_en"]
                col_labels = {
                    "tipo_seguimiento": "TIPO", "fecha": "FECHA", "ruta": "RUTA",
                    "placa": "PLACA", "conductor": "CONDUCTOR",
                    "volumen_declarado": "VOL. DECL. (L)", "vol_estaciones": "VOL. EST. (L)",
                    "diferencia": "DIFER. (L)", "solidos_ruta": "ST RUTA (%)",
                    "crioscopia_ruta": "IC RUTA (°C)", "st_pond": "ST POND",
                    "ic_pond": "IC POND", "num_estaciones": "Nº EST.",
                    "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "ST RUTA (%)":      st.column_config.NumberColumn(format="%.2f"),
                    "ST POND":          st.column_config.NumberColumn(format="%.2f"),
                    "IC RUTA (°C)":     st.column_config.NumberColumn(format="%.3f"),
                    "IC POND":          st.column_config.NumberColumn(format="%.3f"),
                    "VOL. DECL. (L)":   st.column_config.NumberColumn(format="%d"),
                    "VOL. EST. (L)":    st.column_config.NumberColumn(format="%d"),
                    "DIFER. (L)":       st.column_config.NumberColumn(format="%d"),
                    "Nº EST.":          st.column_config.NumberColumn(format="%d"),
                }
                def resaltar_celdas(row):
                    styles = [""] * len(row)
                    cols = list(row.index)
                    desv_st = desv_ic = False
                    try:
                        v = float(str(row.get("ST RUTA (%)", "")).replace(",", "."))
                        if 0 < v < 12.60: desv_st = True
                    except Exception: pass
                    try:
                        v = float(str(row.get("IC RUTA (°C)", "")).replace(",", "."))
                        if v > -0.535 or v < -0.550: desv_ic = True
                    except Exception: pass
                    if (desv_st or desv_ic) and "RUTA" in cols:
                        styles[cols.index("RUTA")] = RED
                    if desv_st and "ST RUTA (%)" in cols:
                        styles[cols.index("ST RUTA (%)")] = RED
                    if desv_ic and "IC RUTA (°C)" in cols:
                        styles[cols.index("IC RUTA (°C)")] = RED
                    return styles

            elif filtro_tipo == "SEGUIMIENTOS":
                cols_sel = ["sub_tipo_seguimiento", "fecha", "seg_codigo",
                            "seg_quien_trajo", "ruta",
                            "seg_id_muestra", "seg_grasa", "seg_st", "seg_ic", "seg_agua",
                            "seg_alcohol", "seg_cloruros", "seg_neutralizantes",
                            "seg_observaciones", "guardado_en"]
                col_labels = {
                    "sub_tipo_seguimiento": "SUB-TIPO", "fecha": "FECHA",
                    "seg_codigo": "CÓDIGO", "seg_quien_trajo": "ENTREGADO POR",
                    "ruta": "RUTA",
                    "seg_id_muestra": "ID MUESTRA",
                    "seg_grasa": "GRASA (%)", "seg_st": "ST (%)",
                    "seg_ic": "IC (°C)", "seg_agua": "AGUA (%)",
                    "seg_alcohol": "ALCOHOL", "seg_cloruros": "CLORUROS",
                    "seg_neutralizantes": "NEUTRALIZANTES",
                    "seg_observaciones": "OBSERVACIONES", "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "GRASA (%)": st.column_config.NumberColumn(format="%.2f"),
                    "ST (%)":    st.column_config.NumberColumn(format="%.2f"),
                    "IC (°C)":   st.column_config.NumberColumn(format="%.3f"),
                    "AGUA (%)":  st.column_config.NumberColumn(format="%.2f"),
                }
                def resaltar_celdas(row):
                    return [""] * len(row)

            else:
                # TODOS: mismas columnas que SEGUIMIENTOS
                cols_sel = ["sub_tipo_seguimiento", "fecha", "seg_codigo",
                            "seg_quien_trajo", "ruta",
                            "seg_id_muestra", "seg_grasa", "seg_st", "seg_ic", "seg_agua",
                            "seg_alcohol", "seg_cloruros", "seg_neutralizantes",
                            "seg_observaciones", "guardado_en"]
                col_labels = {
                    "sub_tipo_seguimiento": "SUB-TIPO", "fecha": "FECHA",
                    "seg_codigo": "CÓDIGO", "seg_quien_trajo": "ENTREGADO POR",
                    "ruta": "RUTA",
                    "seg_id_muestra": "ID MUESTRA",
                    "seg_grasa": "GRASA (%)", "seg_st": "ST (%)",
                    "seg_ic": "IC (°C)", "seg_agua": "AGUA (%)",
                    "seg_alcohol": "ALCOHOL", "seg_cloruros": "CLORUROS",
                    "seg_neutralizantes": "NEUTRALIZANTES",
                    "seg_observaciones": "OBSERVACIONES", "guardado_en": "GUARDADO EN",
                }
                col_config_map = {
                    "GRASA (%)": st.column_config.NumberColumn(format="%.2f"),
                    "ST (%)":    st.column_config.NumberColumn(format="%.2f"),
                    "IC (°C)":   st.column_config.NumberColumn(format="%.3f"),
                    "AGUA (%)":  st.column_config.NumberColumn(format="%.2f"),
                }
                def resaltar_celdas(row):
                    return [""] * len(row)

            # ── Tabla única + acciones (solo cuando NO hay filtro por código) ──
            if not _buscar_por_codigo:
                # Filtrar solo columnas que existen
                cols_data = [c for c in CSV_COLS if c in df_filtrado.columns]
                cols_vis   = [c for c in cols_sel if c in df_filtrado.columns]
                df_display = df_filtrado[cols_vis].rename(columns=col_labels).reset_index(drop=True)

                sel = st.dataframe(
                    df_display.style.apply(resaltar_celdas, axis=1),
                    width='stretch',
                    hide_index=True,
                    on_select="rerun",
                    selection_mode="multi-row",
                    column_config=col_config_map,
                )

                # ── Calcular selección ANTES del botón de descarga ────────────
                orig_indices  = df_filtrado.index.tolist()
                filas_sel     = (sel.selection.rows
                                 if sel and hasattr(sel, "selection") else [])
                sel_orig_idxs = [orig_indices[i] for i in filas_sel
                                 if i < len(orig_indices)]
                sel_orig_idx  = sel_orig_idxs[0] if len(sel_orig_idxs) == 1 else None

                # Si hay filas seleccionadas → exportar solo esas;
                # si no → exportar todas las filtradas
                df_para_excel = (df_filtrado.loc[sel_orig_idxs]
                                 if sel_orig_idxs else df_filtrado)

                # ── Descarga Excel (filas seleccionadas o todas) ──────────────
                if not df_filtrado.empty:
                    _ts = now_col().strftime('%Y%m%d_%H%M')
                    _n  = len(df_para_excel)
                    _msg = (f"📋 {_n} fila(s) seleccionada(s)"
                            if sel_orig_idxs
                            else f"📋 {_n} registro(s) filtrados")
                    _cx, _mx, _ = st.columns([1, 2, 3])
                    with _cx:
                        st.download_button(
                            label="⬇️ DESCARGAR",
                            data=historial_to_excel_filtrado(
                                df_para_excel, fecha_desde, fecha_hasta,
                                filtro_tipo, filtro_subtipo
                            ),
                            file_name=f"historial_qualilact_{_ts}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            width='stretch',
                        )
                    with _mx:
                        st.markdown(
                            f"<div style='padding:8px 0 0 4px;font-size:13px;"
                            f"color:#374151;'>{_msg}</div>",
                            unsafe_allow_html=True,
                        )

            if sel_orig_idxs:
                n_sel = len(sel_orig_idxs)
                _accion_hint = " — elige una acción:" if _is_admin() else " — seleccionada para ver detalle:"
                st.markdown(
                    f"<div style='font-size:12px;color:#6B7280;margin:6px 0 4px 0;'>"
                    f"{'1 fila seleccionada' if n_sel == 1 else f'{n_sel} filas seleccionadas'}"
                    f"{_accion_hint}</div>",
                    unsafe_allow_html=True,
                )
                if _is_admin():
                    ab1, ab2, _ = st.columns([1, 1, 5])
                    with ab1:
                        if st.button("✏️ Modificar", key="btn_modificar",
                                     width='stretch',
                                     help="Editar este registro",
                                     disabled=(n_sel != 1)):
                            st.session_state.admin_accion = "modificar"
                            st.session_state.admin_idx    = sel_orig_idx
                            st.rerun()
                    with ab2:
                        if st.button("🗑️ Eliminar", key="btn_eliminar",
                                     width='stretch',
                                     help=f"Eliminar {n_sel} registro(s) seleccionado(s)"):
                            st.session_state.admin_accion    = "eliminar"
                            st.session_state.admin_idxs      = sel_orig_idxs
                            st.session_state.admin_idx       = sel_orig_idx
                            st.session_state.admin_from_seg  = filtro_tipo in ("SEGUIMIENTOS", "TODOS")
                            st.rerun()

            # ── Detalle de ruta al seleccionar una fila ────────────────────
            # Si estamos en modo edición y la selección se perdió en el rerun,
            # recuperar el índice desde la sesión para mantener el panel abierto.
            _ss_accion = st.session_state.get("admin_accion")
            _ss_idx    = st.session_state.get("admin_idx")
            if (sel_orig_idx is None and _ss_accion == "modificar"
                    and _ss_idx is not None):
                sel_orig_idx = _ss_idx

            # ── Panel de detalle ACOMPAÑAMIENTOS ───────────────────────────
            if (sel_orig_idx is not None and filtro_tipo == "SEGUIMIENTOS"
                    and st.session_state.hist_buscar_ok):
                if sel_orig_idx in df_filtrado.index:
                    _srow = df_filtrado.loc[sel_orig_idx]
                else:
                    _srow = {}
                _s_sub = str(_srow.get("sub_tipo_seguimiento", "")).strip()
                if _s_sub == "ACOMPAÑAMIENTOS":
                    _mj_raw = str(_srow.get("muestras_json", "") or "").strip()
                    try:
                        _mj_data = json.loads(_mj_raw) if _mj_raw else []
                    except Exception:
                        _mj_data = []
                    if _mj_data:
                        st.markdown("---")
                        _ac_entreg  = str(_srow.get("seg_quien_trajo", "") or "").strip() or "—"
                        _ac_ruta    = str(_srow.get("ruta", "") or "").strip() or "—"
                        _ac_resp    = str(_srow.get("seg_responsable", "") or "").strip() or "—"
                        _ac_cod     = str(_srow.get("seg_codigo", "") or "").strip() or "—"
                        _ac_fecha   = str(_srow.get("fecha", "") or "").strip() or "—"
                        st.markdown(
                            f"""<div style="background:#0056A3;border-radius:8px;
                                           padding:10px 16px;margin-bottom:14px;">
                                  <span style="font-size:1rem;font-weight:700;color:#fff;
                                               letter-spacing:.05em;">
                                    👥 DETALLE ACOMPAÑAMIENTO
                                  </span>
                                  <span style="font-size:.85rem;color:#cce0f5;margin-left:12px;">
                                    {_ac_cod} &nbsp;·&nbsp; {_ac_fecha}
                                  </span>
                                </div>""",
                            unsafe_allow_html=True,
                        )

                        # ── Detección de modo edición ──────────────────────────────
                        _det_accion_ac = st.session_state.get("admin_accion")
                        _det_idx_ac    = st.session_state.get("admin_idx")
                        _edit_mode_ac  = (
                            _det_accion_ac == "modificar"
                            and _det_idx_ac == sel_orig_idx
                        )

                        if _edit_mode_ac:
                            # ── MODO EDICIÓN ACOMPAÑAMIENTOS ───────────────────────
                            def _pn_ed(x, default=0.0):
                                try: return float(str(x).replace(",","."))
                                except: return default

                            try:
                                _fe_ac = datetime.strptime(
                                    str(_srow.get("fecha", "")), "%d/%m/%Y"
                                ).date()
                            except Exception:
                                _fe_ac = date.today()
                            try:
                                _ac_vol_d_v = int(_pn_ed(
                                    _srow.get("seg_vol_declarado", "") or 0
                                ))
                            except Exception:
                                _ac_vol_d_v = 0

                            # ── Encabezado ─────────────────────────────────────────
                            st.markdown(
                                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                "letter-spacing:.05em;margin-bottom:6px;'>📋 ENCABEZADO</div>",
                                unsafe_allow_html=True,
                            )
                            aci1, aci2 = st.columns(2)
                            _ace_fecha = aci1.date_input(
                                "📅 FECHA", value=_fe_ac,
                                format="DD/MM/YYYY", key="edit_ac_fecha",
                            )
                            _ace_ruta = aci2.text_input(
                                "📍 NOMBRE DE LA RUTA",
                                value=_ac_ruta if _ac_ruta != "—" else "",
                                key="edit_ac_ruta",
                                on_change=convertir_a_mayusculas,
                                args=("edit_ac_ruta",),
                            )
                            acj1, acj2, acj3, acj4 = st.columns(4)
                            _ace_ent = acj1.text_input(
                                "👤 ENTREGADO POR",
                                value=_ac_entreg if _ac_entreg != "—" else "",
                                key="edit_ac_ent",
                                on_change=convertir_a_mayusculas,
                                args=("edit_ac_ent",),
                            )
                            _ace_vol = acj2.number_input(
                                "🪣 VOL. DECLARADO (L)",
                                value=_ac_vol_d_v, min_value=0, step=1,
                                key="edit_ac_vol",
                            )
                            _ace_str = acj3.number_input(
                                "📊 ST RUTA (%)",
                                value=_pn_ed(_srow.get("seg_solidos_ruta", "")),
                                min_value=0.0, max_value=100.0,
                                step=0.01, format="%.2f", key="edit_ac_str",
                            )
                            _ace_icr = acj4.number_input(
                                "🌡 IC RUTA (°C)",
                                value=_pn_ed(_srow.get("seg_crioscopia_ruta", ""), -0.530),
                                step=0.001, format="%.3f", key="edit_ac_icr",
                            )

                            # ── Muestras por estación (data_editor) ────────────────
                            st.markdown(
                                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                "letter-spacing:.05em;margin:10px 0 4px;'>🔬 ESTACIONES</div>",
                                unsafe_allow_html=True,
                            )
                            st.caption("💡 Tab / → / Enter → siguiente celda · ← → anterior · Enter en última celda → guarda")

                            _AC_COLS = ["ID","_volumen","_grasa","_st","_proteina",
                                        "_ic","_agua","_alcohol","_cloruros",
                                        "_neutralizantes","_obs"]
                            _ac_cache_key = f"_ac_m_cache_{sel_orig_idx}"
                            _cat_ac_ed    = load_catalogo()
                            _cat_ac_ed_m  = dict(zip(_cat_ac_ed["codigo"], _cat_ac_ed["nombre"]))

                            if _ac_cache_key in st.session_state:
                                _df_ac_ed = st.session_state[_ac_cache_key].copy()
                            else:
                                _df_ac_ed = pd.DataFrame(
                                    [{k: m.get(k, "") for k in _AC_COLS} for m in _mj_data],
                                    columns=_AC_COLS,
                                )
                                for _nc in ["_volumen"]:
                                    _df_ac_ed[_nc] = pd.to_numeric(_df_ac_ed[_nc], errors="coerce")
                                for _nc in ["_grasa","_st","_proteina","_agua","_ic"]:
                                    _df_ac_ed[_nc] = pd.to_numeric(_df_ac_ed[_nc], errors="coerce")

                            _df_ac_ed["nombre_estacion"] = _df_ac_ed["ID"].apply(
                                lambda c: _cat_ac_ed_m.get(str(c).strip(), "") if pd.notna(c) else ""
                            )
                            _tri_ac_de = ["N/A", "NEGATIVO (−)", "POSITIVO (+)"]
                            _edited_ac = st.data_editor(
                                _df_ac_ed, num_rows="dynamic", width='stretch',
                                key="edit_ac_de",
                                column_config={
                                    "ID":              st.column_config.TextColumn("CÓDIGO"),
                                    "nombre_estacion": st.column_config.TextColumn("NOMBRE", disabled=True),
                                    "_volumen":        st.column_config.NumberColumn("VOL (L)",    format="%.0f", min_value=0, step=1),
                                    "_grasa":          st.column_config.NumberColumn("GRASA (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                    "_st":             st.column_config.NumberColumn("ST (%)",     format="%.2f", min_value=0.0, max_value=100.0),
                                    "_proteina":       st.column_config.NumberColumn("PROT. (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                    "_ic":             st.column_config.NumberColumn("IC (°C)",    format="%.3f"),
                                    "_agua":           st.column_config.NumberColumn("AGUA (%)",   format="%.2f", min_value=0.0, max_value=100.0),
                                    "_alcohol":        st.column_config.SelectboxColumn("ALCOHOL",        options=_tri_ac_de, required=True),
                                    "_cloruros":       st.column_config.SelectboxColumn("CLORUROS",       options=_tri_ac_de, required=True),
                                    "_neutralizantes": st.column_config.SelectboxColumn("NEUT.",           options=_tri_ac_de, required=True),
                                    "_obs":            st.column_config.TextColumn("OBS"),
                                },
                                hide_index=True,
                            )
                            # Caché + refresh si cambió código
                            _ac_prev_ids = list(_df_ac_ed["ID"].fillna("").astype(str))
                            _ac_new_ids  = list(_edited_ac["ID"].fillna("").astype(str))
                            _ac_cache_df = _edited_ac.drop(columns=["nombre_estacion"], errors="ignore").copy()
                            st.session_state[_ac_cache_key] = _ac_cache_df
                            if _ac_prev_ids != _ac_new_ids:
                                st.rerun()

                            activar_siguiente_con_enter()
                            ac_s1, ac_s2, _ = st.columns([1.5, 1, 3])
                            with ac_s1:
                                if st.button(
                                    "💾 GUARDAR CAMBIOS", type="primary",
                                    width='stretch', key="btn_save_edit_ac",
                                ):
                                    # Reconstruir muestras_json desde data_editor
                                    def _pf(v, d=0.0):
                                        try:
                                            f = float(str(v).replace(",","."))
                                            return f if f == f else d  # descarta NaN
                                        except: return d
                                    def _pi(v, d=0):
                                        try:
                                            f = float(str(v).replace(",","."))
                                            return int(f) if f == f else d
                                        except: return d
                                    _save_df = _edited_ac.drop(columns=["nombre_estacion"], errors="ignore")
                                    _new_muestras = []
                                    for _, _row in _save_df.iterrows():
                                        _nid  = str(_row.get("ID","") or "").upper().strip()
                                        _nvol = _pi(_row.get("_volumen", 0))
                                        _ngr  = round(_pf(_row.get("_grasa",0)), 2)
                                        _nst  = round(_pf(_row.get("_st",0)), 2)
                                        _nic = round(_pf(_row.get("_ic", -0.530), -0.530), 3)
                                        _nag  = round(_pf(_row.get("_agua",0)), 2)
                                        _nprot= round(_pf(_row.get("_proteina",0)), 2)
                                        _nalc = str(_row.get("_alcohol","N/A") or "N/A")
                                        _nclor= str(_row.get("_cloruros","N/A") or "N/A")
                                        _nneut= str(_row.get("_neutralizantes","N/A") or "N/A")
                                        _nobs = str(_row.get("_obs","") or "")
                                        _new_muestras.append({
                                            "ID": _nid, "_volumen": _nvol,
                                            "_grasa": _ngr, "_st": _nst,
                                            "_ic": _nic, "_agua": _nag,
                                            "_proteina": _nprot,
                                            "_alcohol": _nalc, "_cloruros": _nclor,
                                            "_neutralizantes": _nneut, "_obs": _nobs,
                                        })
                                    # Recalcular ponderados
                                    _nv_tot   = sum(m["_volumen"] for m in _new_muestras)
                                    _npond_st = [m["_volumen"]*m["_st"] for m in _new_muestras
                                                 if m["_volumen"] and m["_st"] is not None]
                                    _npond_ic = [m["_volumen"]*m["_ic"] for m in _new_muestras
                                                 if m["_volumen"] and m["_ic"] is not None]
                                    _nst_p = round(sum(_npond_st)/_nv_tot,2) if _nv_tot and _npond_st else ""
                                    _nic_p = round(sum(_npond_ic)/_nv_tot,3) if _nv_tot and _npond_ic else ""
                                    _n_vol_dec = _pi(st.session_state.get("edit_ac_vol", _ac_vol_d_v) or _ac_vol_d_v)
                                    update_seg_row_in_csv(sel_orig_idx, {
                                        "fecha":              _ace_fecha.strftime("%d/%m/%Y"),
                                        "ruta":               str(st.session_state.get("edit_ac_ruta","")).upper().strip(),
                                        "seg_quien_trajo":    str(st.session_state.get("edit_ac_ent","")).upper().strip(),
                                        "seg_vol_declarado":  _n_vol_dec,
                                        "seg_vol_muestras":   _pi(_nv_tot) if _nv_tot else "",
                                        "seg_diferencia_vol": _pi(_n_vol_dec - _nv_tot) if _nv_tot else "",
                                        "seg_solidos_ruta":   round(_pf(st.session_state.get("edit_ac_str",0)), 2),
                                        "seg_crioscopia_ruta":round(_pf(st.session_state.get("edit_ac_icr",-0.530),-0.530), 3),
                                        "seg_st_pond":        _nst_p,
                                        "seg_ic_pond":        _nic_p,
                                        "muestras_json":      json.dumps(_new_muestras, ensure_ascii=False),
                                    })
                                    st.session_state.pop(_ac_cache_key, None)
                                    st.session_state.admin_accion = None
                                    st.session_state.admin_idx    = None
                                    st.rerun()
                            with ac_s2:
                                if st.button(
                                    "✖ CANCELAR", width='stretch',
                                    key="btn_cancel_edit_ac",
                                ):
                                    st.session_state.pop(_ac_cache_key, None)
                                    st.session_state.admin_accion = None
                                    st.session_state.admin_idx    = None
                                    st.rerun()

                        else:
                            # ── MODO VISTA ACOMPAÑAMIENTOS ─────────────────────────
                            def _kpi_card_ac(label, value, badge=None, badge_ok=True):
                                badge_html = ""
                                if badge:
                                    bg  = "#D4EDDA" if badge_ok else "#F8D7DA"
                                    col = "#155724" if badge_ok else "#721C24"
                                    badge_html = (
                                        f'<div style="margin-top:4px;font-size:.65rem;font-weight:700;'
                                        f'color:{col};background:{bg};border-radius:4px;'
                                        f'padding:1px 5px;display:inline-block;">{badge}</div>'
                                    )
                                return (
                                    f'<div style="background:#fff;border:1px solid #dde6f0;'
                                    f'border-radius:8px;padding:10px 12px;text-align:center;height:100%;">'
                                    f'<div style="font-size:.62rem;font-weight:700;color:#6c8ca8;'
                                    f'letter-spacing:.06em;margin-bottom:4px;">{label}</div>'
                                    f'<div style="font-size:1.05rem;font-weight:800;color:#0056A3;">{value}</div>'
                                    f'{badge_html}</div>'
                                )

                            def _pnac(x):
                                try: return float(str(x).replace(",", "."))
                                except: return None

                            _ac_n_muestras  = len(_mj_data)
                            _entreg_short   = (_ac_entreg[:16]+"…") if len(_ac_entreg) > 16 else _ac_entreg
                            _resp_short     = (_ac_resp[:16]+"…")   if len(_ac_resp)   > 16 else _ac_resp

                            _ac_vol_dec = _srow.get("seg_vol_declarado", "")
                            _ac_vol_m   = _srow.get("seg_vol_muestras",  "")
                            _ac_dif_vol = _srow.get("seg_diferencia_vol","")
                            _ac_st_r    = _srow.get("seg_solidos_ruta",  "")
                            _ac_ic_r    = _srow.get("seg_crioscopia_ruta","")
                            _ac_st_p    = _srow.get("seg_st_pond",       "")
                            _ac_ic_p    = _srow.get("seg_ic_pond",       "")

                            try:    _v_ac_vol_dec = f"{int(float(_ac_vol_dec)):,} L"
                            except: _v_ac_vol_dec = "—"
                            try:    _v_ac_vol_m   = f"{int(float(_ac_vol_m)):,} L"
                            except: _v_ac_vol_m   = "—"
                            try:
                                _dif_v = int(float(_ac_dif_vol))
                                _dif_ok = abs(_dif_v) <= 20
                                _v_ac_dif = f"{_dif_v:+,} L"
                                _b_ac_dif = ("✔ OK" if _dif_ok else "⚠ DIFERENCIA", _dif_ok)
                            except: _v_ac_dif = "—"; _b_ac_dif = (None, True)
                            try:
                                _st_rv = float(_ac_st_r)
                                _st_ok = _st_rv >= 12.60
                                _v_ac_st_r = f"{_st_rv:.2f} %"
                                _b_ac_st   = ("✔ CONFORME" if _st_ok else "✖ DESVIACIÓN", _st_ok)
                            except: _v_ac_st_r = "—"; _b_ac_st = (None, True)
                            try:
                                _ic_rv = float(_ac_ic_r)
                                _ic_ok = -0.550 <= _ic_rv <= -0.535
                                _v_ac_ic_r = f"{_ic_rv:.3f} °C"
                                _b_ac_ic   = ("✔ CONFORME" if _ic_ok else "✖ DESVIACIÓN", _ic_ok)
                            except: _v_ac_ic_r = "—"; _b_ac_ic = (None, True)
                            try:    _v_ac_st_p = f"{float(_ac_st_p):.2f} %"
                            except: _v_ac_st_p = "—"
                            try:    _v_ac_ic_p = f"{float(_ac_ic_p):.3f} °C"
                            except: _v_ac_ic_p = "—"
                            try:    _v_ac_dif_st = f"{float(_ac_st_r) - float(_ac_st_p):+.2f} %"
                            except: _v_ac_dif_st = "—"
                            try:    _v_ac_dif_ic = f"{float(_ac_ic_r) - float(_ac_ic_p):+.3f} °C"
                            except: _v_ac_dif_ic = "—"

                            _ac_row1_html = (
                                '<div style="display:grid;grid-template-columns:repeat(3,1fr);'
                                'gap:8px;margin-bottom:8px;">'
                                + _kpi_card_ac("ENTREGADO POR",  _entreg_short)
                                + _kpi_card_ac("RUTA",           _ac_ruta)
                                + _kpi_card_ac("Nº MUESTRAS",    str(_ac_n_muestras))
                                + '</div>'
                            )
                            _ac_row2_html = (
                                '<div style="display:grid;grid-template-columns:repeat(3,1fr);'
                                'gap:8px;margin-bottom:8px;">'
                                + _kpi_card_ac("VOL. DECLARADO",     _v_ac_vol_dec)
                                + _kpi_card_ac("VOL. SUMA MUESTRAS", _v_ac_vol_m)
                                + _kpi_card_ac("DIFERENCIA VOL.",    _v_ac_dif, _b_ac_dif[0], _b_ac_dif[1])
                                + '</div>'
                            )
                            _ac_row3_html = (
                                '<div style="display:grid;grid-template-columns:repeat(6,1fr);'
                                'gap:8px;margin-bottom:14px;">'
                                + _kpi_card_ac("ST RUTA (%)",     _v_ac_st_r, _b_ac_st[0], _b_ac_st[1])
                                + _kpi_card_ac("ST PONDERADO",    _v_ac_st_p)
                                + _kpi_card_ac("ΔST (RUTA−POND)", _v_ac_dif_st)
                                + _kpi_card_ac("IC RUTA (°C)",    _v_ac_ic_r, _b_ac_ic[0], _b_ac_ic[1])
                                + _kpi_card_ac("IC PONDERADO",    _v_ac_ic_p)
                                + _kpi_card_ac("ΔIC (RUTA−POND)", _v_ac_dif_ic)
                                + '</div>'
                            )
                            st.markdown(_ac_row1_html + _ac_row2_html + _ac_row3_html, unsafe_allow_html=True)

                            st.markdown(
                                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                "letter-spacing:.05em;margin-bottom:4px;'>"
                                "📋 CALIDAD POR MUESTRA</div>",
                                unsafe_allow_html=True,
                            )
                            _cat_ac = load_catalogo()
                            _cat_ac_map = dict(zip(_cat_ac["codigo"], _cat_ac["nombre"]))
                            _acomp_det_rows = []
                            for _am in _mj_data:
                                _cod_am = str(_am.get("ID", "") or "").strip()
                                _vol_am = _pnac(_am.get("_volumen"))
                                _st_am  = _pnac(_am.get("_st"))
                                _ic_am  = _pnac(_am.get("_ic"))
                                _pst_am = round(_vol_am * _st_am, 2) if _vol_am is not None and _st_am is not None else None
                                _pic_am = round(_vol_am * _ic_am, 3) if _vol_am is not None and _ic_am is not None else None
                                _acomp_det_rows.append({
                                    "CÓDIGO":          _cod_am,
                                    "NOMBRE ESTACIÓN": _cat_ac_map.get(_cod_am, ""),
                                    "GRASA (%)":       _pnac(_am.get("_grasa")),
                                    "ST (%)":          _st_am,
                                    "PROTEÍNA (%)":    _pnac(_am.get("_proteina")),
                                    "IC (°C)":         _ic_am,
                                    "AGUA (%)":        _pnac(_am.get("_agua")),
                                    "VOLUMEN (L)":     int(_vol_am) if _vol_am is not None else None,
                                    "ALCOHOL":         _am.get("_alcohol", "N/A") or "N/A",
                                    "CLORUROS":        _am.get("_cloruros", "N/A") or "N/A",
                                    "NEUTRALIZANTES":  _am.get("_neutralizantes", "N/A") or "N/A",
                                    "OBS":             _am.get("_obs", "") or "",
                                })
                            _df_ac_det = pd.DataFrame(_acomp_det_rows)
                            _RED_AC = "background-color:#FFC7CE;color:#9C0006;font-weight:700"
                            def _color_ac(row):
                                styles = [""] * len(row)
                                cols = list(row.index)
                                try:
                                    _ic_v = row.get("IC (°C)")
                                    if _ic_v is not None:
                                        _icf = float(_ic_v)
                                        if (_icf > -0.530 or _icf < -0.550) and "IC (°C)" in cols:
                                            styles[cols.index("IC (°C)")] = _RED_AC
                                except Exception: pass
                                for _qc in ("ALCOHOL", "CLORUROS", "NEUTRALIZANTES"):
                                    try:
                                        if row.get(_qc) == "+" and _qc in cols:
                                            styles[cols.index(_qc)] = _RED_AC
                                    except Exception: pass
                                return styles
                            _fmt_ac = {
                                "GRASA (%)": "{:.2f}", "ST (%)": "{:.2f}",
                                "PROTEÍNA (%)": "{:.2f}",
                                "IC (°C)": "{:.3f}",   "AGUA (%)": "{:.2f}",
                            }
                            st.dataframe(
                                _df_ac_det.style.apply(_color_ac, axis=1).format(_fmt_ac, na_rep="—"),
                                width='stretch', hide_index=True,
                                height=min(38 + 35 * len(_acomp_det_rows), 420),
                                column_config={
                                    "CÓDIGO":           st.column_config.TextColumn("CÓDIGO",          width="small"),
                                    "NOMBRE ESTACIÓN":  st.column_config.TextColumn("NOMBRE ESTACIÓN", width="medium"),
                                    "GRASA (%)":        st.column_config.NumberColumn("GRASA (%)",      width="small", format="%.2f"),
                                    "ST (%)":           st.column_config.NumberColumn("ST (%)",         width="small", format="%.2f"),
                                    "PROTEÍNA (%)":     st.column_config.NumberColumn("PROTEÍNA (%)",   width="small", format="%.2f"),
                                    "IC (°C)":          st.column_config.NumberColumn("IC (°C)",        width="small", format="%.3f"),
                                    "AGUA (%)":         st.column_config.NumberColumn("AGUA (%)",       width="small", format="%.2f"),
                                    "VOLUMEN (L)":      st.column_config.NumberColumn("VOL. (L)",       width="small", format="%d"),
                                    "ALCOHOL":          st.column_config.TextColumn("ALC.",             width="small"),
                                    "CLORUROS":         st.column_config.TextColumn("CLOR.",            width="small"),
                                    "NEUTRALIZANTES":   st.column_config.TextColumn("NEUT.",            width="small"),
                                    "OBS":              st.column_config.TextColumn("OBSERVACIONES",    width="medium"),
                                },
                            )
                            # Fotos asociadas
                            _s_fotos_raw = str(_srow.get("fotos_json", "") or "").strip()
                            if _s_fotos_raw and _s_fotos_raw not in ("[]", ""):
                                try:
                                    _s_fotos_list = json.loads(_s_fotos_raw)
                                except Exception:
                                    _s_fotos_list = []
                                _s_fotos_ok = [p for p in _s_fotos_list if os.path.exists(p)]
                                if _s_fotos_ok:
                                    st.markdown(
                                        "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                        "letter-spacing:.05em;margin:10px 0 6px;'>"
                                        "📷 IMÁGENES DE MUESTRAS</div>",
                                        unsafe_allow_html=True,
                                    )
                                    _cols_sfoto = st.columns(min(len(_s_fotos_ok), 4))
                                    for _sfi, _sfp in enumerate(_s_fotos_ok):
                                        with _cols_sfoto[_sfi % 4]:
                                            st.image(_sfp, width='stretch')

                elif _s_sub == "ESTACIONES":
                    st.markdown("---")
                    _est_cod    = str(_srow.get("seg_codigo", "") or "").strip() or "—"
                    _est_fecha  = str(_srow.get("fecha", "") or "").strip() or "—"
                    _cat_est_df = load_catalogo()
                    _cat_est_m  = dict(zip(_cat_est_df["codigo"], _cat_est_df["nombre"]))
                    _est_nombre = _cat_est_m.get(_est_cod, "")
                    _est_grasa  = _srow.get("seg_grasa", "")
                    _est_st     = _srow.get("seg_st", "")
                    _est_ic     = _srow.get("seg_ic", "")
                    _est_agua   = _srow.get("seg_agua", "")
                    _est_alc    = str(_srow.get("seg_alcohol", "")        or "").strip() or "N/A"
                    _est_clor   = str(_srow.get("seg_cloruros", "")       or "").strip() or "N/A"
                    _est_neut   = str(_srow.get("seg_neutralizantes", "") or "").strip() or "N/A"
                    _est_obs    = str(_srow.get("seg_observaciones", "")  or "").strip()

                    _nom_label = f" — {_est_nombre}" if _est_nombre else ""
                    st.markdown(
                        f"""<div style="background:#0056A3;border-radius:8px;
                                       padding:10px 16px;margin-bottom:14px;">
                              <span style="font-size:1rem;font-weight:700;color:#fff;
                                           letter-spacing:.05em;">
                                🏭 DETALLE ESTACIÓN
                              </span>
                              <span style="font-size:.85rem;color:#cce0f5;margin-left:12px;">
                                {_est_cod}{_nom_label} &nbsp;·&nbsp; {_est_fecha}
                              </span>
                            </div>""",
                        unsafe_allow_html=True,
                    )

                    _det_accion_est = st.session_state.get("admin_accion")
                    _det_idx_est    = st.session_state.get("admin_idx")
                    _edit_mode_est  = (
                        _det_accion_est == "modificar"
                        and _det_idx_est == sel_orig_idx
                    )

                    if _edit_mode_est:
                        try:
                            _fe_e = datetime.strptime(
                                str(_srow.get("fecha", "")), "%d/%m/%Y"
                            ).date()
                        except Exception:
                            _fe_e = date.today()
                        try:    _ee_grasa_v = float(str(_est_grasa or 0).replace(",", "."))
                        except: _ee_grasa_v = 0.0
                        try:    _ee_st_v    = float(str(_est_st    or 0).replace(",", "."))
                        except: _ee_st_v    = 0.0
                        try:    _ee_ic_v    = float(str(_est_ic    or 0).replace(",", "."))
                        except: _ee_ic_v    = 0.0
                        try:    _ee_agua_v  = float(str(_est_agua  or 0).replace(",", "."))
                        except: _ee_agua_v  = 0.0

                        _cat_ed_df  = load_catalogo()
                        _cat_ed_cod = dict(zip(_cat_ed_df["codigo"], _cat_ed_df["nombre"]))
                        _cat_ed_nom = dict(zip(_cat_ed_df["nombre"], _cat_ed_df["codigo"]))

                        def _ed_fill_nom():
                            cod = st.session_state.get("edit_est_cod", "").strip().upper()
                            st.session_state["edit_est_cod"] = cod
                            nom = _cat_ed_cod.get(cod)
                            if nom:
                                st.session_state["edit_est_nombre"] = nom

                        def _ed_fill_cod():
                            raw = st.session_state.get("edit_est_nombre", "").upper()
                            st.session_state["edit_est_nombre"] = raw
                            cod = _cat_ed_nom.get(raw.strip())
                            if cod:
                                st.session_state["edit_est_cod"] = cod

                        st.session_state.setdefault(
                            "edit_est_nombre",
                            _cat_ed_cod.get(_est_cod if _est_cod != "—" else "", _est_nombre),
                        )

                        ee1, ee2, ee3 = st.columns([1, 1, 2])
                        _ee_fecha = ee1.date_input(
                            "📅 FECHA", value=_fe_e,
                            format="DD/MM/YYYY", key="edit_est_fecha",
                        )
                        _ee_cod = ee2.text_input(
                            "🔖 CÓDIGO",
                            value=_est_cod if _est_cod != "—" else "",
                            key="edit_est_cod",
                            on_change=_ed_fill_nom,
                        )
                        ee3.text_input(
                            "📍 NOMBRE ESTACIÓN",
                            key="edit_est_nombre",
                            on_change=_ed_fill_cod,
                        )
                        ep1, ep2, ep3, ep4 = st.columns(4)
                        _ee_grasa = ep1.number_input(
                            "GRASA (%)", value=_ee_grasa_v,
                            min_value=0.0, max_value=100.0,
                            step=0.01, format="%.2f", key="edit_est_grasa",
                        )
                        _ee_st = ep2.number_input(
                            "ST (%)", value=_ee_st_v,
                            min_value=0.0, max_value=100.0,
                            step=0.01, format="%.2f", key="edit_est_st",
                        )
                        _ee_ic = ep3.number_input(
                            "IC (°C)", value=_ee_ic_v,
                            step=0.001, format="%.3f", key="edit_est_ic",
                        )
                        _ee_agua = ep4.number_input(
                            "AGUA (%)", value=_ee_agua_v,
                            min_value=0.0, max_value=100.0,
                            step=0.01, format="%.2f", key="edit_est_agua",
                        )
                        _tri = ["N/A", "NEGATIVO (−)", "POSITIVO (+)"]
                        ea1, ea2, ea3 = st.columns(3)
                        _ee_alc = ea1.selectbox(
                            "ALCOHOL", _tri,
                            index=_tri.index(_est_alc) if _est_alc in _tri else 0,
                            key="edit_est_alc",
                        )
                        _ee_clor = ea2.selectbox(
                            "CLORUROS", _tri,
                            index=_tri.index(_est_clor) if _est_clor in _tri else 0,
                            key="edit_est_clor",
                        )
                        _ee_neut = ea3.selectbox(
                            "NEUTRALIZANTES", _tri,
                            index=_tri.index(_est_neut) if _est_neut in _tri else 0,
                            key="edit_est_neut",
                        )
                        _ee_obs = st.text_area(
                            "📝 OBSERVACIONES", value=_est_obs,
                            height=80, key="edit_est_obs",
                        )
                        activar_siguiente_con_enter()
                        es1, es2, _ = st.columns([1.5, 1, 3])
                        with es1:
                            if st.button(
                                "💾 GUARDAR CAMBIOS", type="primary",
                                width='stretch', key="btn_save_edit_est",
                            ):
                                update_seg_row_in_csv(_det_idx_est, {
                                    "fecha":              _ee_fecha.strftime("%d/%m/%Y"),
                                    "seg_codigo":         st.session_state.get("edit_est_cod", _ee_cod).upper().strip(),
                                    "seg_grasa":          round(float(_ee_grasa), 2),
                                    "seg_st":             round(float(_ee_st), 2),
                                    "seg_ic":             round(float(_ee_ic), 3),
                                    "seg_agua":           round(float(_ee_agua), 2),
                                    "seg_alcohol":        _ee_alc,
                                    "seg_cloruros":       _ee_clor,
                                    "seg_neutralizantes": _ee_neut,
                                    "seg_observaciones":  _ee_obs,
                                })
                                st.session_state.pop("edit_est_nombre", None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()
                        with es2:
                            if st.button(
                                "✖ CANCELAR", width='stretch',
                                key="btn_cancel_edit_est",
                            ):
                                st.session_state.pop("edit_est_nombre", None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()

                    else:
                        def _kpi_e(label, value, badge=None, badge_ok=True):
                            bh = ""
                            if badge:
                                bg = "#D4EDDA" if badge_ok else "#F8D7DA"
                                cl = "#155724" if badge_ok else "#721C24"
                                bh = (
                                    f'<div style="margin-top:4px;font-size:.65rem;font-weight:700;'
                                    f'color:{cl};background:{bg};border-radius:4px;'
                                    f'padding:1px 5px;display:inline-block;">{badge}</div>'
                                )
                            return (
                                f'<div style="background:#fff;border:1px solid #dde6f0;'
                                f'border-radius:8px;padding:10px 12px;text-align:center;height:100%;">'
                                f'<div style="font-size:.62rem;font-weight:700;color:#6c8ca8;'
                                f'letter-spacing:.06em;margin-bottom:4px;">{label}</div>'
                                f'<div style="font-size:1.05rem;font-weight:800;color:#0056A3;">{value}</div>'
                                f'{bh}</div>'
                            )

                        try:    _v_eg  = f"{float(_est_grasa):.2f} %"
                        except: _v_eg  = "—"
                        try:
                            _sv = float(_est_st)
                            _v_est_st = f"{_sv:.2f} %"
                            _b_est_st = (
                                "✔ CONFORME" if _sv >= 12.60 else "✖ DESVIACIÓN",
                                _sv >= 12.60,
                            )
                        except:
                            _v_est_st = "—"; _b_est_st = (None, True)
                        try:
                            _iv = float(_est_ic)
                            _v_eic = f"{_iv:.3f} °C"
                            _ic_ok_e = -0.550 <= _iv <= -0.530
                            _b_eic = (
                                "✔ CONFORME" if _ic_ok_e else "✖ DESVIACIÓN",
                                _ic_ok_e,
                            )
                        except:
                            _v_eic = "—"; _b_eic = (None, True)
                        try:    _v_eag = f"{float(_est_agua):.2f} %"
                        except: _v_eag = "—"

                        _er1 = (
                            '<div style="display:grid;grid-template-columns:repeat(4,1fr);'
                            'gap:8px;margin-bottom:8px;">'
                            + _kpi_e("NOMBRE ESTACIÓN", _est_nombre or "—")
                            + _kpi_e("GRASA (%)", _v_eg)
                            + _kpi_e("ST (%)", _v_est_st, _b_est_st[0], _b_est_st[1])
                            + _kpi_e("IC (°C)", _v_eic, _b_eic[0], _b_eic[1])
                            + '</div>'
                        )
                        _er2 = (
                            '<div style="display:grid;grid-template-columns:repeat(4,1fr);'
                            'gap:8px;margin-bottom:8px;">'
                            + _kpi_e("AGUA (%)", _v_eag)
                            + _kpi_e("ALCOHOL", _est_alc)
                            + _kpi_e("CLORUROS", _est_clor)
                            + _kpi_e("NEUTRALIZANTES", _est_neut)
                            + '</div>'
                        )
                        st.markdown(_er1 + _er2, unsafe_allow_html=True)
                        if _est_obs:
                            st.markdown(f"**Observaciones:** {_est_obs}")

                elif _s_sub == "CONTRAMUESTRAS SOLICITADAS":
                    st.markdown("---")
                    _ct_fecha  = str(_srow.get("fecha", "") or "").strip() or "—"
                    _ct_resp   = str(_srow.get("seg_responsable", "") or "").strip() or "—"
                    _ct_id     = str(_srow.get("seg_id_muestra", "") or "").strip() or "—"
                    _ct_cat     = load_catalogo()
                    _ct_cat_map = dict(zip(_ct_cat["codigo"].astype(str), _ct_cat["nombre"]))
                    _ct_nombre  = _ct_cat_map.get(_ct_id, "")
                    _ct_grasa   = _srow.get("seg_grasa",  "")
                    _ct_st      = _srow.get("seg_st",     "")
                    _ct_ic      = _srow.get("seg_ic",     "")
                    _ct_agua    = _srow.get("seg_agua",   "")
                    _ct_alc     = str(_srow.get("seg_alcohol",        "") or "").strip() or "N/A"
                    _ct_clor    = str(_srow.get("seg_cloruros",       "") or "").strip() or "N/A"
                    _ct_neut    = str(_srow.get("seg_neutralizantes", "") or "").strip() or "N/A"
                    _ct_obs     = str(_srow.get("seg_observaciones",  "") or "").strip()

                    # Parsear muestras_json (formato nuevo) — retrocompat. si vacío
                    _ct_mj_raw  = str(_srow.get("muestras_json", "") or "").strip()
                    try:
                        _ct_mj_data = json.loads(_ct_mj_raw) if _ct_mj_raw else []
                    except Exception:
                        _ct_mj_data = []
                    _ct_is_multi = bool(_ct_mj_data)  # True = nuevo formato

                    # Si es formato viejo (campos planos), construir lista de una muestra
                    if not _ct_is_multi:
                        _ct_mj_data = [{
                            "ID":             _ct_id if _ct_id != "—" else "",
                            "_grasa":         _ct_grasa, "_st": _ct_st,
                            "_ic":            _ct_ic,   "_agua": _ct_agua,
                            "_alcohol":       _ct_alc,  "_cloruros": _ct_clor,
                            "_neutralizantes":_ct_neut, "_obs": _ct_obs,
                        }]

                    _ct_n_muestras = len(_ct_mj_data)
                    _ct_header_sub = (f"{_ct_n_muestras} muestra(s)"
                                      if _ct_is_multi else
                                      f"{_ct_id}{f' — {_ct_nombre}' if _ct_nombre else ''}")

                    st.markdown(
                        f"""<div style="background:#0056A3;border-radius:8px;
                                       padding:10px 16px;margin-bottom:14px;">
                              <span style="font-size:1rem;font-weight:700;color:#fff;
                                           letter-spacing:.05em;">
                                🧪 DETALLE CONTRAMUESTRA
                              </span>
                              <span style="font-size:.85rem;color:#cce0f5;margin-left:12px;">
                                {_ct_header_sub} &nbsp;·&nbsp; {_ct_fecha}
                              </span>
                            </div>""",
                        unsafe_allow_html=True,
                    )

                    # ── Detección de modo edición ───────────────────────────────
                    _det_accion_ct = st.session_state.get("admin_accion")
                    _det_idx_ct    = st.session_state.get("admin_idx")
                    _edit_mode_ct  = (
                        _det_accion_ct == "modificar"
                        and _det_idx_ct == sel_orig_idx
                    )

                    if _edit_mode_ct:
                        # ── MODO EDICIÓN CONTRAMUESTRA ──────────────────────────
                        try:
                            _fe_ct = datetime.strptime(
                                str(_srow.get("fecha", "")), "%d/%m/%Y"
                            ).date()
                        except Exception:
                            _fe_ct = date.today()

                        # Encabezado
                        cta1, cta2 = st.columns(2)
                        _cte_fecha = cta1.date_input(
                            "📅 FECHA", value=_fe_ct,
                            format="DD/MM/YYYY", key="edit_ct_fecha",
                        )
                        _cte_resp = cta2.text_input(
                            "👤 ENTREGADO POR",
                            value=_ct_resp if _ct_resp != "—" else "",
                            key="edit_ct_resp",
                            on_change=convertir_a_mayusculas,
                            args=("edit_ct_resp",),
                        )

                        # Tabla de muestras (data_editor)
                        st.markdown(
                            "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                            "letter-spacing:.05em;margin:10px 0 4px;'>🔬 MUESTRAS</div>",
                            unsafe_allow_html=True,
                        )
                        st.caption("💡 Tab / → / Enter → siguiente celda · ← → anterior · Enter en última celda → guarda")

                        _CT_COLS    = ["ID","_proveedor","_grasa","_st","_ic","_agua",
                                       "_alcohol","_cloruros","_neutralizantes","_obs"]
                        _ct_cache_k = f"_ct_m_cache_{sel_orig_idx}"
                        _cat_ct_ed  = load_catalogo()
                        _cat_ct_ed_m = dict(zip(_cat_ct_ed["codigo"], _cat_ct_ed["nombre"]))

                        if _ct_cache_k in st.session_state:
                            _df_ct_ed = st.session_state[_ct_cache_k].copy()
                        else:
                            _df_ct_ed = pd.DataFrame(
                                [{k: m.get(k, "") for k in _CT_COLS} for m in _ct_mj_data],
                                columns=_CT_COLS,
                            )
                            for _nc in ["_grasa","_st","_ic","_agua"]:
                                _df_ct_ed[_nc] = pd.to_numeric(_df_ct_ed[_nc], errors="coerce")

                        _df_ct_ed["nombre_estacion"] = _df_ct_ed["ID"].apply(
                            lambda c: _cat_ct_ed_m.get(str(c).strip(), "") if pd.notna(c) else ""
                        )
                        _tri_ct_de = ["N/A", "NEGATIVO (−)", "POSITIVO (+)"]
                        _edited_ct = st.data_editor(
                            _df_ct_ed, num_rows="dynamic", width='stretch',
                            key="edit_ct_de",
                            column_config={
                                "ID":              st.column_config.TextColumn("CÓDIGO"),
                                "nombre_estacion": st.column_config.TextColumn("NOMBRE", disabled=True),
                                "_proveedor":      st.column_config.TextColumn("PROVEEDOR"),
                                "_grasa":          st.column_config.NumberColumn("GRASA (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                "_st":             st.column_config.NumberColumn("ST (%)",     format="%.2f", min_value=0.0, max_value=100.0),
                                "_ic":             st.column_config.NumberColumn("IC (°C)",    format="%.3f"),
                                "_agua":           st.column_config.NumberColumn("AGUA (%)",   format="%.2f", min_value=0.0, max_value=100.0),
                                "_alcohol":        st.column_config.SelectboxColumn("ALCOHOL",        options=_tri_ct_de, required=True),
                                "_cloruros":       st.column_config.SelectboxColumn("CLORUROS",       options=_tri_ct_de, required=True),
                                "_neutralizantes": st.column_config.SelectboxColumn("NEUT.",           options=_tri_ct_de, required=True),
                                "_obs":            st.column_config.TextColumn("OBS"),
                            },
                            hide_index=True,
                        )
                        _ct_prev_ids = list(_df_ct_ed["ID"].fillna("").astype(str))
                        _ct_new_ids  = list(_edited_ct["ID"].fillna("").astype(str))
                        _ct_save_df  = _edited_ct.drop(columns=["nombre_estacion"], errors="ignore").copy()
                        st.session_state[_ct_cache_k] = _ct_save_df
                        if _ct_prev_ids != _ct_new_ids:
                            st.rerun()

                        activar_siguiente_con_enter()
                        ct_s1, ct_s2, _ = st.columns([1.5, 1, 3])
                        with ct_s1:
                            if st.button(
                                "💾 GUARDAR CAMBIOS", type="primary",
                                width='stretch', key="btn_save_edit_ct",
                            ):
                                def _pf_ct(v, d=0.0):
                                    try: return float(str(v).replace(",","."))
                                    except: return d
                                _new_ct_ms = []
                                for _, _ctrow in _ct_save_df.iterrows():
                                    _new_ct_ms.append({
                                        "ID":              str(_ctrow.get("ID","") or "").upper().strip(),
                                        "_proveedor":      str(_ctrow.get("_proveedor","") or "").upper().strip(),
                                        "_grasa":          round(_pf_ct(_ctrow.get("_grasa",0)), 2),
                                        "_st":             round(_pf_ct(_ctrow.get("_st",0)), 2),
                                        "_ic":             round(_pf_ct(_ctrow.get("_ic",-0.530) or -0.530), 3),
                                        "_agua":           round(_pf_ct(_ctrow.get("_agua",0)), 2),
                                        "_alcohol":        str(_ctrow.get("_alcohol","N/A") or "N/A"),
                                        "_cloruros":       str(_ctrow.get("_cloruros","N/A") or "N/A"),
                                        "_neutralizantes": str(_ctrow.get("_neutralizantes","N/A") or "N/A"),
                                        "_obs":            str(_ctrow.get("_obs","") or ""),
                                    })
                                update_seg_row_in_csv(sel_orig_idx, {
                                    "fecha":           _cte_fecha.strftime("%d/%m/%Y"),
                                    "seg_responsable": str(_cte_resp).upper().strip(),
                                    "seg_id_muestra":  _new_ct_ms[0].get("ID","") if _new_ct_ms else "",
                                    "muestras_json":   json.dumps(_new_ct_ms, ensure_ascii=False),
                                })
                                st.session_state.pop(_ct_cache_k, None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()
                        with ct_s2:
                            if st.button(
                                "✖ CANCELAR", width='stretch',
                                key="btn_cancel_edit_ct",
                            ):
                                st.session_state.pop(_ct_cache_k, None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()

                    else:
                        # ── MODO VISTA CONTRAMUESTRA ────────────────────────────
                        if _ct_resp and _ct_resp != "—":
                            st.markdown(
                                f'<div style="font-size:.85rem;color:#555;margin-bottom:8px;">'
                                f'<b>Entregado por:</b> {_ct_resp}</div>',
                                unsafe_allow_html=True,
                            )

                        # Tabla de todas las muestras (nuevo y viejo formato)
                        _ct_view_rows = []
                        _RED_CT = "background-color:#FFC7CE;color:#9C0006;font-weight:700"
                        for _mv in _ct_mj_data:
                            def _pnv(x):
                                try:
                                    s = str(x).strip()
                                    return float(s.replace(",",".")) if s not in ("","None","nan") else None
                                except: return None
                            _cod_v = str(_mv.get("ID","") or "").strip()
                            _nom_v = _ct_cat_map.get(_cod_v, "")
                            _ct_view_rows.append({
                                "CÓDIGO":          _cod_v,
                                "NOMBRE ESTACIÓN": _nom_v,
                                "PROVEEDOR":       str(_mv.get("_proveedor","") or ""),
                                "GRASA (%)":       _pnv(_mv.get("_grasa")),
                                "ST (%)":          _pnv(_mv.get("_st")),
                                "IC (°C)":         _pnv(_mv.get("_ic")),
                                "AGUA (%)":        _pnv(_mv.get("_agua")),
                                "ALCOHOL":         str(_mv.get("_alcohol","N/A") or "N/A"),
                                "CLORUROS":        str(_mv.get("_cloruros","N/A") or "N/A"),
                                "NEUTRALIZANTES":  str(_mv.get("_neutralizantes","N/A") or "N/A"),
                                "OBS":             str(_mv.get("_obs","") or ""),
                            })
                        _df_ct_v = pd.DataFrame(_ct_view_rows)

                        def _color_ct_v(row):
                            styles = [""] * len(row)
                            cols = list(row.index)
                            try:
                                _icf = float(row.get("IC (°C)") or 0)
                                if (_icf > -0.530 or _icf < -0.550) and "IC (°C)" in cols:
                                    styles[cols.index("IC (°C)")] = _RED_CT
                            except: pass
                            try:
                                _stf = float(row.get("ST (%)") or 0)
                                if 0 < _stf < 12.60 and "ST (%)" in cols:
                                    styles[cols.index("ST (%)")] = _RED_CT
                            except: pass
                            for _qc in ("ALCOHOL","CLORUROS","NEUTRALIZANTES"):
                                try:
                                    if row.get(_qc) == "POSITIVO (+)" and _qc in cols:
                                        styles[cols.index(_qc)] = _RED_CT
                                except: pass
                            return styles

                        _fmt_ct_v = {
                            "GRASA (%)":"{:.2f}", "ST (%)":"{:.2f}",
                            "IC (°C)":"{:.3f}",   "AGUA (%)":"{:.2f}",
                        }
                        st.dataframe(
                            _df_ct_v.style.apply(_color_ct_v, axis=1).format(_fmt_ct_v, na_rep="—"),
                            width='stretch', hide_index=True,
                            height=min(38 + 35 * len(_ct_view_rows), 420),
                            column_config={
                                "CÓDIGO":           st.column_config.TextColumn("CÓDIGO",          width="small"),
                                "NOMBRE ESTACIÓN":  st.column_config.TextColumn("NOMBRE ESTACIÓN", width="medium"),
                                "PROVEEDOR":        st.column_config.TextColumn("PROVEEDOR",       width="medium"),
                                "GRASA (%)":        st.column_config.NumberColumn("GRASA (%)",      width="small", format="%.2f"),
                                "ST (%)":           st.column_config.NumberColumn("ST (%)",         width="small", format="%.2f"),
                                "IC (°C)":          st.column_config.NumberColumn("IC (°C)",        width="small", format="%.3f"),
                                "AGUA (%)":         st.column_config.NumberColumn("AGUA (%)",       width="small", format="%.2f"),
                                "ALCOHOL":          st.column_config.TextColumn("ALC.",             width="small"),
                                "CLORUROS":         st.column_config.TextColumn("CLOR.",            width="small"),
                                "NEUTRALIZANTES":   st.column_config.TextColumn("NEUT.",            width="small"),
                                "OBS":              st.column_config.TextColumn("OBSERVACIONES",    width="medium"),
                            },
                        )

            if (sel_orig_idx is not None and filtro_tipo in ("RUTAS", "TODOS", "TRANSUIZA")
                    and st.session_state.hist_buscar_ok):
                if sel_orig_idx in df_filtrado.index:
                    _drow = df_filtrado.loc[sel_orig_idx]
                else:
                    _df_all = load_historial()
                    _drow   = (_df_all.loc[sel_orig_idx]
                               if sel_orig_idx in _df_all.index else {})
                _tipo_reg = str(_drow.get("tipo_seguimiento", "RUTAS")).strip()
                _d_nombre = str(_drow.get("ruta", "")).strip() or "—"
                _d_fecha  = str(_drow.get("fecha", "")).strip() or "—"
                _d_placa  = str(_drow.get("placa",     "")).strip() or "—"
                _d_cond   = str(_drow.get("conductor", "")).strip() or "—"
                _d_vold   = _drow.get("volumen_declarado", "")
                _d_vole   = _drow.get("vol_estaciones",    "")
                _d_dif    = _drow.get("diferencia",        "")
                _d_st     = _drow.get("solidos_ruta",      "")
                _d_ic     = _drow.get("crioscopia_ruta",   "")
                _d_stpond = _drow.get("st_pond",           "")
                _d_icpond = _drow.get("ic_pond",           "")
                _d_nest   = _drow.get("num_estaciones",    "")
                # Campos exclusivos TRANSUIZA
                _d_st_car = _drow.get("st_carrotanque",   "")
                _d_grasa  = _drow.get("grasa_muestra",    "")
                _d_prot   = _drow.get("proteina_muestra", "")
                _d_dif_s  = _drow.get("diferencia_solidos","")

                st.markdown("---")

                # ── Encabezado del panel ──────────────────────────────────
                _panel_titulo = "🔍 DETALLE TRANSUIZA" if _tipo_reg == "TRANSUIZA" else "🔍 DETALLE DE RUTA"
                st.markdown(
                    f"""<div style="background:#0056A3;border-radius:8px;
                                   padding:10px 16px;margin-bottom:14px;">
                          <span style="font-size:1rem;font-weight:700;color:#fff;
                                       letter-spacing:.05em;">
                            {_panel_titulo}
                          </span>
                          <span style="font-size:.85rem;color:#cce0f5;margin-left:12px;">
                            {_d_placa} &nbsp;·&nbsp; {_d_fecha}
                          </span>
                        </div>""",
                    unsafe_allow_html=True,
                )

                # ── Detección de modo edición ─────────────────────────────
                _det_accion = st.session_state.get("admin_accion")
                _det_idx    = st.session_state.get("admin_idx")
                _edit_mode  = (_det_accion == "modificar" and _det_idx == sel_orig_idx)

                if _edit_mode:
                    # ══ MODO EDICIÓN: ramas por tipo de registro ════
                    try:
                        _fe_orig = datetime.strptime(str(_drow.get("fecha", "")), "%d/%m/%Y").date()
                    except Exception:
                        _fe_orig = date.today()

                    if _tipo_reg == "TRANSUIZA":
                        # ── Edición TRANSUIZA ─────────────────────────────
                        try:
                            _stcar_orig = float(str(_d_st_car or 0).replace(",","."))
                        except Exception: _stcar_orig = 0.0
                        try:
                            _grasa_orig = float(str(_d_grasa or 0).replace(",","."))
                        except Exception: _grasa_orig = 0.0
                        try:
                            _stm_orig = float(str(_d_st or 0).replace(",","."))
                        except Exception: _stm_orig = 0.0
                        try:
                            _prot_orig = float(str(_d_prot or 0).replace(",","."))
                        except Exception: _prot_orig = 0.0

                        te1, te2 = st.columns(2)
                        with te1:
                            _te_fecha = st.date_input("FECHA", value=_fe_orig,
                                                      format="DD/MM/YYYY", key="edit_t_fecha")
                            _te_placa = st.text_input("PLACA", value=_d_placa, key="edit_t_placa")
                            _te_stcar = st.number_input("ST DEL CARROTANQUE (%)",
                                                        value=_stcar_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_stcar")
                        with te2:
                            _te_grasa = st.number_input("GRASA (%)",
                                                        value=_grasa_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_grasa")
                            _te_stm   = st.number_input("ST MUESTRA (%)",
                                                        value=_stm_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_stm")
                            _te_prot  = st.number_input("PROTEÍNA (%)",
                                                        value=_prot_orig,
                                                        min_value=0.0, max_value=100.0,
                                                        step=0.01, format="%.2f",
                                                        key="edit_t_prot")
                        _te_dif = round(_te_stcar - _te_stm, 2)
                        _dif_col = "#9C0006" if abs(_te_dif) > 0.5 else "#006100"
                        st.markdown(
                            f"<div style='text-align:center;padding:8px;background:#F8FAFC;"
                            f"border-radius:8px;border:1px solid #D1D5DB;'>"
                            f"<div style='font-size:11px;font-weight:600;color:#6B7280;'>"
                            f"DIFERENCIA DE SÓLIDOS</div>"
                            f"<div style='font-size:1.5rem;font-weight:800;color:{_dif_col};'>"
                            f"{_te_dif:+.2f} %</div></div>",
                            unsafe_allow_html=True,
                        )
                        tec1, tec2, _ = st.columns([1.5, 1, 3])
                        with tec1:
                            if st.button("💾 GUARDAR CAMBIOS", type="primary",
                                         key="btn_save_edit_t", width='stretch'):
                                update_row_in_csv(_det_idx, {
                                    "fecha":              _te_fecha.strftime("%d/%m/%Y"),
                                    "placa":              str(_te_placa).upper(),
                                    "st_carrotanque":     round(float(_te_stcar), 2),
                                    "solidos_ruta":       round(float(_te_stm), 2),
                                    "grasa_muestra":      round(float(_te_grasa), 2),
                                    "proteina_muestra":   round(float(_te_prot), 2),
                                    "diferencia_solidos": _te_dif,
                                })
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()
                        with tec2:
                            if st.button("✖ CANCELAR", key="btn_cancel_edit_t",
                                         width='stretch'):
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()

                    else:
                        # ── Edición RUTAS ─────────────────────────────────
                        try:
                            _vol_orig = int(float(str(_drow.get("volumen_declarado", 0) or 0)))
                        except Exception:
                            _vol_orig = 0
                        try:
                            _st_orig = float(str(_drow.get("solidos_ruta", "0") or 0).replace(",", "."))
                        except Exception:
                            _st_orig = 0.0
                        try:
                            _ic_orig = float(str(_drow.get("crioscopia_ruta", "0") or 0).replace(",", "."))
                        except Exception:
                            _ic_orig = 0.0

                        ef1, ef2 = st.columns(2)
                        with ef1:
                            _e_fecha = st.date_input("FECHA", value=_fe_orig,
                                                     format="DD/MM/YYYY", key="edit_fecha")
                            _e_ruta  = st.text_input("RUTA",
                                                     value=str(_drow.get("ruta", "")), key="edit_ruta")
                            _e_placa = st.text_input("PLACA",
                                                     value=str(_drow.get("placa", "")), key="edit_placa")
                            _e_cond  = st.text_input("CONDUCTOR",
                                                     value=str(_drow.get("conductor", "")), key="edit_cond")
                        with ef2:
                            _e_vol = st.number_input("VOLUMEN DECLARADO (L)",
                                                     value=_vol_orig, min_value=0, step=1, key="edit_vol")
                            _e_st  = st.number_input("ST RUTA (%)", value=_st_orig,
                                                     step=0.01, format="%.2f", key="edit_st")
                            _e_ic  = st.number_input("IC RUTA (°C)", value=_ic_orig,
                                                     step=0.001, format="%.3f", key="edit_ic")

                        st.markdown(
                            "<div style='font-weight:700;color:#0056A3;margin:12px 0 6px;"
                            "font-size:.9rem;border-left:4px solid #0056A3;padding-left:8px;'>"
                            "🏭 Estaciones</div>",
                            unsafe_allow_html=True,
                        )
                        _ECOLS_E   = ["codigo", "grasa", "solidos", "proteina",
                                       "crioscopia", "agua_pct", "volumen", "alcohol",
                                       "cloruros", "neutralizantes", "obs"]
                        _est_json_e = str(_drow.get("estaciones_json", "") or "").strip()
                        try:
                            _est_data_e = json.loads(_est_json_e) if _est_json_e else []
                        except Exception:
                            _est_data_e = []
                        # ── Caché de edición para preservar cambios entre reruns ──
                        _h_cache_key = f"_h_est_cache_{st.session_state.get('admin_idx', 0)}"
                        _cat_h = load_catalogo()
                        _cat_h_map = dict(zip(_cat_h["codigo"], _cat_h["nombre"]))

                        if _h_cache_key in st.session_state:
                            # Usar la versión cacheada (conserva ediciones previas)
                            _df_est_e = st.session_state[_h_cache_key].copy()
                        else:
                            _df_est_e = (pd.DataFrame(_est_data_e, columns=_ECOLS_E)
                                         if _est_data_e else pd.DataFrame(columns=_ECOLS_E))
                            for _nc in ["grasa", "solidos", "proteina", "agua_pct"]:
                                _df_est_e[_nc] = pd.to_numeric(_df_est_e[_nc], errors="coerce")
                            _df_est_e["volumen"] = pd.to_numeric(
                                _df_est_e["volumen"], errors="coerce")

                        # Siempre recomputar nombre desde el catálogo (refleja código actual)
                        _df_est_e["nombre_estacion"] = _df_est_e["codigo"].apply(
                            lambda c: _cat_h_map.get(str(c).strip(), "")
                                      if pd.notna(c) else ""
                        )

                        st.caption("💡 Tab / → / Enter → siguiente celda · ← → anterior · Enter en última celda → guarda")
                        _edited_df_e = st.data_editor(
                            _df_est_e, num_rows="dynamic", width='stretch',
                            key="edit_est_editor",
                            column_config={
                                "codigo":         st.column_config.TextColumn("CÓDIGO"),
                                "grasa":          st.column_config.NumberColumn("GRASA (%)",     format="%.2f", min_value=0.0, max_value=100.0),
                                "solidos":        st.column_config.NumberColumn("SÓL.TOT. (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                "proteina":       st.column_config.NumberColumn("PROTEÍNA (%)",  format="%.2f", min_value=0.0, max_value=100.0),
                                "crioscopia":     st.column_config.TextColumn("CRIOSCOPIA (°C)"),
                                "volumen":        st.column_config.NumberColumn("VOLUMEN (L)",   format="%.0f", min_value=0,   step=1),
                                "alcohol":        st.column_config.SelectboxColumn("ALCOHOL",        options=["N/A", "+", "-"], required=True),
                                "cloruros":       st.column_config.SelectboxColumn("CLORUROS",       options=["N/A", "+", "-"], required=True),
                                "neutralizantes": st.column_config.SelectboxColumn("NEUTRALIZANTES", options=["N/A", "+", "-"], required=True),
                                "agua_pct":       st.column_config.NumberColumn("% AGUA",        format="%.1f", min_value=0.0, max_value=100.0),
                                "obs":            st.column_config.TextColumn("OBSERVACIONES"),
                                "nombre_estacion": st.column_config.TextColumn(
                                                     "NOMBRE ESTACIÓN", disabled=True),
                            },
                            hide_index=True,
                        )

                        # Guardar edición en caché (sin nombre_estacion) y refrescar si cambió código
                        _h_prev_codes = list(_df_est_e["codigo"].fillna("").astype(str).str.strip())
                        _h_new_codes  = list(_edited_df_e["codigo"].fillna("").astype(str).str.strip())
                        _cache_df = _edited_df_e.drop(columns=["nombre_estacion"], errors="ignore").copy()
                        st.session_state[_h_cache_key] = _cache_df
                        if _h_prev_codes != _h_new_codes:
                            st.rerun()

                        # ── Serialización robusta con sanitización ─────────────
                        def _sanitize_est_df(df):
                            """Uppercase + limpia espacios y chars especiales en texto."""
                            import re
                            df = df.copy()
                            if "codigo" in df.columns:
                                df["codigo"] = (df["codigo"].fillna("").astype(str)
                                                .str.strip().str.upper()
                                                .apply(lambda x: re.sub(r"[^A-Z0-9ÁÉÍÓÚÑ\-/]", "", x)))
                            if "obs" in df.columns:
                                df["obs"] = (df["obs"].fillna("").astype(str)
                                             .str.strip().str.upper())
                            if "crioscopia" in df.columns:
                                df["crioscopia"] = df["crioscopia"].fillna("").astype(str).str.strip()
                            return df

                        _clean_df = _sanitize_est_df(_edited_df_e)
                        _est_records = []
                        for _, _er in _clean_df.iterrows():
                            _rec = {}
                            for _ck in _clean_df.columns:
                                if _ck == "nombre_estacion":
                                    continue
                                _v = _er[_ck]
                                try:
                                    if pd.isna(_v):
                                        _rec[_ck] = None
                                    elif _ck == "volumen":
                                        _rec[_ck] = int(float(_v))
                                    elif hasattr(_v, "item"):
                                        _rec[_ck] = _v.item()
                                    else:
                                        _rec[_ck] = _v
                                except (TypeError, ValueError):
                                    _rec[_ck] = str(_v) if _v is not None else None
                            _est_records.append(_rec)
                        _edited_est_json = json.dumps(_est_records, ensure_ascii=False)

                        # ── Precalcular ponderados desde las estaciones editadas ──
                        _e_vols_p, _e_sum_st, _e_sum_ic = [], [], []
                        for _er2 in _est_records:
                            try:   _ev2 = float(_er2.get("volumen") or 0)
                            except: _ev2 = 0.0
                            try:   _es2 = float(str(_er2.get("solidos","") or "").replace(",","."))
                            except: _es2 = None
                            try:
                                _ec2_raw = str(_er2.get("crioscopia","") or "").strip()
                                _ec2 = float(_ec2_raw.replace(",",".")) if _ec2_raw else None
                            except Exception: _ec2 = None
                            if _ev2 > 0:
                                _e_vols_p.append(_ev2)
                                if _es2 is not None: _e_sum_st.append(_ev2 * _es2)
                                if _ec2 is not None: _e_sum_ic.append(_ev2 * _ec2)
                        _e_vol_total   = sum(_e_vols_p)
                        _e_vol_ests    = int(_e_vol_total) if _e_vol_total else 0
                        _e_diferencia  = int(_e_vol) - _e_vol_ests
                        _e_st_pond     = round(sum(_e_sum_st) / _e_vol_total, 2)  if _e_vol_total and _e_sum_st else ""
                        _e_ic_pond     = round(sum(_e_sum_ic) / _e_vol_total, 3)  if _e_vol_total and _e_sum_ic else ""

                        ec1, ec2, _ = st.columns([1.5, 1, 3])
                        with ec1:
                            if st.button("💾 GUARDAR CAMBIOS", type="primary",
                                         key="btn_save_edit", width='stretch'):
                                try:
                                    _ests_cnt = len(json.loads(_edited_est_json) or [])
                                except Exception:
                                    _ests_cnt = 0
                                update_row_in_csv(_det_idx, {
                                    "fecha":             _e_fecha.strftime("%d/%m/%Y"),
                                    "ruta":              str(_e_ruta).upper(),
                                    "placa":             str(_e_placa).upper(),
                                    "conductor":         str(_e_cond).upper(),
                                    "volumen_declarado": int(_e_vol),
                                    "solidos_ruta":      round(float(_e_st), 2),
                                    "crioscopia_ruta":   round(float(_e_ic), 3),
                                    "vol_estaciones":    _e_vol_ests,
                                    "diferencia":        _e_diferencia,
                                    "st_pond":           _e_st_pond,
                                    "ic_pond":           _e_ic_pond,
                                    "estaciones_json":   _edited_est_json,
                                    "num_estaciones":    _ests_cnt,
                                })
                                st.session_state.pop(f"_h_est_cache_{st.session_state.get('admin_idx', 0)}", None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()
                        with ec2:
                            if st.button("✖ CANCELAR", key="btn_cancel_edit",
                                         width='stretch'):
                                st.session_state.pop(f"_h_est_cache_{st.session_state.get('admin_idx', 0)}", None)
                                st.session_state.admin_accion = None
                                st.session_state.admin_idx    = None
                                st.rerun()

                if not _edit_mode:
                    # ══ MODO VISTA: tarjetas KPI + tabla estática ════════

                    # ── Fila 1 de KPIs (tarjetas HTML) ───────────────────────
                    def _kpi_card(label, value, badge=None, badge_ok=True):  # noqa: E301
                        badge_html = ""
                        if badge:
                            bg = "#D4EDDA" if badge_ok else "#F8D7DA"
                            col = "#155724" if badge_ok else "#721C24"
                            badge_html = (f'<div style="margin-top:4px;font-size:.65rem;'
                                          f'font-weight:700;color:{col};background:{bg};'
                                          f'border-radius:4px;padding:1px 5px;display:inline-block;">'
                                          f'{badge}</div>')
                        return (
                            f'<div style="background:#fff;border:1px solid #dde6f0;border-radius:8px;'
                            f'padding:10px 12px;text-align:center;height:100%;">'
                            f'<div style="font-size:.62rem;font-weight:700;color:#6c8ca8;'
                            f'letter-spacing:.06em;margin-bottom:4px;">{label}</div>'
                            f'<div style="font-size:1.05rem;font-weight:800;color:#0056A3;">{value}</div>'
                            f'{badge_html}</div>'
                        )

                    try:    _v_vold = f"{int(float(_d_vold)):,} L"
                    except: _v_vold = "—"
                    try:
                        _st_v  = float(_d_st)
                        _st_ok = _st_v >= 12.60
                        _v_st  = f"{_st_v:.2f} %"
                        _b_st  = ("✔ CONFORME" if _st_ok else "✖ DESVIACIÓN", _st_ok)
                    except: _v_st = "—"; _b_st = (None, True)
                    try:
                        _ic_v  = float(_d_ic)
                        _ic_ok = -0.550 <= _ic_v <= -0.535
                        _v_ic  = f"{_ic_v:.3f} °C"
                        _b_ic  = ("✔ CONFORME" if _ic_ok else "✖ DESVIACIÓN", _ic_ok)
                    except: _v_ic = "—"; _b_ic = (None, True)
                    try:    _v_nest = str(int(float(_d_nest)))
                    except: _v_nest = str(_d_nest)

                    # Valores auxiliares para filas 1 y 2
                    try:    _v_vole = f"{int(float(_d_vole)):,} L"
                    except: _v_vole = "—"
                    try:
                        _dif_v  = int(float(_d_dif))
                        _dif_ok = abs(_dif_v) <= 20
                        _v_dif  = f"{_dif_v:+,} L"
                        _b_dif  = ("✔ OK" if _dif_ok else "⚠ DIFERENCIA", _dif_ok)
                    except: _v_dif = "—"; _b_dif = (None, True)
                    try:
                        _stpv  = float(_d_stpond)
                        _v_stp = f"{_stpv:.2f} %"
                    except: _v_stp = "—"
                    try:
                        _dif_st_v = float(_d_st) - float(_d_stpond)
                        _v_dif_st = f"{_dif_st_v:+.2f} %"
                    except: _v_dif_st = "—"

                    # IC PONDERADO solo si TODAS las estaciones tienen crioscopia válida
                    _todas_con_ic = False
                    try:
                        _raw_check = str(_drow.get("estaciones_json", "") or "").strip()
                        if _raw_check:
                            _ests_check = json.loads(_raw_check)
                            if _ests_check:
                                _todas_con_ic = all(
                                    (lambda v: v is not None)(
                                        (lambda s: float(s.replace(",", "."))
                                         if s and str(s).strip() not in ("", "None", "nan") else None
                                        )(str(_ec.get("crioscopia", "") or ""))
                                    )
                                    for _ec in _ests_check
                                )
                    except Exception:
                        _todas_con_ic = False

                    # ── Filas KPI según tipo de registro ─────────────────────
                    if _tipo_reg == "TRANSUIZA":
                        # ── Vista TRANSUIZA ────────────────────────────────────
                        try:
                            _v_stcar = f"{float(str(_d_st_car or 0).replace(',','.')):.2f} %"
                        except: _v_stcar = "—"
                        try:
                            _stm_val = float(str(_d_st or 0).replace(",","."))
                            _stm_ok  = _stm_val >= 12.60
                            _v_stm   = f"{_stm_val:.2f} %"
                            _b_stm   = ("✔ CONFORME" if _stm_ok else "✖ DESVIACIÓN", _stm_ok)
                        except: _v_stm = "—"; _b_stm = (None, True)
                        try:
                            _v_grasa_t = f"{float(str(_d_grasa or 0).replace(',','.')):.2f} %"
                        except: _v_grasa_t = "—"
                        try:
                            _v_prot_t = f"{float(str(_d_prot or 0).replace(',','.')):.2f} %"
                        except: _v_prot_t = "—"
                        try:
                            _dif_s_v  = float(str(_d_dif_s or 0).replace(",","."))
                            _dif_s_ok = abs(_dif_s_v) <= 0.5
                            _v_dif_s  = f"{_dif_s_v:+.2f} %"
                            _b_dif_s  = ("✔ OK" if _dif_s_ok else "⚠ DIFERENCIA", _dif_s_ok)
                        except: _v_dif_s = "—"; _b_dif_s = (None, True)

                        _trans_kpi_html = (
                            '<div style="display:grid;grid-template-columns:repeat(5,1fr);'
                            'gap:8px;margin-bottom:14px;">'
                            + _kpi_card("PLACA",              _d_placa or "—")
                            + _kpi_card("ST CARROTANQUE",     _v_stcar)
                            + _kpi_card("ST MUESTRA",         _v_stm, _b_stm[0], _b_stm[1])
                            + _kpi_card("GRASA",              _v_grasa_t)
                            + _kpi_card("PROTEÍNA",           _v_prot_t)
                            + '</div>'
                            + '<div style="display:grid;grid-template-columns:repeat(1,1fr);'
                            'gap:8px;margin-bottom:14px;">'
                            + _kpi_card("DIF. DE SÓLIDOS (CARROTANQUE − MUESTRA)", _v_dif_s, _b_dif_s[0], _b_dif_s[1])
                            + '</div>'
                        )
                        st.markdown(_trans_kpi_html, unsafe_allow_html=True)

                    else:
                        # ── Fila 1: PLACA · CONDUCTOR · VOL. DECL. · ST RUTA · ST POND · ΔST
                        _row1_html = (
                            '<div style="display:grid;grid-template-columns:repeat(6,1fr);'
                            'gap:8px;margin-bottom:8px;">'
                            + _kpi_card("PLACA",          _d_placa or "—")
                            + _kpi_card("CONDUCTOR",      (_d_cond[:16]+"…") if len(_d_cond)>16 else (_d_cond or "—"))
                            + _kpi_card("VOL. DECLARADO", _v_vold)
                            + _kpi_card("ST RUTA",        _v_st,  _b_st[0], _b_st[1])
                            + _kpi_card("ST PONDERADO",   _v_stp)
                            + _kpi_card("ΔST (RUTA−POND)",_v_dif_st)
                            + '</div>'
                        )

                    # ── Fila 2: Nº EST · DIF. VOL. · VOL. EST · IC RUTA · [IC POND · ΔIC]
                        _row2_cards = [
                            _kpi_card("Nº ESTACIONES",  _v_nest),
                            _kpi_card("DIF. VOLUMEN",   _v_dif,  _b_dif[0], _b_dif[1]),
                            _kpi_card("VOL. ESTACIONES",_v_vole),
                            _kpi_card("IC RUTA",        _v_ic,   _b_ic[0],  _b_ic[1]),
                        ]
                        if _todas_con_ic:
                            try:
                                _icpv     = float(_d_icpond)
                                _dif_ic_v = float(_d_ic) - _icpv
                                _row2_cards.append(_kpi_card("IC PONDERADO",   f"{_icpv:.3f} °C"))
                                _row2_cards.append(_kpi_card("ΔIC (RUTA−POND)",f"{_dif_ic_v:+.3f} °C"))
                            except Exception:
                                pass

                        _ncols2 = len(_row2_cards)
                        _row2_html = (
                            f'<div style="display:grid;grid-template-columns:repeat({_ncols2},1fr);'
                            f'gap:8px;margin-bottom:14px;">'
                            + "".join(_row2_cards)
                            + '</div>'
                        )

                        st.markdown(_row1_html + _row2_html, unsafe_allow_html=True)

                        # ── Tabla de estaciones ───────────────────────────────────
                        _est_json_raw = str(_drow.get("estaciones_json", "") or "").strip()
                        if _est_json_raw:
                            try:
                                _ests = json.loads(_est_json_raw)
                                if _ests:
                                    st.markdown(
                                        "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                        "letter-spacing:.05em;margin-bottom:4px;'>"
                                        "📋 CALIDAD POR ESTACIÓN</div>",
                                        unsafe_allow_html=True,
                                    )

                                    def _pn(x):
                                        try: return float(str(x).replace(",", "."))
                                        except: return None

                                    _cat_det = load_catalogo()
                                    _cat_det_map = dict(zip(_cat_det["codigo"], _cat_det["nombre"]))
                                    _det_rows = []
                                    for _e in _ests:
                                        _ev  = _pn(_e.get("volumen"))
                                        _est_v = _pn(_e.get("solidos"))
                                        _eic = _pn(_e.get("crioscopia"))
                                        _pst = round(_ev * _est_v, 2) if _ev is not None and _est_v is not None else None
                                        _pic = round(_ev * _eic,   3) if _ev is not None and _eic is not None else None
                                        _cod_e = _e.get("codigo", "") or ""
                                        _det_rows.append({
                                            "CÓDIGO":           _cod_e,
                                            "GRASA (%)":        _pn(_e.get("grasa")),
                                            "SÓLIDOS TOT. (%)": _est_v,
                                            "PROTEÍNA (%)":     _pn(_e.get("proteina")),
                                            "CRIOSCOPIA":       _eic,
                                            "AGUA (%)":         _pn(_e.get("agua_pct")),
                                            "VOLUMEN (L)":      int(_ev) if _ev is not None else None,
                                            "ALC.":             _e.get("alcohol",        "N/A") or "N/A",
                                            "CLOR.":            _e.get("cloruros",       "N/A") or "N/A",
                                            "NEUT.":            _e.get("neutralizantes", "N/A") or "N/A",
                                            "OBSERVACIONES":    _e.get("obs", "") or "",
                                        })

                                    _df_det = pd.DataFrame(_det_rows)

                                    _RED_EST = "background-color:#FFC7CE;color:#9C0006;font-weight:700"

                                    def _color_est(row):
                                        styles = [""] * len(row)
                                        cols = list(row.index)
                                        try:
                                            if row.get("SÓLIDOS TOT. (%)") is not None and 0 < float(row["SÓLIDOS TOT. (%)"]) < 12.60:
                                                if "SÓLIDOS TOT. (%)" in cols:
                                                    styles[cols.index("SÓLIDOS TOT. (%)")] = _RED_EST
                                        except Exception: pass
                                        try:
                                            _icv = row.get("CRIOSCOPIA")
                                            if _icv is not None:
                                                _icf = float(_icv)
                                                if _icf > -0.530:
                                                    if "CRIOSCOPIA" in cols:
                                                        styles[cols.index("CRIOSCOPIA")] = _RED_EST
                                                    if "AGUA (%)" in cols:
                                                        styles[cols.index("AGUA (%)")] = _RED_EST
                                                elif _icf < -0.550:
                                                    if "CRIOSCOPIA" in cols:
                                                        styles[cols.index("CRIOSCOPIA")] = _RED_EST
                                        except Exception: pass
                                        for _qcol in ("ALC.", "CLOR.", "NEUT."):
                                            try:
                                                if row.get(_qcol) == "+" and _qcol in cols:
                                                    styles[cols.index(_qcol)] = _RED_EST
                                            except Exception: pass
                                        return styles

                                    _fmt_det = {
                                        "GRASA (%)":        "{:.2f}",
                                        "SÓLIDOS TOT. (%)": "{:.2f}",
                                        "PROTEÍNA (%)":     "{:.2f}",
                                        "CRIOSCOPIA":       "{:.3f}",
                                        "AGUA (%)":         "{:.1f}",
                                    }
                                    st.dataframe(
                                        _df_det.style
                                               .apply(_color_est, axis=1)
                                               .format(_fmt_det, na_rep="—"),
                                        width='stretch',
                                        hide_index=True,
                                        height=min(38 + 35 * len(_det_rows), 420),
                                        column_config={
                                            "CÓDIGO":           st.column_config.TextColumn("CÓDIGO",          width="small"),
                                            "GRASA (%)":        st.column_config.NumberColumn("GRASA (%)",      width="small",  format="%.2f"),
                                            "SÓLIDOS TOT. (%)": st.column_config.NumberColumn("ST (%)",         width="small",  format="%.2f"),
                                            "PROTEÍNA (%)":     st.column_config.NumberColumn("PROT. (%)",      width="small",  format="%.2f"),
                                            "CRIOSCOPIA":       st.column_config.NumberColumn("CRIOS.",         width="small",  format="%.3f"),
                                            "AGUA (%)":         st.column_config.NumberColumn("AGUA (%)",       width="small",  format="%.1f"),
                                            "VOLUMEN (L)":      st.column_config.NumberColumn("VOL. (L)",       width="small",  format="%d"),
                                            "ALC.":             st.column_config.TextColumn("ALC.",             width="small"),
                                            "CLOR.":            st.column_config.TextColumn("CLOR.",            width="small"),
                                            "NEUT.":            st.column_config.TextColumn("NEUT.",            width="small"),
                                            "OBSERVACIONES":    st.column_config.TextColumn("OBSERVACIONES",    width="medium"),
                                        },
                                    )
                            except Exception:
                                pass
                        else:
                            st.caption("Esta ruta no tiene datos de estaciones registrados.")

                    # ── Galería de fotos almacenadas ──────────────────────────
                    _fotos_raw = str(_drow.get("fotos_json", "") or "").strip()
                    if _fotos_raw and _fotos_raw not in ("[]", ""):
                        try:
                            _fotos_list = json.loads(_fotos_raw)
                        except Exception:
                            _fotos_list = []
                        _fotos_existentes = [p for p in _fotos_list if os.path.exists(p)]
                        if _fotos_existentes:
                            st.markdown(
                                "<div style='font-size:11px;font-weight:700;color:#0056A3;"
                                "letter-spacing:.05em;margin:10px 0 6px;'>"
                                "📷 IMÁGENES DE MUESTRAS</div>",
                                unsafe_allow_html=True,
                            )
                            _cols_fotos = st.columns(min(len(_fotos_existentes), 4))
                            for _fi, _fp in enumerate(_fotos_existentes):
                                with _cols_fotos[_fi % 4]:
                                    st.image(_fp, width='stretch')

            # ── Acción ELIMINAR ────────────────────────────────────────────
            accion_activa  = st.session_state.get("admin_accion")
            idx_activo     = st.session_state.get("admin_idx")
            _from_seg      = st.session_state.get("admin_from_seg", False)

            _tiene_indices = (idx_activo is not None or
                              bool(st.session_state.get("admin_idxs")))
            if accion_activa == "eliminar" and _tiene_indices:
                with st.container(border=True):
                    # ── Confirmación de borrado (1 o varios) ─────
                    _idxs_del = st.session_state.get("admin_idxs") or (
                        [idx_activo] if idx_activo is not None else []
                    )
                    n_del = len(_idxs_del)
                    st.markdown(
                        f"<div style='font-weight:700;color:#9C0006;margin-bottom:4px;'>"
                        f"🗑️ ¿Confirmar eliminación de "
                        f"{'1 registro' if n_del == 1 else f'{n_del} registros'}?</div>",
                        unsafe_allow_html=True,
                    )
                    # Cargar la fuente correcta para mostrar el detalle
                    _df_src_del = load_seguimientos() if _from_seg else df_hist
                    if n_del == 1 and idx_activo is not None:
                        row_a = _df_src_del.loc[idx_activo] if idx_activo in _df_src_del.index else {}
                        _lbl_extra = (
                            f"**Código:** {row_a.get('seg_codigo','')} &nbsp;·&nbsp; "
                            f"**Sub-tipo:** {row_a.get('sub_tipo_seguimiento','')}"
                            if _from_seg else
                            f"**Ruta:** {row_a.get('ruta','')} &nbsp;·&nbsp; "
                            f"**Placa:** {row_a.get('placa','')}"
                        )
                        st.markdown(
                            f"**Fecha:** {row_a.get('fecha','')} &nbsp;·&nbsp; {_lbl_extra}"
                        )
                    else:
                        for _di in _idxs_del:
                            if _di in _df_src_del.index:
                                _r = _df_src_del.loc[_di]
                                _r_extra = (
                                    f"{_r.get('seg_codigo','')} / {_r.get('sub_tipo_seguimiento','')}"
                                    if _from_seg else
                                    f"{_r.get('ruta','')} / {_r.get('placa','')}"
                                )
                                st.markdown(
                                    f"· **{_r.get('fecha','')}** — {_r_extra}"
                                )
                    dc1, dc2, _ = st.columns([1.5, 1, 3])
                    with dc1:
                        if st.button("🗑️ CONFIRMAR", type="primary",
                                     key="btn_confirm_del", width='stretch'):
                            if _from_seg:
                                delete_seg_rows(_idxs_del)
                            else:
                                delete_rows_from_csv(_idxs_del)
                            st.session_state.admin_accion   = None
                            st.session_state.admin_idx      = None
                            st.session_state.admin_idxs     = []
                            st.session_state.admin_from_seg = False
                            st.rerun()
                    with dc2:
                        if st.button("✖ CANCELAR", key="btn_cancel_del",
                                     width='stretch'):
                            st.session_state.admin_accion   = None
                            st.session_state.admin_idx      = None
                            st.session_state.admin_idxs     = []
                            st.session_state.admin_from_seg = False
                            st.rerun()



elif st.session_state.pagina_activa == "DASHBOARD":
    # ══════════════════════════════════════════════════════════════════════════════
    # DASHBOARD — QualiLact / Nestlé
    # ══════════════════════════════════════════════════════════════════════════════

    # ── Constantes ────────────────────────────────────────────────────────────
    _ST_MIN     = 12.60
    _IC_REF     = -0.530
    _TRANS_ALRT = 0.5
    _CLR_AZUL   = "#0056A3"
    _CLR_AZUL2  = "#3A7CC8"
    _CLR_ROJO   = "#EF4444"
    _CLR_VERDE  = "#10B981"
    _CLR_AMBAR  = "#F59E0B"
    _CLR_GRIS   = "#6B7280"
    _PLT_TMPL   = "plotly_white"

    # ── helpers ───────────────────────────────────────────────────────────────
    def _sh(txt, icon=""):
        st.markdown(
            f"<div style='font-size:.93rem;font-weight:700;color:{_CLR_AZUL};"
            f"border-left:4px solid {_CLR_AZUL};padding-left:10px;"
            f"margin:16px 0 6px 0;letter-spacing:.02em'>{icon} {txt}</div>",
            unsafe_allow_html=True,
        )

    def _base_layout(fig, h=300, mb=60, mr=80):
        fig.update_layout(
            template=_PLT_TMPL, height=h,
            margin=dict(l=20, r=mr, t=34, b=mb),
            paper_bgcolor="white", plot_bgcolor="white",
            font=dict(family="Arial", size=11),
            legend=dict(orientation="h", yanchor="bottom", y=1.02,
                        xanchor="right", x=1),
        )
        return fig

    # ── Carga de datos ────────────────────────────────────────────────────────
    _df_raw   = load_historial()
    _df_seg_raw = load_seguimientos()
    _df_rutas = _df_raw[_df_raw["tipo_seguimiento"] == "RUTAS"].copy()
    _df_trans = _df_raw[_df_raw["tipo_seguimiento"] == "TRANSUIZA"].copy()

    # ── Pre-process RUTAS / TRANSUIZA ─────────────────────────────────────────
    for _df in [_df_rutas, _df_trans]:
        _df["_fecha_dt"] = pd.to_datetime(_df["fecha"], format="%d/%m/%Y", errors="coerce")
        for _c in ["solidos_ruta", "crioscopia_ruta", "st_carrotanque",
                   "diferencia_solidos", "volumen_declarado"]:
            if _c in _df.columns:
                _df[_c] = pd.to_numeric(_df[_c], errors="coerce")

    _df_rutas = _df_rutas.dropna(subset=["_fecha_dt"])
    _df_trans = _df_trans.dropna(subset=["_fecha_dt"])

    # ── Parsear estaciones embebidas en cada registro RUTAS ───────────────────
    _estac_rows = []
    for _, _rrow in _df_rutas.iterrows():
        _js_raw = str(_rrow.get("estaciones_json", "") or "")
        if _js_raw.strip() in ("", "[]", "nan"):
            continue
        try:
            _est_list = json.loads(_js_raw)
        except Exception:
            continue
        for _e in _est_list:
            _cod = str(_e.get("codigo", "")).strip()
            if not _cod:
                continue
            _estac_rows.append({
                "_fecha_dt": _rrow["_fecha_dt"],
                "fecha":     _rrow["fecha"],
                "ruta":      _rrow.get("ruta", ""),
                "placa":     _rrow.get("placa", ""),
                "codigo":    _cod,
                "solidos":   pd.to_numeric(_e.get("solidos"), errors="coerce"),
                "crioscopia": pd.to_numeric(
                    str(_e.get("crioscopia", "")).replace(",", "."), errors="coerce"
                ),
            })

    _df_estac_ok = pd.DataFrame(_estac_rows) if _estac_rows else pd.DataFrame(
        columns=["_fecha_dt", "fecha", "ruta", "placa",
                 "codigo", "solidos", "crioscopia"]
    )

    _hay_rutas = not _df_rutas.empty
    _hay_trans = not _df_trans.empty
    _hay_estac = not _df_estac_ok.empty

    # ══════════════════════════════════════════════════════════════════════════
    # PESTAÑAS PRINCIPALES
    # ══════════════════════════════════════════════════════════════════════════
    _tab_r, _tab_e, _tab_t = st.tabs(
        ["  📊  GESTIÓN DE RUTAS  ",
         "  🏠  ESTACIONES  ",
         "  🏭  AUDITORÍA TRANSUIZA  "]
    )

    # ──────────────────────────────────────────────────────────────────────────
    # PESTAÑA 1 — GESTIÓN DE RUTAS
    # ──────────────────────────────────────────────────────────────────────────
    with _tab_r:
        if not _hay_rutas and not _hay_estac:
            st.info("No hay registros de RUTAS o ESTACIONES disponibles.")
        else:
            # ── Filtros Gestión de Rutas ───────────────────────────────────
            _sh("FILTROS", "🔎")
            _rf1, _rf2, _rf3, _rf4 = st.columns([1.5, 1.5, 2, 2])

            _r_f_min = _df_rutas["_fecha_dt"].min().date() if _hay_rutas else date.today()
            _r_f_max = _df_rutas["_fecha_dt"].max().date() if _hay_rutas else date.today()

            with _rf1:
                _r_desde = st.date_input("DESDE", value=_r_f_min,
                    min_value=_r_f_min, max_value=_r_f_max,
                    key="r_desde", format="DD/MM/YYYY")
            with _rf2:
                _r_hasta = st.date_input("HASTA", value=_r_f_max,
                    min_value=_r_f_min, max_value=_r_f_max,
                    key="r_hasta", format="DD/MM/YYYY")
            with _rf3:
                _r_ruta_opts = (
                    sorted(_df_rutas["ruta"].dropna().unique().tolist())
                    if _hay_rutas else []
                )
                _r_sel_rutas = st.multiselect("RUTA", _r_ruta_opts,
                    key="r_sel_rutas", placeholder="Todas las rutas")
            with _rf4:
                _r_placa_opts = (
                    sorted(_df_rutas["placa"].dropna().unique().tolist())
                    if _hay_rutas else []
                )
                _r_sel_placas = st.multiselect("PLACA", _r_placa_opts,
                    key="r_sel_placas", placeholder="Todas las placas")

            # Aplicar filtros
            _dfr = _df_rutas[
                (_df_rutas["_fecha_dt"].dt.date >= _r_desde) &
                (_df_rutas["_fecha_dt"].dt.date <= _r_hasta)
            ].copy()
            if _r_sel_rutas:
                _dfr = _dfr[_dfr["ruta"].isin(_r_sel_rutas)]
            if _r_sel_placas:
                _dfr = _dfr[_dfr["placa"].isin(_r_sel_placas)]

            st.markdown(
                "<hr style='border-color:#E5E7EB;margin:4px 0 10px;'>",
                unsafe_allow_html=True,
            )

            # ── KPIs ──────────────────────────────────────────────────────
            if not _dfr.empty:
                _n_r      = len(_dfr)
                _avg_st   = float(_dfr["solidos_ruta"].mean())
                _avg_ic   = float(_dfr["crioscopia_ruta"].mean())
                _pct_st   = (_dfr["solidos_ruta"] >= _ST_MIN).mean() * 100
                _pct_ic   = (_dfr["crioscopia_ruta"] <= _IC_REF).mean() * 100
                _k1, _k2, _k3, _k4 = st.columns(4)
                _k1.metric("REGISTROS", str(_n_r))
                _k2.metric("ST PROM.", f"{_avg_st:.2f}%",
                    f"{'✔ OK' if _avg_st >= _ST_MIN else '⚠ BAJO'} — mín {_ST_MIN}%",
                    delta_color="off")
                _k3.metric("IC PROM.", f"{_avg_ic:.3f} °H",
                    f"{'✔ OK' if _avg_ic <= _IC_REF else '⚠ AGUA'} — ref {_IC_REF}",
                    delta_color="off")
                _k4.metric("CONF. ST/IC",
                    f"{_pct_st:.0f}% / {_pct_ic:.0f}%",
                    "ST conforme / IC conforme", delta_color="off")

            # ══════════════════════════════════════════════════════════════
            # SUB-SECCIÓN — Comportamiento por Rutas
            # ══════════════════════════════════════════════════════════════
            _sh("COMPORTAMIENTO POR RUTAS", "🚛")

            if _dfr.empty:
                st.info("Sin registros RUTAS para el período/filtros seleccionados.")
            else:
                # Tendencia temporal ST + IC
                _dfr_d = (
                    _dfr.groupby(_dfr["_fecha_dt"].dt.date)
                    .agg(st_avg=("solidos_ruta", "mean"),
                         ic_avg=("crioscopia_ruta", "mean"),
                         n=("solidos_ruta", "count"))
                    .reset_index().rename(columns={"_fecha_dt": "fecha"})
                )
                _dfr_d["fecha"] = pd.to_datetime(_dfr_d["fecha"])

                _tc1, _tc2 = st.columns(2)
                with _tc1:
                    _fig_r_st = go.Figure()
                    _fig_r_st.add_trace(go.Scatter(
                        x=_dfr_d["fecha"], y=_dfr_d["st_avg"],
                        mode="lines+markers", name="ST prom.",
                        line=dict(color=_CLR_AZUL, width=2),
                        marker=dict(size=7,
                            color=[_CLR_ROJO if v < _ST_MIN else _CLR_AZUL
                                   for v in _dfr_d["st_avg"]]),
                        hovertemplate="%{x|%d/%m/%Y}<br>ST: %{y:.2f}%<extra></extra>",
                    ))
                    _fig_r_st.add_hline(
                        y=_ST_MIN, line_dash="dash", line_color=_CLR_ROJO,
                        annotation_text=f"mín {_ST_MIN}%",
                        annotation_position="right",
                    )
                    _base_layout(_fig_r_st, h=270, mb=40, mr=80)
                    _fig_r_st.update_layout(
                        title=dict(text="TENDENCIA ST — RUTAS (%)",
                                   font=dict(size=12, color=_CLR_AZUL)),
                        yaxis_title="ST (%)", xaxis_title=None, showlegend=False,
                    )
                    st.plotly_chart(_fig_r_st, width='stretch')

                with _tc2:
                    _fig_r_ic = go.Figure()
                    _fig_r_ic.add_trace(go.Scatter(
                        x=_dfr_d["fecha"], y=_dfr_d["ic_avg"],
                        mode="lines+markers", name="IC prom.",
                        line=dict(color=_CLR_VERDE, width=2),
                        marker=dict(size=7,
                            color=[_CLR_ROJO if v > _IC_REF else _CLR_VERDE
                                   for v in _dfr_d["ic_avg"]]),
                        hovertemplate="%{x|%d/%m/%Y}<br>IC: %{y:.3f} °H<extra></extra>",
                    ))
                    _fig_r_ic.add_hline(
                        y=_IC_REF, line_dash="dash", line_color=_CLR_ROJO,
                        annotation_text=f"ref {_IC_REF} °H",
                        annotation_position="right",
                    )
                    _base_layout(_fig_r_ic, h=270, mb=40, mr=80)
                    _fig_r_ic.update_layout(
                        title=dict(text="TENDENCIA IC — RUTAS (°H)",
                                   font=dict(size=12, color=_CLR_AZUL)),
                        yaxis_title="IC (°H)", xaxis_title=None, showlegend=False,
                    )
                    st.plotly_chart(_fig_r_ic, width='stretch')

                # Barras comparativas por Ruta
                _dfr_ruta = (
                    _dfr.groupby("ruta")
                    .agg(st_avg=("solidos_ruta", "mean"),
                         ic_avg=("crioscopia_ruta", "mean"),
                         n=("solidos_ruta", "count"))
                    .reset_index()
                    .sort_values("st_avg", ascending=False)
                )

                _fig_ruta_bar = go.Figure()
                _fig_ruta_bar.add_trace(go.Bar(
                    x=_dfr_ruta["ruta"],
                    y=_dfr_ruta["st_avg"],
                    name="ST prom. (%)",
                    marker_color=[
                        _CLR_ROJO if v < _ST_MIN else _CLR_AZUL
                        for v in _dfr_ruta["st_avg"]
                    ],
                    text=_dfr_ruta["st_avg"].round(2),
                    textposition="outside",
                    yaxis="y1",
                    hovertemplate="<b>%{x}</b><br>ST: %{y:.2f}%<extra></extra>",
                ))
                _fig_ruta_bar.add_trace(go.Scatter(
                    x=_dfr_ruta["ruta"],
                    y=_dfr_ruta["ic_avg"],
                    mode="lines+markers",
                    name="IC prom. (°H)",
                    line=dict(color=_CLR_AMBAR, width=2),
                    marker=dict(size=8, color=_CLR_AMBAR),
                    yaxis="y2",
                    hovertemplate="<b>%{x}</b><br>IC: %{y:.3f} °H<extra></extra>",
                ))
                _fig_ruta_bar.add_hline(
                    y=_ST_MIN, line_dash="dash", line_color=_CLR_ROJO,
                    annotation_text=f"ST mín {_ST_MIN}%",
                    annotation_position="right",
                )
                _base_layout(_fig_ruta_bar, h=320, mb=80, mr=100)
                _fig_ruta_bar.update_layout(
                    title=dict(text="ST E IC PROMEDIO POR RUTA",
                               font=dict(size=12, color=_CLR_AZUL)),
                    barmode="group",
                    xaxis=dict(tickangle=-30, title=None),
                    yaxis=dict(title="ST (%)", range=[11.5, 14.2]),
                    yaxis2=dict(title="IC (°H)", overlaying="y",
                                side="right", range=[-0.58, -0.49]),
                    legend=dict(orientation="h", yanchor="bottom",
                                y=1.02, xanchor="right", x=1),
                )
                st.plotly_chart(_fig_ruta_bar, width='stretch')

    # ──────────────────────────────────────────────────────────────────────────
    # PESTAÑA 2 — ESTACIONES
    # ──────────────────────────────────────────────────────────────────────────
    with _tab_e:
        _sh("FILTROS", "🔎")
        _e_all_fechas = (
            _df_estac_ok["_fecha_dt"] if _hay_estac and not _df_estac_ok.empty
            else pd.Series(dtype="datetime64[ns]")
        )
        _e_f_min = _e_all_fechas.min().date() if not _e_all_fechas.empty else date.today()
        _e_f_max = _e_all_fechas.max().date() if not _e_all_fechas.empty else date.today()

        _ef1, _ef2, _ef3 = st.columns([1.5, 1.5, 3])
        with _ef1:
            _e_desde = st.date_input("DESDE", value=_e_f_min,
                min_value=_e_f_min, max_value=_e_f_max,
                key="e_desde", format="DD/MM/YYYY")
        with _ef2:
            _e_hasta = st.date_input("HASTA", value=_e_f_max,
                min_value=_e_f_min, max_value=_e_f_max,
                key="e_hasta", format="DD/MM/YYYY")
        with _ef3:
            _e_ruta_opts = (
                sorted(_df_estac_ok["ruta"].dropna().unique().tolist())
                if _hay_estac else []
            )
            _e_sel_rutas = st.multiselect("RUTA", _e_ruta_opts,
                key="e_sel_rutas", placeholder="Todas las rutas")

        if not _hay_estac:
            st.info("No hay datos de estaciones disponibles.")
        else:
            _dfe_base = _df_estac_ok[
                (_df_estac_ok["_fecha_dt"].dt.date >= _e_desde) &
                (_df_estac_ok["_fecha_dt"].dt.date <= _e_hasta)
            ].copy()
            if _e_sel_rutas:
                _dfe_base = _dfe_base[_dfe_base["ruta"].isin(_e_sel_rutas)]

            # Filtro adicional por código de estación
            _est_opts2 = sorted(
                _dfe_base["codigo"].dropna().astype(str).unique().tolist()
            ) if not _dfe_base.empty else []
            _sel_est2 = st.multiselect(
                "CÓDIGO DE ESTACIÓN", _est_opts2,
                key="e_sel_est", placeholder="Todas las estaciones"
            )
            if _sel_est2:
                _dfe_base = _dfe_base[
                    _dfe_base["codigo"].astype(str).isin(_sel_est2)
                ]

            st.markdown(
                "<hr style='border-color:#E5E7EB;margin:4px 0 10px;'>",
                unsafe_allow_html=True,
            )

            if _dfe_base.empty:
                st.warning("Sin datos para los filtros seleccionados.")
            else:
                # KPIs rápidos
                _ne  = len(_dfe_base)
                _ast = _dfe_base["solidos"].mean()
                _aic = _dfe_base["crioscopia"].mean()
                _pst = (_dfe_base["solidos"] >= _ST_MIN).mean() * 100
                _ek1, _ek2, _ek3, _ek4 = st.columns(4)
                _ek1.metric("LECTURAS", str(_ne))
                _ek2.metric("ST PROM.", f"{_ast:.2f}%",
                    f"{'✔ OK' if _ast >= _ST_MIN else '⚠ BAJO'} — mín {_ST_MIN}%",
                    delta_color="off")
                _ek3.metric("IC PROM.", f"{_aic:.3f} °H",
                    f"{'✔ OK' if _aic <= _IC_REF else '⚠ AGUA'} — ref {_IC_REF}",
                    delta_color="off")
                _ek4.metric("CONF. ST", f"{_pst:.0f}%",
                    "Lecturas ≥ 12.60%", delta_color="off")

                _dfe_base["Estación"] = _dfe_base["codigo"].astype(str)

                # ── Gráfica ST por Estación ──────────────────────────────
                _sh("SÓLIDOS TOTALES POR ESTACIÓN", "📈")
                _fig_est_st2 = go.Figure()
                for _cod in sorted(_dfe_base["Estación"].unique()):
                    _sub = (_dfe_base[_dfe_base["Estación"] == _cod]
                            .sort_values("_fecha_dt"))
                    _fig_est_st2.add_trace(go.Scatter(
                        x=_sub["_fecha_dt"], y=_sub["solidos"],
                        mode="lines+markers", name=f"Est. {_cod}",
                        hovertemplate=(
                            f"<b>Estación {_cod}</b><br>"
                            "Fecha: %{x|%d/%m/%Y}<br>"
                            "ST: %{y:.2f}%<extra></extra>"
                        ),
                    ))
                _fig_est_st2.add_hline(
                    y=_ST_MIN, line_dash="dash", line_color=_CLR_ROJO,
                    annotation_text=f"mín {_ST_MIN}%",
                    annotation_position="right",
                )
                _base_layout(_fig_est_st2, h=310, mb=40, mr=80)
                _fig_est_st2.update_layout(
                    yaxis_title="ST (%)", xaxis_title=None,
                )
                st.plotly_chart(_fig_est_st2, width='stretch')

                # ── Gráfica IC por Estación ──────────────────────────────
                _sh("CRIOSCOPIA POR ESTACIÓN", "🌡️")
                _fig_est_ic2 = go.Figure()
                for _cod in sorted(_dfe_base["Estación"].unique()):
                    _sub = (_dfe_base[_dfe_base["Estación"] == _cod]
                            .sort_values("_fecha_dt"))
                    _fig_est_ic2.add_trace(go.Scatter(
                        x=_sub["_fecha_dt"], y=_sub["crioscopia"],
                        mode="lines+markers", name=f"Est. {_cod}",
                        hovertemplate=(
                            f"<b>Estación {_cod}</b><br>"
                            "Fecha: %{x|%d/%m/%Y}<br>"
                            "IC: %{y:.3f} °H<extra></extra>"
                        ),
                    ))
                _fig_est_ic2.add_hline(
                    y=_IC_REF, line_dash="dash", line_color=_CLR_ROJO,
                    annotation_text=f"ref {_IC_REF} °H",
                    annotation_position="right",
                )
                _base_layout(_fig_est_ic2, h=310, mb=40, mr=80)
                _fig_est_ic2.update_layout(
                    yaxis_title="IC (°H)", xaxis_title=None,
                )
                st.plotly_chart(_fig_est_ic2, width='stretch')

                # ── Tabla resumen por estación ───────────────────────────
                _sh("RESUMEN POR ESTACIÓN", "📋")
                _df_est_agg = (
                    _dfe_base.groupby("codigo")
                    .agg(
                        lecturas   =("solidos",    "count"),
                        st_prom    =("solidos",    "mean"),
                        st_min     =("solidos",    "min"),
                        st_max     =("solidos",    "max"),
                        ic_prom    =("crioscopia", "mean"),
                        ic_min     =("crioscopia", "min"),
                        ic_max     =("crioscopia", "max"),
                    )
                    .reset_index()
                    .rename(columns={"codigo": "CÓDIGO"})
                )
                _df_est_agg["ST OK"] = (
                    _df_est_agg["st_prom"] >= _ST_MIN
                ).map({True: "✔", False: "✖"})
                _df_est_agg["IC OK"] = (
                    _df_est_agg["ic_prom"] <= _IC_REF
                ).map({True: "✔", False: "✖"})

                def _est_row_style(row):
                    if row["ST OK"] == "✖" or row["IC OK"] == "✖":
                        return ["background-color:#FEF2F2;color:#B91C1C"] * len(row)
                    return ["background-color:#F0FDF4;color:#15803D"] * len(row)

                st.dataframe(
                    _df_est_agg.style.apply(_est_row_style, axis=1)
                    .format({
                        "st_prom": "{:.2f}", "st_min": "{:.2f}", "st_max": "{:.2f}",
                        "ic_prom": "{:.3f}", "ic_min": "{:.3f}", "ic_max": "{:.3f}",
                    }),
                    width='stretch',
                    hide_index=True,
                    column_config={
                        "CÓDIGO":   "CÓDIGO",
                        "lecturas": "LECTURAS",
                        "st_prom":  "ST PROM. (%)",
                        "st_min":   "ST MÍN. (%)",
                        "st_max":   "ST MÁX. (%)",
                        "ic_prom":  "IC PROM. (°H)",
                        "ic_min":   "IC MÍN. (°H)",
                        "ic_max":   "IC MÁX. (°H)",
                        "ST OK":    "ST ≥ 12.60",
                        "IC OK":    "IC ≤ −0.530",
                    },
                )

    # ──────────────────────────────────────────────────────────────────────────
    # PESTAÑA 3 — AUDITORÍA TRANSUIZA
    # ──────────────────────────────────────────────────────────────────────────
    with _tab_t:
        if not _hay_trans:
            st.info("No hay registros TRANSUIZA disponibles.")
        else:
            # ── Filtro Fecha ───────────────────────────────────────────────
            _sh("FILTRO DE FECHA", "📅")
            _t_f_min = _df_trans["_fecha_dt"].min().date()
            _t_f_max = _df_trans["_fecha_dt"].max().date()

            _tf1, _tf2 = st.columns([2, 2])
            with _tf1:
                _t_desde = st.date_input("DESDE", value=_t_f_min,
                    min_value=_t_f_min, max_value=_t_f_max,
                    key="t_desde", format="DD/MM/YYYY")
            with _tf2:
                _t_hasta = st.date_input("HASTA", value=_t_f_max,
                    min_value=_t_f_min, max_value=_t_f_max,
                    key="t_hasta", format="DD/MM/YYYY")

            _dft = _df_trans[
                (_df_trans["_fecha_dt"].dt.date >= _t_desde) &
                (_df_trans["_fecha_dt"].dt.date <= _t_hasta)
            ].copy()

            st.markdown(
                "<hr style='border-color:#E5E7EB;margin:4px 0 10px;'>",
                unsafe_allow_html=True,
            )

            if _dft.empty:
                st.warning("No hay datos TRANSUIZA para el período seleccionado.")
            else:
                _dft["ST MUESTRA (%)"]     = pd.to_numeric(_dft["solidos_ruta"],   errors="coerce")
                _dft["ST CARROTANQUE (%)"] = pd.to_numeric(_dft["st_carrotanque"], errors="coerce")
                _dft["DIFERENCIA (%)"]     = (
                    _dft["ST MUESTRA (%)"] - _dft["ST CARROTANQUE (%)"]
                ).abs().round(3)
                _dft["ALERTA"]  = _dft["DIFERENCIA (%)"] > _TRANS_ALRT
                _dft["_eje_x"]  = (
                    _dft["fecha"].astype(str) + " — " + _dft["placa"].fillna("").astype(str)
                )

                # KPIs Transuiza
                _kt1, _kt2, _kt3 = st.columns(3)
                _kt1.metric("REGISTROS TRANSUIZA", str(len(_dft)))
                _kt2.metric("ST MUESTRA PROM.",
                    f"{_dft['ST MUESTRA (%)'].mean():.2f}%")
                _kt3.metric("ALERTAS DIFERENCIA",
                    str(int(_dft["ALERTA"].sum())),
                    f"Diferencia > {_TRANS_ALRT}%",
                    delta_color="inverse" if _dft["ALERTA"].any() else "off")

                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

                # ── Gráfica comparativa ST Muestra vs ST Carrotanque ──────
                _sh("ST MUESTRA vs. ST CARROTANQUE", "📊")

                _fig_t = go.Figure()
                _fig_t.add_trace(go.Bar(
                    name="ST MUESTRA (%)",
                    x=_dft["_eje_x"],
                    y=_dft["ST MUESTRA (%)"],
                    marker_color=_CLR_AZUL,
                    text=_dft["ST MUESTRA (%)"].round(2),
                    textposition="outside",
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "ST Muestra: %{y:.2f}%<extra></extra>"
                    ),
                ))
                _fig_t.add_trace(go.Bar(
                    name="ST CARROTANQUE (%)",
                    x=_dft["_eje_x"],
                    y=_dft["ST CARROTANQUE (%)"],
                    marker_color=_CLR_AMBAR,
                    text=_dft["ST CARROTANQUE (%)"].round(2),
                    textposition="outside",
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "ST Carrotanque: %{y:.2f}%<extra></extra>"
                    ),
                ))
                # Línea de diferencia (scatter secundario)
                _fig_t.add_trace(go.Scatter(
                    x=_dft["_eje_x"],
                    y=_dft["DIFERENCIA (%)"],
                    mode="lines+markers",
                    name="DIFERENCIA (%)",
                    yaxis="y2",
                    line=dict(color=_CLR_ROJO, width=2, dash="dot"),
                    marker=dict(size=8,
                        color=[_CLR_ROJO if a else _CLR_GRIS
                               for a in _dft["ALERTA"]],
                    ),
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "Diferencia: %{y:.3f}%<extra></extra>"
                    ),
                ))
                _fig_t.add_hline(
                    y=_ST_MIN, line_dash="dash", line_color=_CLR_ROJO,
                    annotation_text=f"ST mín {_ST_MIN}%",
                    annotation_position="right",
                )
                _base_layout(_fig_t, h=340, mb=80, mr=100)
                _fig_t.update_layout(
                    barmode="group",
                    xaxis=dict(tickangle=-35, title=None),
                    yaxis=dict(title="Sólidos Totales (%)", range=[11.0, 15.0]),
                    yaxis2=dict(title="Diferencia (%)", overlaying="y",
                                side="right", range=[0, 3],
                                showgrid=False),
                    legend=dict(orientation="h", yanchor="bottom",
                                y=1.02, xanchor="right", x=1),
                    title=dict(
                        text="COMPARACIÓN ST MUESTRA vs. CARROTANQUE POR REGISTRO",
                        font=dict(size=12, color=_CLR_AZUL),
                    ),
                )
                st.plotly_chart(_fig_t, width='stretch')

                # Alertas destacadas
                _n_alt = int(_dft["ALERTA"].sum())
                if _n_alt > 0:
                    st.markdown(
                        f"<div style='background:#FEF2F2;border:1px solid #FECACA;"
                        f"border-radius:6px;padding:8px 14px;margin:6px 0;"
                        f"color:#B91C1C;font-weight:600;font-size:.88rem;'>"
                        f"⚠️ {_n_alt} registro(s) con diferencia ST &gt; {_TRANS_ALRT}%"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.success("✅ Ningún registro supera la diferencia límite de "
                               f"{_TRANS_ALRT}%.")

                # Tabla de detalle
                _sh("DETALLE DE REGISTROS", "📋")
                _tbl = (
                    _dft[["fecha", "placa", "ST MUESTRA (%)",
                           "ST CARROTANQUE (%)", "DIFERENCIA (%)", "ALERTA"]]
                    .rename(columns={"fecha": "FECHA", "placa": "PLACA"})
                    .reset_index(drop=True)
                )

                def _style_alerta(row):
                    if _tbl.loc[row.name, "ALERTA"]:
                        return ["background-color:#FEE2E2;color:#B91C1C"] * len(row)
                    return [""] * len(row)

                st.dataframe(
                    _tbl.drop(columns=["ALERTA"])
                    .style.apply(_style_alerta, axis=1)
                    .format({"ST MUESTRA (%)": "{:.2f}",
                             "ST CARROTANQUE (%)": "{:.2f}",
                             "DIFERENCIA (%)": "{:.3f}"}),
                    width='stretch',
                    hide_index=True,
                    height=min(70 + len(_tbl) * 36, 380),
                )

                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

                # ── Reporte de Alertas ─────────────────────────────────────
                _sh("REPORTE DE ALERTAS", "🚨")

                # Construir reporte: diferencia > umbral O IC fuera de rango
                _rep_trans = _dft[
                    _dft["DIFERENCIA (%)"] > _TRANS_ALRT
                ][["fecha", "placa", "ST MUESTRA (%)",
                   "ST CARROTANQUE (%)", "DIFERENCIA (%)"]].copy()
                _rep_trans["TIPO ALERTA"] = (
                    "DIFERENCIA ST > " + _TRANS_ALRT.__str__() + "%"
                )
                _rep_trans.rename(
                    columns={"fecha": "FECHA", "placa": "PLACA"}, inplace=True
                )

                # IC fuera de rango (si hay crioscopia en RUTAS del mismo período)
                _rep_ic = pd.DataFrame()
                if _hay_rutas and not _df_rutas.empty:
                    _dfr_ic = _df_rutas[
                        (_df_rutas["_fecha_dt"].dt.date >= _t_desde) &
                        (_df_rutas["_fecha_dt"].dt.date <= _t_hasta) &
                        (_df_rutas["crioscopia_ruta"] > _IC_REF)
                    ][["fecha", "ruta", "placa",
                       "solidos_ruta", "crioscopia_ruta"]].copy()
                    if not _dfr_ic.empty:
                        _dfr_ic.rename(columns={
                            "fecha": "FECHA",
                            "ruta": "RUTA",
                            "placa": "PLACA",
                            "solidos_ruta": "ST MUESTRA (%)",
                            "crioscopia_ruta": "IC (°H)",
                        }, inplace=True)
                        _dfr_ic["TIPO ALERTA"] = (
                            "IC FUERA DE RANGO (> " + str(_IC_REF) + " °H)"
                        )
                        _rep_ic = _dfr_ic

                _dfs_concat = [df for df in [_rep_trans, _rep_ic] if not df.empty]
                _reporte_final = (
                    pd.concat(_dfs_concat, ignore_index=True)
                    if _dfs_concat
                    else pd.DataFrame()
                )

                _n_rep = len(_reporte_final)
                if _n_rep == 0:
                    st.success(
                        "✅ No se encontraron alertas en el período seleccionado."
                    )
                else:
                    st.error(f"⚠️ {_n_rep} alerta(s) detectada(s) en el período.")
                    st.dataframe(
                        _reporte_final,
                        width='stretch',
                        hide_index=True,
                        height=min(70 + _n_rep * 36, 320),
                    )

                _csv_alertas = _reporte_final.to_csv(
                    index=False, encoding="utf-8-sig"
                ).encode()
                st.download_button(
                    label="🚨  GENERAR REPORTE DE ALERTAS (CSV)",
                    data=_csv_alertas,
                    file_name=(
                        f"alertas_transuiza_{now_col().strftime('%Y%m%d_%H%M')}.csv"
                    ),
                    mime="text/csv",
                    type="primary",
                )


    save_draft_state()
